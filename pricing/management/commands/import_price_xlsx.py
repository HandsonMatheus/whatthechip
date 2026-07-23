"""
import_price_xlsx — F4 do PRECIFICACAO.md: a planilha do comprador vira banco.

    python manage.py import_price_xlsx wuquanprices.xlsx --buyer wuquan            # DRY-RUN
    python manage.py import_price_xlsx wuquanprices.xlsx --buyer wuquan --commit   # grava

É o BOOTSTRAP ÚNICO (PRECIFICACAO §1.10): depois dele a planilha se aposenta —
edição de preço passa a ser admin (dono) + dashboard /partner/ (comprador).
Idempotente: re-rodar faz upsert pela chave (lista, kind, gen, tier).

Regras encodadas (§6 — todas verificadas na planilha REAL):
  1. Uma aba por marca + "Other Brands" (vira a lista GENÉRICA). "Instructions"/
     "Sheet1" são ignoradas. Nome da aba → chips.Brand via normalização
     ("Toshiba Kioxia"/"Toshiba/Kioxia" → "Toshiba-Kioxia"). A coluna A é
     DECORATIVA (decisão 2026-07-07): a marca da linha é sempre a da ABA — na
     "Other Brands" tudo vira linha da genérica (nada de lista-fantasma).
     Depois do import, rode `seed_price_grid` para unificar o grid.
  2. Preço-fonte = coluna F (RMB): número → min=max · faixa "90-110" → min/max
     (achatada no ponto médio) · "NO" → no_buy · vazio → unquoted. **F10 (RMB
     canônico, 2026-07-16): o ¥ é gravado DIRETO** — nada de × câmbio; o USD é
     derivado na leitura pelo engine (Buyer.fx_usd_rate). A célula B2 (câmbio)
     segue OBRIGATÓRIA só como validação de estrutura da planilha; a coluna E
     (USD) é IGNORADA. (Pré-F10 este comando gravava RMB × B2 — foi assim que
     os USD "nasceram a 0.15"; o migrate_prices_to_rmb desfaz isso ÷0.15.)
  3. eMCP/uMCP: "64+4" → tier = 64 GB (NAND) + gen = subtipo LPDDR; os combos de
     RAM da MESMA faixa COLAPSAM numa linha. Combos com preços DIFERENTES na
     mesma (gen, faixa) = CONFLITO → o import ABORTA com relatório (contradiz a
     regra "cota por faixa"; quem decide é o dono, nunca o código).
  4. DDR: "2Gb" → tier em Gb (die). eMMC/UFS/LPDDR: GB ("1TB" = 1024). A
     unidade da célula TEM que casar o tipo (Gb≠GB — case-sensitive); não casou
     → linha pulada com motivo.
  5. Linha malformada NÃO derruba o import (vira "pulada" no relatório);
     CONFLITO eMCP e aba sem câmbio DERRUBAM (segurança > conveniência).
  6. `inherits_from` (SK espelha Samsung, genérica → Nanya…) NÃO é setado aqui —
     é regra do comprador, configurada no admin (PRECIFICACAO §4). A aba SK vem
     com linhas PRÓPRIAS (dado real ≠ espelho — decisão registrada no §11).

Escopo: roda FORA de request → `scope_command_to_company()` (auto-resolve com
uma empresa ativa; 2+ exige --company). O Buyer é criado sob demanda na empresa
escopada. Dry-run por padrão (regra de ouro #1: o dono roda o --commit).
"""

import re

from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from chips.models import Brand
from pricing.models import (Buyer, KIND_UNIT, Price, PriceList,
                            STATUS_NO_BUY, STATUS_QUOTED, STATUS_UNQUOTED,
                            fold_gen)
from tenancy.scope import scope_command_to_company

# Abas que não são de preço.
_SKIP_SHEETS = {'instructions', 'sheet1'}
# Nome (de aba OU da coluna A) → nome EXATO do chips.Brand no banco.
_BRAND_ALIASES = {
    'toshiba kioxia':  'Toshiba-Kioxia',
    'toshiba/kioxia':  'Toshiba-Kioxia',
    'toshiba-kioxia':  'Toshiba-Kioxia',
}
_GENERIC_SHEET = 'other brands'

# (GDDR fora do mercado desde 2026-07-23: linha GDDR da planilha é PULADA
#  com motivo — nunca vira linha de grid.)
_TYPE_TO_KIND = {'emmc': 'emmc', 'ufs': 'ufs', 'emcp': 'emcp', 'umcp': 'umcp',
                 'lpddr': 'lpddr', 'ddr': 'ddr'}

_RANGE_RE = re.compile(r'^\s*(\d+(?:\.\d+)?)\s*[-–]\s*(\d+(?:\.\d+)?)\s*$')
_NUM_RE   = re.compile(r'^\s*(\d+(?:\.\d+)?)\s*$')
_EMCP_RE  = re.compile(r'^\s*(\d+(?:\.\d+)?)\s*\+')          # "64+4" → 64
# ⚠ case-sensitive de propósito: GB (byte, pacote) ≠ Gb (bit, die).
_CAP_RE   = re.compile(r'^\s*(\d+(?:\.\d+)?)\s*(TB|GB|Gb|MB)\s*$')

_CENT = Decimal('0.01')


def _clean(v) -> str:
    if v is None:
        return ''
    return str(v).strip()


class Command(BaseCommand):
    help = ('Importa a planilha de preços do comprador (bootstrap único; depois '
            'a planilha se aposenta). DRY-RUN por padrão; --commit grava.')

    def add_arguments(self, parser):
        parser.add_argument('xlsx', help='Caminho da planilha (ex.: wuquanprices.xlsx).')
        parser.add_argument('--buyer', required=True,
                            help='Slug do comprador (ex.: wuquan). Criado se não existir.')
        parser.add_argument('--company', default=None,
                            help='Slug da empresa (obrigatório se houver 2+ ativas).')
        parser.add_argument('--commit', action='store_true',
                            help='Grava de verdade (sem isto: dry-run, nada muda).')

    # ── Parsing ──────────────────────────────────────────────────────────────

    def _parse_rmb(self, raw):
        """Coluna F → (status, rmb_min, rmb_max) ou None (malformada).

        PREÇO FIXO (decisão 2026-07-07): faixa da planilha ("90-110") é
        ACHATADA no ponto médio — o sistema não guarda variação."""
        s = _clean(raw)
        if not s:
            return STATUS_UNQUOTED, None, None
        if s.upper() == 'NO':
            return STATUS_NO_BUY, None, None
        m = _NUM_RE.match(s)
        if m:
            v = Decimal(m.group(1))
            return STATUS_QUOTED, v, v
        m = _RANGE_RE.match(s)
        if m:
            lo, hi = Decimal(m.group(1)), Decimal(m.group(2))
            mid = (lo + hi) / 2
            return STATUS_QUOTED, mid, mid
        return None

    def _parse_tier(self, kind: str, raw):
        """Coluna D → (tier_value: Decimal, motivo_de_erro: str|None)."""
        s = _clean(raw)
        if kind in ('emcp', 'umcp'):
            m = _EMCP_RE.match(s)
            if not m:
                return None, f'capacidade eMCP {s!r} fora do formato NAND+RAM'
            return Decimal(m.group(1)), None
        m = _CAP_RE.match(s)
        if not m:
            return None, f'capacidade {s!r} ilegível'
        value, unit = Decimal(m.group(1)), m.group(2)
        if unit == 'TB':
            value, unit = value * 1024, 'GB'
        elif unit == 'MB':
            value, unit = value / 1024, 'GB'
        if unit != KIND_UNIT[kind]:
            return None, (f'unidade {unit} não casa {kind} '
                          f'(esperado {KIND_UNIT[kind]}; Gb≠GB)')
        return value, None

    def _parse_date(self, raw):
        if isinstance(raw, datetime):
            return raw.date()
        if isinstance(raw, date):
            return raw
        s = _clean(raw)
        if not s:
            return None
        for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d'):
            try:
                return datetime.strptime(s[:19], fmt).date()
            except ValueError:
                continue
        return None

    def _brand_for(self, name: str):
        """Nome cru (aba ou coluna A) → chips.Brand ou None (não cadastrada)."""
        canonical = _BRAND_ALIASES.get(name.strip().lower(), name.strip())
        return Brand.objects.filter(name__iexact=canonical).first()

    # ── Pipeline ─────────────────────────────────────────────────────────────

    def handle(self, *args, **opts):
        try:
            import openpyxl
        except ImportError:
            raise CommandError('openpyxl não instalado (pip install openpyxl).')

        company = scope_command_to_company(opts['company'], self.stdout)
        wb = openpyxl.load_workbook(opts['xlsx'], data_only=True)

        # plan[(brand_name|None, kind, gen, tier)] = dict(linha) — colapso eMCP
        plan: dict = {}
        conflicts, skipped = [], []
        rates: dict = {}

        for sheet in wb.sheetnames:
            low = sheet.strip().lower()
            if low in _SKIP_SHEETS:
                continue
            ws = wb[sheet]
            generic_tab = (low == _GENERIC_SHEET)
            tab_brand_name = None if generic_tab else sheet.strip()

            rate_raw = ws['B2'].value
            try:
                rate = Decimal(str(rate_raw))
                if rate <= 0:
                    raise ValueError
            except Exception:
                raise CommandError(
                    f'Aba {sheet!r}: câmbio ausente/ilegível na célula B2 '
                    f'({rate_raw!r}) — sem ele não há conversão RMB→USD.')
            rates[sheet] = rate

            for i, row in enumerate(ws.iter_rows(min_row=4, values_only=True),
                                    start=4):
                row = list(row) + [None] * (9 - len(row))
                col_type, col_gen, col_cap = (
                    _clean(row[1]), _clean(row[2]), _clean(row[3]))
                if not col_type and not col_cap:
                    continue                        # linha em branco/título

                where = f'{sheet}!L{i}'
                kind = _TYPE_TO_KIND.get(col_type.lower())
                if kind is None:
                    skipped.append((where, f'tipo {col_type!r} desconhecido'))
                    continue

                # Fold (dono 2026-07-21): planilha grafada na variante
                # (DDR3L/LPDDR4X em ddr/lpddr) entra como a geração-BASE —
                # o grid é canônico (fold_gen; eMCP/uMCP mantêm a RAM).
                gen = '' if col_gen in ('—', '-', '') else col_gen
                gen = fold_gen(kind, gen)
                tier, err = self._parse_tier(kind, col_cap)
                if err:
                    skipped.append((where, err))
                    continue

                parsed = self._parse_rmb(row[5])
                if parsed is None:
                    skipped.append((where, f'RMB {row[5]!r} ilegível'))
                    continue
                status, rmb_min, rmb_max = parsed
                val_min = val_max = None
                if status == STATUS_QUOTED:
                    # F10 (RMB canônico): grava o ¥ DIRETO — sem × câmbio.
                    val_min = rmb_min.quantize(_CENT, ROUND_HALF_UP)
                    val_max = rmb_max.quantize(_CENT, ROUND_HALF_UP)

                # Marca da LINHA = a da ABA (decisão 2026-07-07): a coluna A é
                # decorativa — na "Other Brands" TUDO é linha da GENÉRICA (nada
                # de lista-fantasma por marca citada na coluna A, ex.: Rayson).
                brand_name = tab_brand_name

                key = (brand_name, kind, gen, tier)
                entry = dict(
                    status=status, val_min=val_min, val_max=val_max,
                    quote_date=self._parse_date(row[6]),
                    source=_clean(row[7])[:200], notes=_clean(row[8]),
                    where=where, collapsed=0,
                )
                if key not in plan:
                    plan[key] = entry
                    continue
                # ── Colapso (combos eMCP da mesma faixa; ou linha repetida) ──
                # Informação vence ausência: cotado > não-compra > vazio (na
                # planilha REAL a aba SK tem "64+4 cotado" + "64+6 vazio" na
                # MESMA faixa — não é contradição, é célula por preencher).
                # CONFLITO de verdade: cotado×cotado divergente, ou cotado×NO.
                kept = plan[key]
                rank = {STATUS_QUOTED: 2, STATUS_NO_BUY: 1, STATUS_UNQUOTED: 0}

                def _merge_date(a, b):
                    return max(d for d in (a, b) if d) if (a or b) else None

                if kept['status'] == status:
                    same_value = (kept['val_min'] == val_min and
                                  kept['val_max'] == val_max)
                    if same_value:
                        kept['collapsed'] += 1
                        kept['quote_date'] = _merge_date(kept['quote_date'],
                                                         entry['quote_date'])
                        continue
                elif {kept['status'], status} != {STATUS_QUOTED, STATUS_NO_BUY}:
                    vencedor = kept if rank[kept['status']] >= rank[status] else entry
                    vencedor['collapsed'] = kept['collapsed'] + 1
                    vencedor['quote_date'] = _merge_date(kept['quote_date'],
                                                         entry['quote_date'])
                    plan[key] = vencedor
                    continue
                conflicts.append(
                    (kept['where'], where,
                     f'{brand_name or "Genérica"} {kind}/{gen} '
                     f'{tier.normalize():f}{KIND_UNIT[kind]}: '
                     f'{kept["status"]}:¥{kept["val_min"]}–{kept["val_max"]} '
                     f'≠ {status}:¥{val_min}–{val_max}'))

        # ── Marcas precisam existir ANTES (load_brands) ──────────────────────
        brand_cache, missing = {}, set()
        for (brand_name, *_rest) in plan:
            if brand_name is None or brand_name in brand_cache:
                continue
            b = self._brand_for(brand_name)
            if b is None:
                missing.add(brand_name)
            else:
                brand_cache[brand_name] = b

        # ── Relatório ────────────────────────────────────────────────────────
        by_status = {STATUS_QUOTED: 0, STATUS_NO_BUY: 0, STATUS_UNQUOTED: 0}
        collapsed_total = 0
        for entry in plan.values():
            by_status[entry['status']] += 1
            collapsed_total += entry['collapsed']

        self.stdout.write(
            f"Planilha: {opts['xlsx']}  ·  câmbio por aba: "
            + ', '.join(f'{s}={r}' for s, r in rates.items()))
        self.stdout.write(
            f"Linhas de preço: {len(plan)}  "
            f"(cotadas {by_status[STATUS_QUOTED]} · não-compra "
            f"{by_status[STATUS_NO_BUY]} · aguardando {by_status[STATUS_UNQUOTED]}"
            f" · combos eMCP colapsados: {collapsed_total})")
        for where, reason in skipped:
            self.stdout.write(self.style.WARNING(f'  ⚠ pulada {where}: {reason}'))
        for w1, w2, msg in conflicts:
            self.stdout.write(self.style.ERROR(f'  ✗ CONFLITO {w1} × {w2}: {msg}'))
        if conflicts:
            raise CommandError(
                f'{len(conflicts)} conflito(s) de preço na MESMA faixa — a regra '
                '"cota por faixa" foi contrariada na planilha. Resolva lá (ou '
                'com o comprador) e re-rode. NADA foi gravado.')
        if missing:
            raise CommandError(
                'Marca(s) da planilha sem cadastro no catálogo: '
                f'{sorted(missing)}. Rode load_brands/verifique o nome. '
                'NADA foi gravado.')

        if not opts['commit']:
            self.stdout.write(self.style.WARNING(
                'DRY-RUN — nada gravado. Revise e re-rode com --commit.'))
            return

        # ── Escrita (upsert idempotente, tudo-ou-nada) ───────────────────────
        with transaction.atomic():
            buyer, buyer_created = Buyer.all_companies.get_or_create(
                slug=opts['buyer'],
                defaults={'name': opts['buyer'].replace('-', ' ').title(),
                          'company': company})
            # Cura comprador ÓRFÃO (company=NULL, criado antes do escopo):
            # get_or_create não atualiza defaults no "get" — e um Buyer sem
            # empresa é INVISÍVEL ao Buyer.objects escopado (o admin não vê
            # preço nenhum no card). NULL fica reservado ao marketplace futuro
            # (§3.1), que será atribuição EXPLÍCITA, nunca acidente de import.
            if not buyer_created and buyer.company_id is None:
                buyer.company = company
                buyer.save(update_fields=['company'])
                self.stdout.write(self.style.WARNING(
                    f'⚠ Comprador {buyer.slug!r} estava SEM empresa (órfão) — '
                    f'atribuído à {company.name}.'))
            lists_created = rows_created = rows_updated = 0
            list_cache: dict = {}

            def _list_for(brand_name):
                if brand_name not in list_cache:
                    pl, created = PriceList.all_companies.get_or_create(
                        buyer=buyer,
                        brand=brand_cache[brand_name] if brand_name else None)
                    list_cache[brand_name] = pl
                    nonlocal lists_created
                    lists_created += created
                return list_cache[brand_name]

            for (brand_name, kind, gen, tier), e in plan.items():
                pl = _list_for(brand_name)
                fields = dict(
                    status=e['status'], price_min=e['val_min'],
                    price_max=e['val_max'], quote_date=e['quote_date'],
                    source=e['source'] or f"import {opts['xlsx']}",
                    notes=e['notes'])
                obj = Price.all_companies.filter(
                    price_list=pl, kind=kind, gen=gen,
                    tier_value=tier, tier_unit=KIND_UNIT[kind]).first()
                if obj is None:
                    Price.all_companies.create(
                        price_list=pl, kind=kind, gen=gen, tier_value=tier,
                        tier_unit=KIND_UNIT[kind], **fields)
                    rows_created += 1
                else:
                    changed = False
                    for f, v in fields.items():
                        if getattr(obj, f) != v:
                            setattr(obj, f, v)
                            changed = True
                    if changed:
                        obj.save()
                        rows_updated += 1

        self.stdout.write(self.style.SUCCESS(
            f"✅ COMMIT: comprador {buyer.name!r}"
            f"{' (criado)' if buyer_created else ''} · listas novas: "
            f"{lists_created} · preços criados: {rows_created} · atualizados: "
            f"{rows_updated}."))
        self.stdout.write(
            '   Herança (regras do comprador — ex.: genérica → Nanya, SK → '
            'Samsung) é configurada no admin, em Listas de preços (§4 do '
            'PRECIFICACAO.md). A planilha está oficialmente APOSENTADA.')
