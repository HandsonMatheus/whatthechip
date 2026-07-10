"""Exporta para .xlsx os known_parts *identity-only* (Problema A do PLANO_QUALIDADE_DADOS.md).

PORQUÊ: `audit_known_parts --empty` já lista no terminal os confirmed/manual SEM spec
própria — mas o dono quer ABRIR a lista numa planilha para analisar (ordenar por marca,
por tipo, separar os que a gramática cobre dos que não cobre). Este comando é a MESMA
seleção do audit, só que materializada num Excel com uma aba de resumo.

READ-ONLY: não escreve NADA no banco. Espelha exatamente a lógica de `_audit_empty`
(chips/management/commands/audit_known_parts.py) para dar o MESMO número (876 no local,
2026-07-06). Roda com o DATABASE_URL apontando ao banco a inspecionar; rode DEPOIS do
`load_brands --commit` do banco-alvo (a gramática precisa estar carregada, senão o
`_match_family` não casa e tudo cai em "sem família").

Uso:
    python manage.py export_identity_only                       # banco local, out padrão
    python manage.py export_identity_only --out /tmp/wtc.xlsx
    python manage.py export_identity_only --brand Samsung
    python manage.py export_identity_only --confidence confirmed,manual
"""
from collections import Counter

from django.core.management.base import BaseCommand

from chips.models import KnownPart
from chips.engine import _match_family


# Colunas da aba principal (ordem = ordem no Excel).
_HEADERS = [
    "part_number", "marca", "coberto_gramatica", "familia", "grupo",
    "chip_type", "subtype", "confidence", "campos_vazios",
    "device", "fbga_code", "notes", "source_url",
]


def _empty(v) -> bool:
    return not (str(v or "").strip()) or str(v).strip() == "None"


class Command(BaseCommand):
    help = "Exporta os known_parts identity-only (confirmed/manual sem spec própria) para .xlsx."

    def add_arguments(self, parser):
        parser.add_argument("--out", default="identity_only.xlsx",
                            help="Caminho do .xlsx de saída (default: identity_only.xlsx).")
        parser.add_argument("--confidence", default="confirmed,manual",
                            help="Confidences a auditar (default: confirmed,manual).")
        parser.add_argument("--brand", default="", help="Filtra por marca (nome).")
        parser.add_argument("--family", default="",
                            help="Limita a prefixos de família (ex.: KMF,KMQ,KMR).")

    def handle(self, *args, **o):
        w = self.stdout.write
        confs = tuple(c.strip() for c in o["confidence"].split(",") if c.strip())
        fam_prefixes = {f.strip().upper() for f in o["family"].split(",") if f.strip()}

        qs = KnownPart.objects.all()
        if confs:
            qs = qs.filter(confidence__in=confs)
        if o["brand"]:
            qs = qs.filter(brand__name__iexact=o["brand"])

        rows = []
        total = 0
        by_brand, by_grupo, by_cobertura = Counter(), Counter(), Counter()

        for kp in qs.select_related("family", "brand").iterator():
            # MESMA lógica de _audit_empty: a família vem do decode ao vivo (ou do FK).
            fam = _match_family(kp.part_number) or kp.family
            grupo = fam.prefix if fam else (kp.chip_type or "?")
            if fam_prefixes and grupo.upper() not in fam_prefixes:
                continue
            total += 1

            emcp_like = (fam.is_emcp if fam else False) or \
                (kp.chip_type or "").lower() in ("emcp", "umcp")
            if emcp_like:
                spec_fields = ("emcp_ram", "emcp_nand")
            else:
                spec_fields = ("capacity", "density_gbit")

            vazios = [f for f in spec_fields if _empty(getattr(kp, f, ""))]
            if len(vazios) != len(spec_fields):
                continue  # tem ao menos uma spec → NÃO é identity-only

            coberto = "SIM" if fam else "NAO"
            by_brand[kp.brand.name if kp.brand else "?"] += 1
            by_grupo[grupo] += 1
            by_cobertura[coberto] += 1

            rows.append([
                kp.part_number,
                kp.brand.name if kp.brand else "",
                coberto,
                fam.prefix if fam else "",
                grupo,
                kp.chip_type or "",
                kp.subtype or "",
                kp.confidence,
                ", ".join(vazios),
                kp.device or "",
                kp.fbga_code or "",
                (kp.notes or "").replace("\n", " ").strip(),
                kp.source_url or "",
            ])

        # Ordena por: sem-cobertura primeiro (os piores), depois marca, grupo, PN.
        rows.sort(key=lambda r: (r[2] == "SIM", r[1], r[4], r[0]))

        self._write_xlsx(o["out"], rows, by_brand, by_grupo, by_cobertura, total)

        w("")
        w(f"Auditados (confirmed/manual): {total}  ·  identity-only: {len(rows)}")
        w(f"  coberto pela gramatica: SIM={by_cobertura.get('SIM',0)} · "
          f"NAO={by_cobertura.get('NAO',0)}")
        if by_brand:
            w("  por marca: " + " · ".join(f"{k}={v}"
              for k, v in sorted(by_brand.items(), key=lambda kv: -kv[1])))
        w(f"\n.xlsx escrito em: {o['out']}  (READ-ONLY, nada foi gravado no banco)")

    # ────────────────────────────────────────────────────────────────────
    def _write_xlsx(self, path, rows, by_brand, by_grupo, by_cobertura, total):
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()

        # ── Aba 1: os PNs ────────────────────────────────────────────────
        ws = wb.active
        ws.title = "identity-only"
        hdr_fill = PatternFill("solid", fgColor="1F3864")
        hdr_font = Font(bold=True, color="FFFFFF")
        ws.append(_HEADERS)
        for c in ws[1]:
            c.fill, c.font = hdr_fill, hdr_font
            c.alignment = Alignment(vertical="center")
        for r in rows:
            ws.append(r)
        # Realça a coluna coberto_gramatica = NAO (os que precisam de pesquisa real).
        nao_fill = PatternFill("solid", fgColor="FCE4D6")
        for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
            if row[0].value == "NAO":
                row[0].fill = nao_fill
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(_HEADERS))}{ws.max_row}"
        widths = [22, 12, 10, 9, 9, 12, 16, 11, 20, 14, 10, 50, 30]
        for i, wd in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = wd

        # ── Aba 2: resumo ────────────────────────────────────────────────
        s = wb.create_sheet("resumo")
        s.append(["RESUMO — identity-only (Problema A do PLANO_QUALIDADE_DADOS.md)"])
        s["A1"].font = Font(bold=True, size=13)
        s.append([])
        s.append(["Auditados (confirmed/manual)", total])
        s.append(["Identity-only (sem spec propria)", sum(by_cobertura.values())])
        s.append(["  coberto pela gramatica (SIM = backfill resolve)", by_cobertura.get("SIM", 0)])
        s.append(["  SEM familia (NAO = precisa pesquisa Tier-1)", by_cobertura.get("NAO", 0)])
        s.append([])
        s.append(["Por marca", "qtd"])
        for k, v in sorted(by_brand.items(), key=lambda kv: -kv[1]):
            s.append([k, v])
        s.append([])
        s.append(["Por grupo (familia ou tipo)", "qtd"])
        for k, v in sorted(by_grupo.items(), key=lambda kv: -kv[1]):
            s.append([k, v])
        s.column_dimensions["A"].width = 48
        s.column_dimensions["B"].width = 10

        wb.save(path)
