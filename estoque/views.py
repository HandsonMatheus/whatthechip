"""
WhatTheChip — Estoque views (com Lotes)
========================================
GET  /estoque/                           → lista de lotes
POST /estoque/novo/                      → cria novo lote
GET  /estoque/lote/<lot_pk>/             → painel do lote
GET  /estoque/lote/<lot_pk>/preview/     → classifica PN (HTMX)
POST /estoque/lote/<lot_pk>/add/         → adiciona chip
POST /estoque/lote/<lot_pk>/remove/<pk>/ → remove entrada
GET  /estoque/lote/<lot_pk>/export/      → exporta .xlsx
POST /estoque/lote/<lot_pk>/fechar/      → fecha lote
POST /estoque/lote/<lot_pk>/reabrir/     → reabre lote
"""

import io
import json
import logging
import re
from datetime import datetime, timedelta
from difflib import get_close_matches
from uuid import uuid4

from django.contrib import messages
from django.core.exceptions import PermissionDenied
from django.db import IntegrityError, transaction
from django.db.models import F, ProtectedError
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.utils.translation import gettext as _
from django.views.decorators.http import require_POST

from tenancy.access import can_sales, role_required
from tenancy.ui import ui   # E5: canary por empresa (§17.7)

from chips.conventions import canonical_gen
from chips.chip_types import canonical_chip_type, generation_of, label_kind
from chips.engine import assess_profitability, classify, is_dead_by_generation
from chips.models import UnknownChip, Brand

from .models import InventoryEntry, Lot, PendingEntry, RejectedEntry, SubmitToken

logger = logging.getLogger(__name__)


# Elegibilidade de entrada no estoque: passa quem TEM REGISTRO no banco — confidence
# confirmed, manual OU distributor. Decisão do dono (2026-07-08): distribuidor ENTRA;
# gramática PURA (sem registro no banco) JAMAIS entra → vai pra fila.
# ⚠ Isto é ELEGIBILIDADE DE ESTOQUE, não AUTORIDADE sobre a gramática: distribuidor
# continua NÃO vencendo a gramática no engine (chips/engine.py::_CONFIRMED_CONFIDENCE
# segue só confirmed/manual). As specs de um distribuidor vêm da gramática; ter o
# registro só o torna elegível pra bancada.
CONFIRMED_SOURCES = {"banco de dados"}
CONFIRMED_CONF = {"confirmed", "manual", "distributor"}


def _is_confirmed(result: dict) -> bool:
    """True se o PN é ELEGÍVEL PRO ESTOQUE: tem registro no banco (confidence
    confirmed/manual/distributor) OU a classificação veio do banco. Gramática pura
    (sem registro) → False → fila. (O nome '_is_confirmed' é histórico; hoje inclui
    distribuidor por decisão do dono — ver comentário acima.) Reavaliado no servidor —
    nunca confia no campo hidden do formulário."""
    return (
        result.get("classification_source") in CONFIRMED_SOURCES
        or result.get("confidence") in CONFIRMED_CONF
    )


def _display_source(result: dict) -> str:
    """Rótulo 'Source' na camada de ESTOQUE (card, tabela e export .xlsx) — FONTE ÚNICA.

    Regra: tudo que é ELEGÍVEL ao estoque passou pelo gate ``_is_confirmed`` — ou
    seja, TEM registro no banco (confidence confirmed/manual/distributor). Do ponto
    de vista do estoque a origem é, portanto, o BANCO — inclusive quando as specs de
    um registro DISTRIBUIDOR foram completadas pela gramática (o motor marca
    ``classification_source='gramática'`` para sinalizar isso no SITE de
    identificação; no estoque esse rótulo confundia — parecia chip SEM registro,
    contradizendo a regra "só entra quem tem registro"). Diagnóstico: lote 41,
    2026-07-13.

    ``known_exact`` cobre o confirmado SEM família casada (ex.: Micron JZ###), cujo
    ``classification_source`` pode vir vazio. Fora desses casos (fila / itens sem
    registro no banco), preserva o rótulo cru do motor.

    ⚠ NÃO altera o motor nem o site: é só a TRADUÇÃO de exibição do estoque. O gate
    (``_is_confirmed``) continua lendo o ``classification_source`` CRU do motor.
    """
    if _is_confirmed(result):
        return "banco de dados"
    src = result.get("classification_source") or ""
    if not src and result.get("known_exact"):
        return "banco de dados"
    return src


def _size_for_entry(result: dict) -> str:
    """Tamanho a GRAVAR no estoque. `capacity` (bytes) tem prioridade; para DRAM
    standalone (DDR/GDDR) a capacity vem vazia e o tamanho está em `dram_density`
    ('2Gb = 256MB por die') — extraímos os bytes ('256MB') para não perder o dado.
    Antes este tamanho era simplesmente perdido (estoque gravava vazio → 'None').
    (Regex _CAP_BYTES_RE/_GBIT_RE definidos abaixo; resolvidos em tempo de chamada.)"""
    ct = (result.get("chip_type") or "").upper()
    # DRAM discreta de 1 die (DDR/GDDR/SDRAM/RDRAM) → DENSIDADE em Gbit, formato '2G'
    # (mesma convenção da etiqueta da caixa 'DDR3+2G'). _density_g lê de dram_density
    # ou deriva da capacity em bytes — robusto aos dois jeitos de gravar no catálogo.
    if (ct.startswith("DDR") or ct.startswith("GDDR") or ct in ("SDRAM", "RDRAM")) and "LPDDR" not in ct:
        g = _density_g(result)
        return f"{g}G" if g else ""
    # LPDDR/eMMC/UFS (pacote) → CAPACIDADE em bytes (GB). 'None' (string) = lixo de
    # catálogo → trata como vazio; normaliza o espaço ('256 MB'→'256MB').
    cap = (result.get("capacity") or "").strip()
    if cap and cap.lower() != "none":
        return re.sub(r"\s+([KMGT]B)\b", r"\1", cap)
    # fallback raro: tamanho só em dram_density. bytes têm 'B' MAIÚSCULO; 'Gb' é gigaBIT —
    # case-SENSITIVE (sem re.I) p/ não ler 'Gb' como 'GB' (bug clássico 8×).
    dd = result.get("dram_density") or ""
    m = re.search(r"(\d+(?:\.\d+)?)\s*([TGM]B)\b", dd)
    if m:
        return f"{m.group(1)}{m.group(2)}"
    g = _GBIT_RE.search(dd)
    if g:
        mb = float(g.group(1)) * 128
        return f"{mb:g}MB" if mb < 1024 else f"{mb / 1024:g}GB"
    return ""


def _clean_interface(result: dict) -> str:
    """`interface` no estoque = bus width (x16) ou versão (eMMC 5.1) — NUNCA a
    geração (essa vive no chip_type/subtype/label). Remove a geração espelhada
    (DDR3, LPDDR4X, GDDR5…) para o campo ficar CONSISTENTE entre os tipos (era a
    origem do 'alguns espelham, outros não'). Não afeta o label da caixa, que usa
    o `result` cru, não este campo gravado."""
    ifc = (result.get("interface") or "").strip()
    if ifc and re.fullmatch(r"(LP)?DDR\d[A-Z]?|GDDR\d", ifc, re.I):
        return ""
    return ifc


def _snapshot(result: dict) -> dict:
    """Snapshot da classificação para InventoryEntry/PendingEntry/RejectedEntry —
    SEMPRE a partir do classify do SERVIDOR (nunca do POST do cliente).

    Coage None → '' porque os CharFields são NOT NULL com default ''. O
    classify() devolve None em emcp_ram/emcp_nand para chips que NÃO são eMCP
    (ex.: LPDDR2 avulso), e `.get(chave, '')` não cobre esse caso (a chave
    existe com valor None) — daí o NotNullViolation no insert.
    `capacity` captura a densidade DRAM via _size_for_entry; `interface` é limpa
    da geração espelhada via _clean_interface."""
    return {
        "chip_type":             result.get("chip_type") or "",
        "brand":                 result.get("brand") or "",
        "capacity":              _size_for_entry(result),
        "emcp_ram":              result.get("emcp_ram") or "",
        "emcp_nand":             result.get("emcp_nand") or "",
        "is_emcp":               bool(result.get("is_emcp")),
        "interface":             _clean_interface(result),
        "classification_source": _display_source(result),
        "confidence":            result.get("confidence") or "",
    }


def _price_key_fields(result: dict) -> dict:
    """F11.1: deriva a CHAVE DE PREÇO do classify e devolve os campos do
    InventoryEntry — gravada no LANÇAMENTO para a valoração resolver contra a
    tabela Price viva sem reclassificar o lote na leitura. Sem chave (NO_KEY)
    grava o MOTIVO (aparece no sem-preço do export/valoração)."""
    from pricing.engine import derive_price_key   # lazy (padrão da F8)
    err, key = derive_price_key(result or {})
    if err is not None:
        return {'price_kind': err.kind or '', 'price_gen': '',
                'price_tier_value': None, 'price_tier_unit': '',
                'price_key_reason': err.reason[:200]}
    kind, gen, tier_value, tier_unit = key
    return {'price_kind': kind, 'price_gen': gen,
            'price_tier_value': tier_value, 'price_tier_unit': tier_unit,
            'price_key_reason': ''}


def _masked_category(result: dict):
    """F12 v3: (código da caixa, é_hold?) do resultado — o rótulo que a
    empresa-CLIENTE vê no lugar de tipo/specs. A categoria deriva do CHIP
    (convenção universal LETRA-##, pricing/convention.py) e existe COM ou
    SEM preço — "preço até pode ficar sem, categoria não" (dono 2026-07-23).
    Categoria inédita é cunhada AQUI (caminho da aprovação — próximo número
    livre da letra). Sem categoria derivável (dado incompleto, raro) →
    **H-00 HOLD**: não está pronto pra prateleira, separar p/ análise (o
    conceito 'Geral/C-000' foi DESFEITO — dono 2026-07-23)."""
    from pricing.convention import HOLD_LABEL
    from pricing.engine import derive_price_key
    from pricing.models import CategoryCode
    err, key = derive_price_key(result or {})
    if err is None:
        label = CategoryCode.label_for_key(*key)
        if label is not None:
            return label, False
    return HOLD_LABEL, True


def _masked_entry_labels(entries):
    """F12 v3: anexa ``entry.category_label`` (LETRA-##) numa passada só —
    lookup em lote (sem N+1) pros renders mascarados da tabela do lote. A
    geração da chave gravada DOBRA na base (fold_gen). LEITURA NUNCA CUNHA
    código (cunhagem é só na aprovação — evita ressuscitar categoria morta
    de entrada legada); sem código/sem chave → '—'."""
    from pricing.convention import KIND_LETTER
    from pricing.models import CategoryCode, fold_gen
    keyed = {(c.kind, c.gen, c.tier_value, c.tier_unit):
             f'{KIND_LETTER[c.kind]}-{c.code:02d}'
             for c in CategoryCode.objects.all() if c.kind in KIND_LETTER}
    for e in entries:
        if e.price_tier_value is None:
            e.category_label = '—'
        else:
            k = (e.price_kind, fold_gen(e.price_kind, e.price_gen),
                 e.price_tier_value, e.price_tier_unit)
            e.category_label = keyed.get(k, '—')
    return entries


def _nearest_in_lot(lot, pn: str) -> str:
    """PN já existente no lote mais parecido — provável original de um typo."""
    pool = list(lot.entries.values_list("part_number", flat=True))
    near = get_close_matches(pn, [p for p in pool if p != pn], n=1, cutoff=0.8)
    return near[0] if near else ""


# ─── helpers ────────────────────────────────────────────────────────────────

def _normalise_pn(raw: str) -> str:
    return re.sub(r'[^A-Z0-9\-]', '', (raw or '').strip().upper())


#: Pseudo-códigos de TIPO aceitos como "PN" na bancada (dono 2026-08-14,
#: HANDOFF_K9): "K9" tem 2 chars — exceção ao mínimo de 4. O operador digita
#: o nome da categoria (triagem por FORMATO, sem PN); _normalise_pn já
#: uppercaseou ("k9" → "K9"). Espelhado no JS da bancada (estoque.html,
#: pnTooShort).
_TYPE_PSEUDO_PNS = frozenset({'K9'})


def _pn_too_short(pn: str) -> bool:
    """Curto demais pra bancada? (mínimo 4 chars, exceto pseudo-códigos)."""
    return len(pn) < 4 and pn not in _TYPE_PSEUDO_PNS


#: Token de idempotência do add_chip (bug Mundo Metal, 2026-08-10) — formato
#: uuid4().hex; TTL da poda lazy em horas (48h cobre folgado qualquer aba velha).
_TOKEN_RE = re.compile(r'[0-9a-f]{32}')
_TOKEN_TTL_H = 48


def _claim_submit_token(request) -> bool:
    """Idempotência do ``add_chip`` (bug Mundo Metal LOT/002/08/26, 2026-08-10).
    True = este POST é REENVIO de um clique JÁ APLICADO — o chamador re-renderiza
    o estado atual SEM escrever nada (não soma quantidade, não duplica log).

    Como funciona: cada render do card de triagem gera um UUID (hidden
    ``submit_token``); a 1ª request CRIA a linha (unique no banco) e ganha o
    direito de escrever; a duplicata leva IntegrityError → True. O atomic()
    interno é SAVEPOINT: não envenena a transação da request (o
    TenancyMiddleware abre o atomic externo). Sem token (página aberta antes
    do deploy) → False, comportamento antigo.

    Por que isso conserta a rede LENTA (Venezuela): (a) duplo clique/Enter+
    clique geravam 2 POSTs e o servidor somava 2×; (b) quando a RESPOSTA se
    perdia, o operador relançava sem saber se chegou. Com o token + o card
    permanecendo em falha (after-request só limpa em sucesso), o re-clique
    reusa o MESMO token: se o 1º POST chegou, vira no-op; se não chegou,
    aplica. Nos dois casos a quantidade fica certa."""
    tok = (request.POST.get('submit_token') or '').strip().lower()
    if not _TOKEN_RE.fullmatch(tok):
        return False
    try:
        with transaction.atomic():
            SubmitToken.objects.create(token=tok)
    except IntegrityError:
        return True
    # Poda lazy dos tokens velhos (indexado; normalmente 0 linhas afetadas).
    SubmitToken.objects.filter(
        created_at__lt=timezone.now() - timedelta(hours=_TOKEN_TTL_H)).delete()
    return False


_PLACEHOLDER_MARKERS = ("não mapead", "nao mapead", "consultar datasheet")


def _real_spec(val) -> bool:
    """True só se o valor é uma spec REAL — não um placeholder de gramática
    incompleta (ex.: "Código 'BG' não mapeado — consultar datasheet")."""
    if not val:
        return False
    low = str(val).lower()
    return not any(m in low for m in _PLACEHOLDER_MARKERS)


def _has_capacity(result: dict) -> bool:
    # Considera o chip "identificável" se tiver capacidade REAL em qualquer campo,
    # OU se for um KnownPart confirmado (known_exact=True) com chip_type definido.
    # ⚠ Placeholder de gramática incompleta ("código não mapeado — consultar
    # datasheet") NÃO conta: senão o operador vê um card confiante (ex.: "DDR4"
    # sem specs) e pode encaixotar um chip que na verdade não foi reconhecido.
    # Esses caem no fluxo de "Não identificado" / UnknownChip.
    return bool(
        _real_spec(result.get('capacity'))
        or _real_spec(result.get('emcp_ram'))
        or _real_spec(result.get('emcp_nand'))
        or _real_spec(result.get('dram_density'))
        or (result.get('known_exact') and result.get('chip_type'))
    )


def _entries_qs(lot, q='', tipo=''):
    qs = InventoryEntry.objects.filter(lot=lot)
    if q:
        qs = qs.filter(part_number__icontains=q)
    if tipo:
        qs = qs.filter(chip_type__icontains=tipo)
    return qs.order_by('-last_updated')


#: F11.0b (2026-07-16): linhas por página do "Estoque do lote". A página
#: renderizava TODAS as entradas (lote 42 = ~700KB de HTML por request).
_PAGE_SIZE = 100


def _paginate_entries(request, entries):
    """(entradas da página, page_obj) — pagina a LISTA já montada (a ordem
    -last_updated garante o recém-lançado na página 1). Filtro/busca resetam
    para a página 1 sozinhos (o hx-get deles não envia ?p=); POSTs de
    add/remove idem. A valoração e o export seguem cobrindo o lote INTEIRO —
    isto é só exibição."""
    from django.core.paginator import Paginator
    page_obj = Paginator(entries, _PAGE_SIZE).get_page(request.GET.get('p'))
    return list(page_obj.object_list), page_obj


def _current_snapshot(pn: str) -> dict:
    """Snapshot ATUAL do servidor para um PN: `_snapshot(classify(pn))` sem o
    campo `confidence`. O rótulo 'Source' (incl. a tradução distribuidor/gramática →
    "banco de dados" e o confirmado SEM família casada) é derivado por
    `_display_source`, dentro do `_snapshot` — fonte única, igual ao intake e ao
    `resnapshot_lote`/`refresh_lote`. Devolve só os campos de exibição
    (chip_type/brand/capacity/emcp_*/is_emcp/interface/Source)."""
    r = classify(pn) or {}
    snap = _snapshot(r)
    snap.pop("confidence", None)
    return snap


#: Teto de recálculo SÍNCRONO por render do on-read. O caminho que persiste em
#: massa é o `resnapshot_lote`; a tela só reconcilia o que estiver à vista, sem
#: varrer o lote inteiro a cada abertura (problema 4.4 do BRIEFING_ESCALABILIDADE).
_ONREAD_CAP = 150


def _entries_for_display(lot, q='', tipo=''):
    """Entradas do lote prontas para a TELA, com **cálculo na leitura** (on-read):
    as DEFASADAS — `snapshot_catalog_version` menor que `CatalogVersion.current()`,
    i.e. o catálogo melhorou desde que o chip foi lançado — são recalculadas EM
    MEMÓRIA para mostrar o valor ATUAL, **sem gravar** (quem persiste é o
    `resnapshot_lote`). Limita-se a `_ONREAD_CAP` recálculos por render para não
    classificar o lote inteiro de forma síncrona. O que está em dia sai do banco
    sem custo."""
    from chips.models import CatalogVersion
    cur = CatalogVersion.current()
    entries = list(_entries_qs(lot, q, tipo))
    recomputed = 0
    for e in entries:
        if e.snapshot_catalog_version < cur and recomputed < _ONREAD_CAP:
            for k, v in _current_snapshot(e.part_number).items():
                setattr(e, k, v)   # override em memória — NÃO chama e.save()
            recomputed += 1
    return entries


def _get_lot(request, lot_pk):
    """Lote acessível ao usuário. T1 (papéis, §8 do plano): o lote é um ATIVO DA
    EMPRESA — o gerente abre, o operador lança nele — então caiu o filtro
    ``operator=request.user`` (cada um só via os próprios lotes, o que impediria
    o operador de trabalhar num lote aberto pelo gerente). O campo ``operator``
    do Lot vira "quem abriu". T3: escopado por empresa via ``Lot.objects``
    (EXPLÍCITO — o _default_manager voltou a ser o cru; ver estoque/models.py):
    lote de outra empresa é 404, como se não existisse."""
    return get_object_or_404(Lot.objects, pk=lot_pk)


def _extract_gb(text: str) -> str:
    """
    Extrai o valor em GB de uma string de capacidade.
    '8GB' → '8'  |  '1.5GB' → '1.5'  |  'eMMC 5.1 16GB' → '16'
    Usa look-behind negativo para não capturar o ".1" de "eMMC 5.1 8GB".
    """
    if not text:
        return ''
    # Captura número decimal ou inteiro seguido de GB,
    # mas exige que não seja precedido por outro dígito (evita pegar "5.1" de "eMMC 5.1")
    m = re.search(r'(?<!\d)(\d+(?:\.\d+)?)\s*GB', text, re.IGNORECASE)
    if not m:
        return ''
    val = m.group(1)
    # Remove ".0" redundante: "8.0" → "8"
    if val.endswith('.0'):
        val = val[:-2]
    return val


# "Gb" = gigabit (≠ "GB" gigabyte). Densidade DRAM por die.
_GBIT_RE = re.compile(r'(\d+(?:\.\d+)?)\s*Gb\b')
_CAP_BYTES_RE = re.compile(r'(\d+(?:\.\d+)?)\s*(TB|GB|MB)\b', re.I)  # capacidade em bytes (TB adicionado 2026-06-26)


def _format_cap(text: str) -> str:
    """Capacidade em MB ou GB, preservando a unidade original.
    '512MB' → '512MB'  |  '8GB' → '8GB'  |  'eMMC 5.1 16GB' → '16GB'
    Não converte MB→GB (512MB é 512MB, não 0.5GB).
    Retorna '' se não encontrar nenhuma unidade reconhecida."""
    if not text:
        return ''
    m = _CAP_BYTES_RE.search(text)
    if not m:
        return ''
    val, unit = m.group(1), m.group(2).upper()
    if val.endswith('.0'):
        val = val[:-2]
    return f"{val}{unit}"


def _density_g(result: dict) -> str:
    """Densidade DRAM por die em Gb, para o rótulo da caixa física (2G/4G/8G).
    Em ordem de prioridade:
      1) `dram_density` (campo canônico, já em Gb): '2Gb = …' → '2';
      2) fallback — deriva da `capacity` em bytes. Em DRAM standalone (1 die por
         chip) a capacidade do CHIP é a própria densidade do die:
         256MB→2G, 512MB→4G, 1GB→8G  (bytes ×8 ÷1024 = Gbit). Muitos registros
         CONFIRMADOS guardam a densidade só como bytes em `capacity`, deixando
         `dram_density` vazio — sem este fallback o rótulo perderia o '+2G'.
    Nunca devolve os bytes crus (256MB confunde — a caixa é por densidade).
    Agnóstico de marca: serve a qualquer fabricante, com ou sem dram_density."""
    m = _GBIT_RE.search(result.get('dram_density') or '')
    if m:
        v = m.group(1)
        return v[:-2] if v.endswith('.0') else v
    c = _CAP_BYTES_RE.search(result.get('capacity') or '')
    if c:
        mb = float(c.group(1)) * (1024 if c.group(2).upper() == 'GB' else 1)
        return f"{mb / 128:g}"   # MB → Gbit  (1Gb = 128 MB)
    return ''


def _capacity_g(result: dict) -> str:
    """Capacidade total do PACOTE em GB, p/ o rótulo de caixa de LPDDR (móvel,
    multi-die): '4GB'→'4', '2GB'→'2', '512MB'→'0.5'. LPDDR é triado pela
    capacidade do pacote (não pela densidade do die), porque empilha vários dies —
    tratar a capacidade como densidade de 1 die daria valor absurdo (4GB→32Gbit)."""
    c = _CAP_BYTES_RE.search(result.get('capacity') or '')
    if not c:
        return ''
    gb = float(c.group(1)) * (1 if c.group(2).upper() == 'GB' else 1 / 1024)
    return f"{gb:g}"


def _compute_destination(result: dict) -> tuple:
    """
    Return (label, category) for the physical storage bin.
    category is used as CSS modifier:
      emcp | umcp | ssd | k9 | lpddr | ddr | gddr | ufs | emmc | nand | unknown
    """
    chip_type = (result.get('chip_type') or '').strip()
    ct = chip_type.lower()
    # Tipo de label da FONTE ÚNICA (chips/chip_types.py) + fallback de substring,
    # que preserva casos legados (ex.: 'onenand' contém 'nand'). A GERAÇÃO do label
    # vem de canonical_gen(subtype) com fallback no chip_type via generation_of.
    kind = label_kind(canonical_chip_type(chip_type, result.get('subtype') or ''))

    if kind == 'umcp' or 'umcp' in ct:
        nand  = _extract_gb(result.get('emcp_nand', ''))
        ram   = _extract_gb(result.get('emcp_ram', ''))
        label = f"UMCP{nand}+{ram}" if nand else 'uMCP'
        return label, 'umcp'

    if kind == 'emcp' or 'emcp' in ct or result.get('is_emcp'):
        nand  = _extract_gb(result.get('emcp_nand', ''))
        ram   = _extract_gb(result.get('emcp_ram', ''))
        label = f"EMCP{nand}+{ram}" if nand else 'eMCP'
        return label, 'emcp'

    if kind == 'ssd':
        # SSD BGA/NVMe (dono 2026-07-24): comprado por GB; caixa SSD+capacidade.
        cap   = _format_cap(result.get('capacity', ''))
        label = f"SSD{cap}" if cap else 'SSD'
        return label, 'ssd'

    if kind == 'k9':
        # K9 (dono 2026-08-14, HANDOFF_K9): caixa ÚNICA do tipo — sem marca,
        # sem capacidade; o nome de mercado É o rótulo (código: K-01).
        return 'K9', 'k9'

    if kind == 'ufs' or 'ufs' in ct:
        # _format_cap preserva a unidade original: "128GB"→"128GB", "1TB"→"1TB".
        # Antes usava _extract_gb + "GB" hardcoded, o que produzia "UFS" (label vazio)
        # para chips 1TB — _extract_gb só reconhecia GB, não TB (fix 2026-06-26).
        cap   = _format_cap(result.get('capacity', ''))
        label = f"UFS{cap}" if cap else 'UFS'
        return label, 'ufs'

    if kind == 'emmc' or 'emmc' in ct:
        cap   = _format_cap(result.get('capacity', ''))
        label = f"EMMC{cap}" if cap else 'eMMC'
        return label, 'emmc'

    # GDDR antes do bloco DDR — 'ddr' in 'gddr3' seria True (substring), causando
    # falso-positivo. Samsung usa chip_type dedicado ("GDDR3"/"GDDR5"/etc.);
    # Hynix H5RS usa chip_type="RAM" com subtype="GDDR3" (coberto pelo elif).
    subtype_lower = (result.get('subtype') or '').lower()
    if kind == 'gddr' or 'gddr' in ct or ('gddr' in subtype_lower and ct in ('ram', 'dram', 'sdram')):
        gen  = canonical_gen(result.get('subtype') or '', chip_type) \
            or generation_of(chip_type, result.get('subtype') or '') or chip_type.upper()
        size = _density_g(result)
        label = f"{gen}+{size}G" if (gen and size) else (gen or 'GDDR')
        return label, 'gddr'

    if kind in ('lpddr', 'ddr', 'sdram') or 'lpddr' in ct or 'ddr' in ct or ct in ('ram', 'dram', 'sdram'):
        # Caixa de RAM = GERAÇÃO + DENSIDADE, no formato impresso na caixa física,
        # ex.: "DDR3+2G", "DDR4+4G", "LPDDR4+8GB". A geração vem de subtype (ex.: "DDR3",
        # "LPDDR4X") — campo específico para o tipo de RAM. `interface` é a config de
        # barramento (ex.: "x16 @ 800MHz (1600MTPS)") e NÃO é usada como label da caixa.
        # Densidade (Gb) vem de dram_density; capacity (bytes) é ignorada de propósito.
        # canonical_gen (chips/conventions.py, FONTE ÚNICA) reduz o subtype ao
        # token de geração para o label: "LPDDR4 Mobile"→"LPDDR4", "DDR3 SDRAM"→
        # "DDR3", "LPDDR4X Multi-Channel"→"LPDDR4X". Generaliza o antigo strip de
        # "SDRAM" e cobre TODAS as marcas e os dois caminhos (banco e gramática),
        # de forma retroativa, sem reescrever o banco. Fallback p/ interface se o
        # subtype estiver vazio (comportamento anterior preservado).
        gen = canonical_gen(result.get('subtype') or '', chip_type) \
            or generation_of(chip_type, result.get('subtype') or '') \
            or (result.get('interface') or '').strip()
        # Tamanho da caixa depende do TIPO (unidade explícita no sufixo):
        #  • LPDDR (móvel) = pacote multi-die → CAPACIDADE do pacote em GB (4GB→"4GB");
        #  • DDR (componente) = 1 die → DENSIDADE do die em Gbit (1GB/die→"8G").
        # "G" sozinho = Gigabits (convenção do setor); "GB" = Gigabytes.
        # UFS/eMMC já usam "GB" explícito; LPDDR alinhado neste padrão.
        if 'lpddr' in ct or gen.upper().startswith('LPDDR'):
            size = _capacity_g(result)
            unit = 'GB'
        else:
            size = _density_g(result)
            unit = 'G'
        if gen and size:
            label = f"{gen}+{size}{unit}"
        elif size:
            label = f"{size}{unit}"
        elif gen:
            label = gen
        else:
            label = 'RAM'
        cat = 'lpddr' if ('lpddr' in ct or gen.upper().startswith('LPDDR')) else 'ddr'
        return label, cat

    if kind == 'nand' or 'nand' in ct:
        # Usa subtype como prefixo do rótulo (ex.: "SLC NAND") + capacidade na
        # unidade original (MB ou GB). _extract_gb só lê GB e perderia "512MB".
        # canonical_gen normaliza a célula: "SLC NAND paralela industrial" → "SLC NAND".
        gen     = canonical_gen(result.get('subtype') or '', chip_type)
        cap_str = _format_cap(result.get('capacity') or '')
        prefix  = gen if gen else 'NAND'
        label   = f"{prefix} {cap_str}" if cap_str else prefix
        return label, 'nand'

    return chip_type or '?', 'unknown'


# ─── gateway de triagem ───────────────────────────────────────────────────────

# Rótulos comerciais da rentabilidade → chave CSS curta (reusados no template).
_PROFIT_KEY = {
    'RENTÁVEL':      'rentavel',
    'NÃO RENTÁVEL':  'nao_rentavel',
    'INDETERMINADO': 'indeterminado',
}


def _compute_gateway(result: dict, has_cap: bool) -> dict:
    """
    Decide o destino de triagem de um chip em 3 etapas de funil (a primeira que
    falha decide), mais um sinal de digitação em paralelo.

    Ordem (importa!): identificação → fonte → rentabilidade.
      1. Identificação: tem specs reais (has_cap)?  Não → 'desconhecido'.
      2. Fonte: confirmado no banco (_is_confirmed)? Não → 'fila' (gestor revisa).
      3. Rentabilidade: assess_profitability. 'NÃO RENTÁVEL' → 'reprovado';
         'RENTÁVEL' ou 'INDETERMINADO' → 'aprovado'.

    Regra de negócio (REVISADA pelo dono, 2026-07-31): INDETERMINADO continua
    'aprovado' no funil (não é descarte — o material fica na bancada), mas NÃO
    pode ser LANÇADO: `can_add=False` desabilita o botão de adicionar e o
    add_chip recusa no servidor. Entrada no estoque exige rentabilidade
    AVALIADA ('RENTÁVEL'). Antes o INDETERMINADO entrava (regra conservadora);
    o lote 042 mostrou o custo: chip dentro do estoque sem avaliação.

    O typo (fuzzy_suggestions) NÃO é uma etapa: é uma rede de segurança exibida à
    parte, válida em qualquer destino.

    Retorna dict com:
      destination   : 'aprovado' | 'fila' | 'desconhecido' | 'reprovado'
      steps         : 3 dicts {id, label, status: pass|fail|skip, detail}
      typo          : {has: bool, suggestions: list}
      profitable    : string crua de assess_profitability ('' quando não avaliada)
      profitable_key: chave CSS ('rentavel' | 'nao_rentavel' | 'indeterminado')
    """
    fuzzy = result.get('fuzzy_suggestions') or []
    typo = {'has': bool(fuzzy), 'suggestions': fuzzy}
    # i18n: 'id' e 'status' são CHAVES (lógica/CSS — nunca traduzir);
    # 'label' e 'detail' são EXIBIÇÃO (gettext, resolve no idioma da request).
    steps = [
        {'id': 'identificacao', 'label': _('Reconheci'),  'status': 'skip', 'detail': ''},
        {'id': 'fonte',         'label': _('Confirmado'), 'status': 'skip', 'detail': ''},
        {'id': 'rentabilidade', 'label': _('Rentável'),   'status': 'skip', 'detail': ''},
    ]

    def _out(destination, profitable='', by_generation=False):
        return {
            'destination':          destination,
            'steps':                steps,
            'typo':                 typo,
            'profitable':           profitable,
            'profitable_key':       _PROFIT_KEY.get(profitable, 'indeterminado'),
            'reject_by_generation': by_generation,
            # Dono 2026-07-31: lançar exige rentabilidade AVALIADA — aprovado
            # com INDETERMINADO fica com o botão de adicionar DESABILITADO
            # (os outros destinos têm botões próprios: fila/descarte/desconhecido).
            'can_add': destination != 'aprovado' or profitable == 'RENTÁVEL',
        }

    # ── Atalho: morto por GERAÇÃO → reprovado direto ─────────────────────────
    # Tecnologia velha (LPDDR2-, DDR2-, MCP legado) é sucata por fato de mercado.
    # Vale mesmo SEM confirmação no banco e SEM capacidade mapeada — a geração é
    # lida da gramática curada. Só geração; nunca capacidade (limite de negócio).
    if is_dead_by_generation(result) and not _is_confirmed(result):
        steps[0].update(status='pass', detail=_('tipo/geração'))
        steps[1].update(status='fail',
                        detail=result.get('classification_source') or _('gramática'))
        steps[2].update(status='fail', detail=_('Geração não rentável'))
        return _out('reprovado', 'NÃO RENTÁVEL', by_generation=True)

    # ── 1. Identificação (specs reais) ───────────────────────────────────────
    if not has_cap:
        steps[0].update(status='fail', detail=_('specs ausentes'))
        return _out('desconhecido')
    steps[0].update(status='pass', detail=_('specs reais'))

    # ── 2. Fonte (confirmado no banco) — NÃO altera _is_confirmed ─────────────
    if not _is_confirmed(result):
        steps[1].update(status='fail',
                        detail=result.get('classification_source') or _('gramática'))
        return _out('fila')
    steps[1].update(status='pass', detail=_('banco de dados'))

    # ── 3. Rentabilidade (conservador: INDETERMINADO → aprovado) ─────────────
    profitable = assess_profitability(result)
    if profitable == 'NÃO RENTÁVEL':
        steps[2].update(status='fail', detail=_('Não rentável'))
        return _out('reprovado', profitable)

    # RENTÁVEL = verde "sim". INDETERMINADO ENTRA no estoque (regra conservadora), mas
    # NÃO é um "sim" confiante — ganha estado próprio ('warn'/âmbar) pra não mentir ao
    # operador (bug: antes recebia status='pass' e o frontend mostrava "Rentável: sim").
    if profitable == 'RENTÁVEL':
        steps[2].update(status='pass', detail=_('Rentável'))
    else:
        steps[2].update(status='warn', detail=_('Indeterminado (não avaliado)'))
    return _out('aprovado', profitable)


# ─── painel (home pós-login) ─────────────────────────────────────────────────

@role_required('operator')
def painel(request):
    """Home pós-login: responde "o que eu faço agora?" antes de jogar o usuário
    na lista de lotes. Padrão lançadeira (decisão de UX 2026-07-06): o lote
    ABERTO vira o CTA principal ("Continuar triagem"); sem lote aberto, o
    empty-state orienta por papel (gerente: abrir lote; operador: pedir ao
    gerente). Stats do dia são contexto, não o foco — o produto é a bancada.
    T3: tudo aqui passa a ser escopado por empresa via manager."""
    today = timezone.localdate()
    open_lots = list(Lot.objects.filter(status=Lot.STATUS_OPEN))
    # F8: valor aproximado (ao vivo) do lote-herói — SÓ admin (mesmo gate do preço).
    if getattr(request, 'company_role', None) == 'admin' and open_lots:
        vals = _lot_valuations(request, open_lots[0])
        open_lots[0].val_mid = vals[0]['total_mid'] if vals else None
        open_lots[0].val_mid_rmb = vals[0].get('total_mid_rmb') if vals else None
    ctx = {
        'open_lots': open_lots,
        'stats': {
            'open_count':     len(open_lots),
            'types_today':    InventoryEntry.objects.filter(added_at__date=today).count(),
            'pending_count':  PendingEntry.objects.count(),
            'rejected_today': RejectedEntry.objects.filter(created_at__date=today).count(),
        },
    }
    return render(request, ui(request, 'estoque/painel.html'), ctx)


# ─── lot list ───────────────────────────────────────────────────────────────

@role_required('operator')
def lot_list(request):
    # Todos os lotes (da empresa — T3 escopa via manager; hoje há uma empresa).
    # Antes filtrava por operator=request.user; ver docstring de _get_lot.
    lots = list(Lot.objects.all())

    # F8 (PRECIFICACAO §7): valoração por lote — SÓ admin da empresa (mesmo gate
    # do lot_detail). Fechado = CONGELADO (LotPricing.total_mid, 1 query em bloco);
    # aberto = ao vivo (poucos lotes abertos). Anexa lot.val_mid (Decimal|None).
    is_admin = getattr(request, 'company_role', None) == 'admin'
    if is_admin:
        from pricing.models import LotPricing
        frozen = {}
        for lp in (LotPricing.objects.filter(lot__in=lots)
                   .order_by('lot_id', '-created_at')):
            frozen.setdefault(lp.lot_id, lp.total_mid)   # mais recente por lote
        for lot in lots:
            if lot.status == Lot.STATUS_CLOSED:
                lot.val_mid = frozen.get(lot.pk)
            else:
                vals = _lot_valuations(request, lot)
                lot.val_mid = vals[0]['total_mid'] if vals else None
                lot.val_mid_rmb = vals[0].get('total_mid_rmb') if vals else None

    return render(request, ui(request, 'estoque/lotes.html'),
                  {'lots': lots, 'show_valuation': is_admin})


# ─── lot create ─────────────────────────────────────────────────────────────

@role_required('manager')   # §8: abrir lote é de gerente+
@require_POST
def lot_create(request):
    description = request.POST.get('description', '').strip()
    # Origem OBRIGATÓRIA e sem default (dono, 2026-08-01): a promessa
    # comercial do lote — celular × PCB — nasce na abertura.
    origin = (request.POST.get('origin') or '').strip()
    if origin not in (Lot.ORIGIN_PHONE, Lot.ORIGIN_PCB):
        messages.error(request, _('Escolha a ORIGEM do lote (celular ou '
                                  'PCB) — obrigatória, sem padrão.'))
        return redirect('estoque:index')
    # T2: numeração atômica por empresa (lock no contador da Company) — elimina
    # a corrida do antigo Max+1. request.company existe: o gate exige Membership.
    # T3: o lote nasce com a empresa (e a filial do gerente, se houver — §9).
    lot = Lot.open_for_company(request.company, request.user, description,
                               branch=request.membership.branch, origin=origin)
    return redirect('estoque:lot_detail', lot_pk=lot.pk)


# ─── lot detail ─────────────────────────────────────────────────────────────

@role_required('operator')
def lot_detail(request, lot_pk):
    lot  = _get_lot(request, lot_pk)
    q    = request.GET.get('q', '').strip()
    tipo = request.GET.get('tipo', '').strip()

    entries   = _entries_for_display(lot, q, tipo)
    total_qty = sum(e.quantity for e in entries)          # do lote/filtro INTEIRO
    entries, page_obj = _paginate_entries(request, entries)   # F11.0b

    # F12: empresa-CLIENTE vê a tabela mascarada (PN + C-### + qtd).
    from tenancy.access import is_unmasked
    masked = not is_unmasked(request)
    if masked:
        _masked_entry_labels(entries)

    from pricing.engine import fx_display
    ctx = {
        'lot':       lot,
        'entries':   entries,
        'page_obj':  page_obj,
        'fx_info':   fx_display(),   # PLANO_FX C: taxa é dado PÚBLICO (modal/selo)
        'masked':    masked,
        'total_qty': total_qty,
        'q':         q,
        'tipo':      tipo,
        # Marcas SUPORTADAS = as que têm família (classificam algo), direto do banco —
        # antes era lista hardcoded de 7 (com 'Toshiba' e sem Nanya/Kingston/Rayson).
        # Agora auto-atualiza: as 10 marcas dos yamls, Toshiba-Kioxia unificada, sem fantasmas.
        'brands': list(Brand.objects.filter(families__isnull=False)
                       .order_by('name').values_list('name', flat=True).distinct()),
    }

    if request.headers.get('HX-Request'):
        return render(request, ui(request, 'estoque/partials/table_body.html'), ctx)

    # F8 (PRECIFICACAO §7): valoração do lote — SÓ admin, só no render completo.
    # Lote FECHADO mostra o CONGELADO (auditoria "vendi com qual tabela");
    # lote aberto calcula on-read da tabela viva (re-classifica cada PN).
    if getattr(request, 'company_role', None) == 'admin':
        ctx['valuations'] = _lot_valuations(request, lot)
        # AUTO-REFRESH: o wrapper vivo renderiza pro admin MESMO com a lista
        # vazia (sem comprador ainda) — o est:added/60s preenche depois.
        ctx['show_live_valuation'] = True
    # F11.2c: smart button (padrão Odoo) — a venda ativa deste lote. Segue o
    # gate do MENU (gerente para cima, dono 2026-08-14): a valoração acima é
    # que continua admin — o botão só leva à OV, que já se mascara sozinha.
    if can_sales(request):
        from vendas.models import STATUS_CANCELLED, SalesOrder
        ctx['sales_order'] = (SalesOrder.all_companies.filter(lot=lot)
                              .exclude(status=STATUS_CANCELLED).first())

    return render(request, ui(request, 'estoque/estoque.html'), ctx)


def _lot_valuations(request, lot):
    """[{buyer, frozen, totals…}] para o painel de valoração (admin-only)."""
    from pricing.engine import price_lot_multi
    from pricing.models import Buyer, LotPricing

    out = []
    if lot.status == Lot.STATUS_CLOSED:
        for lp in LotPricing.objects.filter(lot=lot).select_related('buyer'):
            out.append({
                'buyer': lp.buyer, 'frozen': True, 'created_at': lp.created_at,
                'total_low': lp.total_low, 'total_mid': lp.total_mid,
                'total_high': lp.total_high,
                'priced_units': lp.priced_units, 'total_units': lp.total_units,
                'coverage_units': lp.coverage_units,
            })
        if out:
            return out
    # F11.0: classify 1× por PN, compartilhado entre os compradores.
    for buyer, rep in price_lot_multi(lot, Buyer.objects.filter(active=True)):
        out.append({
            'buyer': buyer, 'frozen': False, 'created_at': None,
            'total_low': rep.totals['low'], 'total_mid': rep.totals['mid'],
            'total_high': rep.totals['high'],
            # PLANO_FX (Fase A): o ¥ é o valor PRIMÁRIO; o US$ vira "≈".
            'total_mid_rmb': rep.totals_rmb['mid'],
            'total_low_rmb': rep.totals_rmb['low'],
            'total_high_rmb': rep.totals_rmb['high'],
            'priced_units': rep.priced_units, 'total_units': rep.total_units,
            'coverage_units': rep.coverage_units,
        })
    return out


# ─── auto-refresh (PLANO_FX — dono 2026-08-01: "sem F5") ────────────────────

@role_required('operator')
def fx_badge(request):
    """Parcial do widget da taxa (header) — alvo do polling HTMX de 60s.
    Taxa é dado público de mercado: qualquer papel logado."""
    from pricing.engine import fx_display
    return render(request, ui(request, 'estoque/partials/fx_badge.html'),
                  {'wtc_fx': fx_display() or {'rate_disp': None}})


@role_required('operator')
def lot_valuation_live(request, lot_pk):
    """Parcial da valoração VIVA do lote — alvo do refresh em est:added e do
    polling de 60s. Mesmo gate do lot_detail: valores SÓ para admin (os
    demais recebem parcial vazio — a view nunca vaza).
    ``?fmt=card`` devolve o formato COMPACTO dos cards da listagem/painel
    (¥ N ≈ US$ M) em vez das linhas 💰 da página do lote."""
    lot = _get_lot(request, lot_pk)
    valuations = (_lot_valuations(request, lot)
                  if getattr(request, 'company_role', None) == 'admin' else [])
    if request.GET.get('fmt') == 'card':
        v = valuations[0] if valuations else None
        return render(request, ui(request, 'estoque/partials/lot_value_card.html'),
                      {'v': v})
    return render(request, ui(request, 'estoque/partials/lot_valuation.html'),
                  {'valuations': valuations})


# ─── lot close / reopen ──────────────────────────────────────────────────────

def _freeze_lot_pricing(request, lot):
    """F8 (PRECIFICACAO §1.7): congela a valoração no FECHAMENTO — o registro
    'vendi com qual tabela'. Um LotPricing por comprador ativo; reabrir+fechar
    cria outro (append). ⚠ Falha de preço NUNCA trava o fechamento do lote
    (operação da bancada > auditoria): loga e segue."""
    try:
        from pricing.engine import price_lot_multi
        from pricing.models import Buyer, LotPricing
        # F11.0: classify 1× por PN, compartilhado entre os compradores.
        for buyer, rep in price_lot_multi(lot, Buyer.objects.filter(active=True)):
            LotPricing.all_companies.create(
                lot=lot, buyer=buyer, closed_by=request.user,
                total_low=rep.totals['low'], total_mid=rep.totals['mid'],
                total_high=rep.totals['high'],
                priced_units=rep.priced_units, total_units=rep.total_units,
                priced_lines=rep.priced_lines, total_lines=rep.total_lines,
                lines=[{
                    'pn': l.part_number, 'qty': l.quantity,
                    'status': l.quote.status,
                    'min': str(l.quote.price_min) if l.quote.price_min is not None else None,
                    'max': str(l.quote.price_max) if l.quote.price_max is not None else None,
                    'reason': l.quote.reason, 'via': l.quote.via,
                } for l in rep.lines])
    except Exception:
        logger.exception('F8: falha ao congelar valoração do lote %s', lot.pk)


@role_required('manager')   # §8: fechar lote é de gerente+
@require_POST
def lot_close(request, lot_pk):
    lot = _get_lot(request, lot_pk)
    # F11.2 (dono, 2026-07-16): fechar exige digitar o CÓDIGO COMPLETO do
    # lote (type-to-confirm) — a barreira é AQUI, o prompt do template é UX.
    if (request.POST.get('confirm_code') or '').strip() != lot.code:
        messages.error(request, _(
            'Código de confirmação não confere — digite o código completo '
            'do lote (ex.: %(code)s). O lote NÃO foi fechado.')
            % {'code': lot.code})
        return redirect('estoque:lot_detail', lot_pk=lot.pk)
    # ── TRAVA DE CÂMBIO (PLANO_FX Fase C, 2026-08-01): o acordo com o
    # comprador é a taxa de MERCADO do dia do fechamento — capturada AQUI,
    # atômica com o CLOSED. Sem taxa no sistema, o fechamento físico NUNCA
    # bloqueia (campos ficam nulos + aviso; o comercial resolve depois). ──
    from pricing.engine import current_fx_rate
    _rate, _fx = current_fx_rate()
    lot.status    = Lot.STATUS_CLOSED
    lot.closed_at = timezone.now()
    if _rate is not None:
        lot.fx_rate      = _rate
        lot.fx_source    = (_fx.source if _fx else 'bootstrap contratual')
        lot.fx_locked_at = timezone.now()
        lot.fx_is_fallback = bool(_fx and _fx.is_fallback)
    else:
        messages.warning(request, _(
            'Lote fechado SEM taxa de câmbio no sistema — rode o '
            'fetch_fx_rate e trave manualmente com o suporte.'))
    lot.save(update_fields=['status', 'closed_at', 'fx_rate', 'fx_source',
                            'fx_locked_at', 'fx_is_fallback'])
    # O snapshot é criado no servidor mesmo quando quem fecha é o GERENTE —
    # ele não VÊ valores (§7); o registro é para o admin/auditoria.
    _freeze_lot_pricing(request, lot)
    # F11.2 (§12.19): fechamento gera a COTAÇÃO draft no menu Vendas (valores
    # vivos até o admin confirmar). Nunca trava o fechamento (padrão F8).
    from vendas.services import create_draft_for_lot
    so = create_draft_for_lot(lot, request.user)
    # F11.2c (dono): ADMIN cai direto na venda recém-criada; o gerente segue
    # no lote — desde 2026-08-14 ele TEM acesso a /vendas/ (mascarado), mas o
    # fluxo dele termina no lote; chega na OV pelo smart button ou pelo menu.
    if so is not None and getattr(request, 'company_role', None) == 'admin':
        return redirect('vendas:so_detail', pk=so.pk)
    return redirect('estoque:lot_detail', lot_pk=lot.pk)


@role_required('manager')   # o papel segue sendo o piso; o teto é o gate abaixo
@require_POST
def lot_reopen(request, lot_pk):
    lot = _get_lot(request, lot_pk)
    # ── FECHOU, TÁ FECHADO (dono, 2026-07-31/PLANO_FX §1.3): a reabertura
    # saiu do produto — só o SUPERUSER (plataforma) reverte, auditado pelo
    # pghistory. Correção comum pós-fechamento = Acerto (padrão Odoo). ──
    if not request.user.is_superuser:
        messages.error(request, _(
            'Lote fechado é definitivo — correções entram como ACERTO na '
            'venda. Reabertura é exclusiva da plataforma.'))
        return redirect('estoque:lot_detail', lot_pk=lot.pk)
    # F11.2 (padrão Odoo, dono 2026-07-16): OV CONFIRMADA bloqueia a
    # reabertura (cancele a ordem no menu Vendas antes — auditável);
    # cotação DRAFT é cancelada automaticamente ao reabrir.
    from vendas.models import STATUS_CONFIRMED, STATUS_DRAFT, SalesOrder
    from vendas.services import cancel as cancel_so
    if SalesOrder.all_companies.filter(lot=lot,
                                       status=STATUS_CONFIRMED).exists():
        messages.error(request, _(
            'Este lote tem uma Ordem de Venda CONFIRMADA — cancele a ordem '
            'no menu Vendas antes de reabrir.'))
        return redirect('estoque:lot_detail', lot_pk=lot.pk)
    for so in SalesOrder.all_companies.filter(lot=lot, status=STATUS_DRAFT):
        cancel_so(so, request.user)
    # Reabriu → o câmbio DESTRAVA (volta ao vivo); o re-fechamento captura
    # taxa NOVA. As duas travas ficam no histórico (pghistory/LotEvent).
    lot.status    = Lot.STATUS_OPEN
    lot.closed_at = None
    lot.fx_rate = None
    lot.fx_source = ''
    lot.fx_locked_at = None
    lot.fx_is_fallback = False
    lot.save(update_fields=['status', 'closed_at', 'fx_rate', 'fx_source',
                            'fx_locked_at', 'fx_is_fallback'])
    return redirect('estoque:lot_detail', lot_pk=lot.pk)


@role_required('manager')   # §8: excluir é de gerente+ (admin/superuser inclusos; operador NÃO)
@require_POST
def lot_delete(request, lot_pk):
    """Exclui o lote DEFINITIVAMENTE (dono 2026-08-05). Piso de papel = gerente
    (``role_required('manager')``): admin e o superuser-da-plataforma (que navega
    com Membership real) alcançam; o OPERADOR não. A barreira REAL é aqui — o
    botão escondido no template nunca é a única trava (§8).

    Confirmação type-to-confirm (digitar o código completo do lote), igual ao
    fechamento; o modal é só UX. Cascata FK: ``InventoryEntry``/``PendingEntry``/
    ``RejectedEntry`` e a valoração congelada (``LotPricing``) somem junto. Já a
    ``vendas.SalesOrder`` referencia o lote com ``on_delete=PROTECT`` — lote com
    venda vinculada NÃO se apaga (evita o 500 e protege o histórico financeiro):
    o comercial cancela a ordem no menu Vendas antes."""
    lot = _get_lot(request, lot_pk)
    if (request.POST.get('confirm_code') or '').strip() != lot.code:
        messages.error(request, _(
            'Código de confirmação não confere — digite o código completo '
            'do lote (ex.: %(code)s). O lote NÃO foi excluído.')
            % {'code': lot.code})
        return redirect('estoque:lot_detail', lot_pk=lot.pk)
    code = lot.code
    try:
        lot.delete()
    except ProtectedError:
        messages.error(request, _(
            'Não é possível excluir %(code)s: há uma venda vinculada a este '
            'lote. Cancele a ordem no menu Vendas e tente de novo.')
            % {'code': code})
        return redirect('estoque:lot_detail', lot_pk=lot.pk)
    messages.success(request, _(
        'Lote %(code)s excluído definitivamente.') % {'code': code})
    return redirect('estoque:index')


# ─── preview chip ────────────────────────────────────────────────────────────

@role_required('operator')
def preview_chip(request, lot_pk):
    lot = _get_lot(request, lot_pk)
    pn  = _normalise_pn(request.GET.get('pn', ''))

    if _pn_too_short(pn):
        return HttpResponse('')

    result  = classify(pn)
    # A camada "não encontrado / em fila de revisão" do classify devolve um dict REDUZIDO
    # (só flags + campos numéricos: known/in_review_queue/nand_gb/… — SEM as chaves de spec
    # string). Os templates confirm_card* acessam result.capacity/.interface/etc. como
    # ARGUMENTO de filtro (ex.: `default:result.capacity`), cuja resolução é ESTRITA:
    # VariableDoesNotExist derruba o preview (500) quando a chave falta — bug exposto ao
    # digitar o prefixo de um PN em fila de revisão (2026-08-05). Garante as chaves-padrão
    # (fill-only: não altera um resultado já completo, só preenche o reduzido com '').
    for _k in ('chip_type', 'subtype', 'capacity', 'dram_density', 'interface', 'brand',
               'emcp_ram', 'emcp_nand', 'classification_source', 'device', 'fbga_code'):
        result.setdefault(_k, '')
    has_cap = _has_capacity(result)

    destination, dest_cat = _compute_destination(result)

    # display_cap = detalhe da sub-linha. Para DRAM, a densidade já está no rótulo
    # da caixa (2G/4G/8G); não mostrar a capacidade em bytes (256MB), que confunde
    # o operador — a caixa de RAM é por densidade.
    if result.get('is_emcp'):
        parts = [p for p in [result.get('emcp_nand', ''), result.get('emcp_ram', '')] if p]
        display_cap = ' / '.join(parts)
    elif dest_cat == 'lpddr':
        # Geração + densidade já estão no rótulo da caixa (ex.: D3+2G). A sub-linha
        # fica só com a marca — sem capacidade em bytes (256MB), que confunde.
        display_cap = ''
    else:
        display_cap = result.get('capacity') or result.get('dram_density') or ''

    try:
        current_qty = InventoryEntry.objects.get(lot=lot, part_number=pn).quantity
    except InventoryEntry.DoesNotExist:
        current_qty = 0

    # Gateway de triagem (3 etapas + typo). Substitui o cálculo solto de
    # profitable/prof_key — a regra de destino agora mora num lugar só.
    gateway = _compute_gateway(result, has_cap)

    ctx = {
        'lot':             lot,
        'pn':              pn,
        'result':          result,
        'has_cap':         has_cap,
        'display_cap':     display_cap,
        'result_json':     json.dumps({**result, 'pn': pn}),
        'current_qty':     current_qty,
        'destination':     destination,
        'destination_cat': dest_cat,
        'gateway':         gateway,
        'gateway_dest':    gateway['destination'],
        'gateway_steps':   gateway['steps'],
        'profitable':      gateway['profitable'],
        'profitable_key':  gateway['profitable_key'],
        # Idempotência (2026-08-10): 1 token por CARD — o add_chip só aplica a
        # escrita para o 1º POST deste token; reenvio (duplo clique, re-clique
        # em rede lenta, retry pós-queda) vira no-op. Ver _claim_submit_token.
        'submit_token':    uuid4().hex,
    }

    # F8 (PRECIFICACAO §7/§12): preço do comprador na bancada — SÓ admin.
    # quotes_for_admin devolve [] para operador/gerente sem nem consultar preço.
    from pricing.engine import fx_display, quotes_for_admin
    ctx['price_quotes'] = quotes_for_admin(request, result, origin=lot.origin)
    ctx['fx_info'] = fx_display() if ctx['price_quotes'] else None

    # F12 (máscara de categoria, dono 2026-07-17): empresa-CLIENTE recebe o
    # card MASCARADO (template whitelist: PN + destino C-###/baldes + qtd +
    # preço p/ admin — sem specs, sem veredito nominal, sem data-debug).
    # Plataforma (is_unmasked) segue no card completo.
    from tenancy.access import is_unmasked
    if not is_unmasked(request):
        masked_code, masked_general = _masked_category(result)
        ctx.update({'masked_code': masked_code,
                    'masked_general': masked_general})
        return render(request,
                      ui(request, 'estoque/partials/confirm_card_masked.html'), ctx)

    return render(request, ui(request, 'estoque/partials/confirm_card.html'), ctx)


# ─── add chip ────────────────────────────────────────────────────────────────

@role_required('operator')   # §8: adicionar a lote ABERTO é o trabalho do operador
@require_POST
def add_chip(request, lot_pk):
    lot = _get_lot(request, lot_pk)

    if not lot.is_open:
        return HttpResponse(
            '<div class="est-msg est-msg--error" style="padding:12px 16px;border:1px solid #da1e28;color:#da1e28;margin-top:12px;">'
            + _('Este lote está fechado. Reabra-o para adicionar chips.')
            + '</div>'
        )

    pn  = _normalise_pn(request.POST.get('pn', ''))
    qty = max(1, int(request.POST.get('qty') or 1))

    if _pn_too_short(pn):
        return HttpResponse(
            '<div class="est-msg est-msg--error" style="padding:12px 16px;">' + _('PN inválido.') + '</div>'
        )

    # Reclassifica no servidor (não confia no hidden do form).
    server_result = classify(pn)
    confirmed = _is_confirmed(server_result)

    # ── Idempotência (bug Mundo Metal, 2026-08-10) ───────────────────────────
    # Reivindica o token DESTE clique ANTES de qualquer escrita. Duplicata
    # (is_dup=True) percorre os MESMOS ramos abaixo, mas só re-renderiza o
    # estado atual — sem somar estoque/fila e sem duplicar linha de auditoria.
    is_dup = _claim_submit_token(request)

    # ── Atalho: morto por GERAÇÃO (não confirmado) → descarte direto ─────────
    # Tecnologia velha (LPDDR2-, DDR2-, MCP legado) é sucata por fato de mercado.
    # Reprovado mesmo SEM confirmação e SEM capacidade mapeada — por isso vem
    # ANTES do ramo has_cap (senão um chip sem capacidade cairia em "desconhecido")
    # e da fila. Razão distinta ("geração") para a auditoria separar do reprovado
    # confirmado. Confirmados seguem o fluxo normal abaixo.
    if is_dead_by_generation(server_result) and not confirmed:
        if not is_dup:
            RejectedEntry.objects.create(
                lot=lot, part_number=pn, quantity=qty,
                **_snapshot(server_result),
                # ⚠ CANÔNICO — persistido p/ auditoria. NUNCA traduzir (i18n só na
                # exibição; dado gravado fica em pt-br — I18N.md §8.2).
                rejection_reason='NÃO RENTÁVEL (geração)',
                operator=request.user,
            )
        return render(request, ui(request, 'estoque/partials/rejected_feedback.html'), {
            'pn': pn, 'qty': qty,
            'chip_type': server_result.get('chip_type', ''),
            'capacity':  server_result.get('capacity', ''),
            'by_generation': True,
        })

    # has_cap recomputado NO SERVIDOR (regra de ouro #3): a identificação é decisão
    # de negócio (manda o chip para UnknownChip) e não pode confiar no hidden do
    # form. classify() é determinístico (lru_cache), então casa com o do preview;
    # um POST forjado não burla mais esta etapa.
    has_cap = _has_capacity(server_result)

    if not has_cap:
        # company: anotação §14.1 (fila global; 1ª empresa a reportar).
        # (get_or_create já é idempotente; o guard só poupa a query na duplicata.)
        if not is_dup:
            UnknownChip.objects.get_or_create(
                part_number=pn, defaults={'company': request.company})
        return render(request, ui(request, 'estoque/partials/unknown_feedback.html'), {'pn': pn})

    # ── Bloqueio "só confirmados" ────────────────────────────────────────────
    # Se o PN não é confirmado no banco, NÃO entra no estoque: vai para a fila de
    # conferência (PendingEntry) para o gestor aprovar/reprovar.
    if not confirmed:
        near = _nearest_in_lot(lot, pn)
        if is_dup:
            # Reenvio: fila NÃO soma de novo — mostra a quantidade atual.
            pend = PendingEntry.objects.filter(lot=lot, part_number=pn).first()
            pend_qty = pend.quantity if pend else qty
        else:
            pend, p_created = PendingEntry.objects.get_or_create(
                lot=lot, part_number=pn,
                defaults={
                    'quantity':          qty,
                    **_snapshot(server_result),
                    'nearest_confirmed': near,
                    'operator':          request.user,
                },
            )
            if not p_created:
                PendingEntry.objects.filter(pk=pend.pk).update(quantity=F('quantity') + qty)
                pend.refresh_from_db()
            pend_qty = pend.quantity
        return render(request, ui(request, 'estoque/partials/pending_feedback.html'), {
            'pn': pn, 'qty': pend_qty, 'near': near,
        })

    # ── Bloqueio DURO de rentabilidade ───────────────────────────────────────
    # Chip confirmado e com specs, mas NÃO RENTÁVEL: não entra no estoque. É
    # desviado para RejectedEntry (log de auditoria) e segue para resíduo
    # eletrônico. Decisão de negócio: bloqueio real no servidor, não só na UI.
    profitable = assess_profitability(server_result)
    if profitable == 'NÃO RENTÁVEL':
        if not is_dup:
            RejectedEntry.objects.create(
                lot=lot, part_number=pn, quantity=qty,
                **_snapshot(server_result),
                # ⚠ CANÔNICO — persistido p/ auditoria. NUNCA traduzir (I18N.md §8.2).
                rejection_reason='NÃO RENTÁVEL',
                operator=request.user,
            )
        return render(request, ui(request, 'estoque/partials/rejected_feedback.html'), {
            'pn': pn, 'qty': qty,
            'chip_type': server_result.get('chip_type', ''),
            'capacity':  server_result.get('capacity', ''),
            'by_generation': False,
        })

    # ── Sem rentabilidade AVALIADA → não lança (dono, 2026-07-31) ────────────
    # O card já desabilita o botão (gateway.can_add=False); aqui é a barreira
    # REAL contra POST forjado/página velha (template nunca é a única barreira).
    # Nada é gravado: o chip fica na bancada até o dado/regra completar.
    if profitable != 'RENTÁVEL':
        return HttpResponse(
            '<div class="est-msg est-msg--error" style="padding:12px 16px;border:1px solid #da1e28;color:#da1e28;margin-top:12px;">'
            + _('Sem avaliação de rentabilidade — este chip não pode entrar no estoque até o dado ficar completo. Sinalize ao gestor.')
            + '</div>'
        )

    # Grava SEMPRE a partir do classify do SERVIDOR (server_result), não do POST
    # do cliente — fonte autoritativa, à prova de form forjado/defasado, e idêntica
    # ao que PendingEntry/RejectedEntry já fazem (linhas acima). _snapshot captura a
    # densidade DRAM em `capacity` (antes perdida → 'None') e limpa a geração do
    # `interface`. `confidence` não existe no InventoryEntry (só em Pending/Rejected).
    if is_dup:
        # Reenvio do MESMO clique (token já usado): não soma — só re-renderiza
        # o estado atual. É exatamente a resposta que o operador perdeu quando
        # a conexão caiu; o total exibido já contém o lançamento original.
        entry = InventoryEntry.objects.filter(lot=lot, part_number=pn).first()
    else:
        snap = _snapshot(server_result)
        snap.pop('confidence', None)
        # Passo 2: carimba a edição do catálogo do snapshot de intake (detecção de defasagem).
        # F11.1: a chave de preço nasce junto (o classify já rodou — custo zero).
        from chips.models import CatalogVersion
        defaults = {**snap, **_price_key_fields(server_result), 'quantity': qty,
                    'snapshot_catalog_version': CatalogVersion.current()}

        entry, created = InventoryEntry.objects.get_or_create(
            lot=lot, part_number=pn, defaults=defaults,
        )

        if not created:
            InventoryEntry.objects.filter(pk=entry.pk).update(
                quantity=F('quantity') + qty,
                last_updated=timezone.now(),
            )

    entries   = _entries_for_display(lot)
    total_qty = sum(e.quantity for e in entries)
    entries, page_obj = _paginate_entries(request, entries)   # F11.0b: pág. 1
    from tenancy.access import is_unmasked
    masked = not is_unmasked(request)                         # F12
    if masked:
        _masked_entry_labels(entries)

    response = render(request, ui(request, 'estoque/partials/table_body.html'), {
        'lot':        lot,
        'entries':    entries,
        'page_obj':   page_obj,
        'masked':     masked,
        'total_qty':  total_qty,
        'just_added': pn,
    })
    # Duplicata REPLICA a resposta original (toast incluso — é a confirmação
    # que o operador perdeu). `entry` só é None numa corrida rara (duplicata
    # chega antes do commit do original): aí não há pk p/ o undo do toast.
    if entry is not None:
        response['HX-Trigger'] = json.dumps({'est:added': {'pn': pn, 'qty': qty, 'pk': entry.pk}})
    return response


# ─── remove entry ────────────────────────────────────────────────────────────

@role_required('operator')
@require_POST
def remove_entry(request, lot_pk, pk):
    lot   = _get_lot(request, lot_pk)
    # Operador só corrige lançamento em lote ABERTO (parte do fluxo de adicionar).
    # Mexer em lote fechado é correção de gestão → gerente+ (comportamento de
    # hoje preservado para o gerente; antes qualquer um removia de lote fechado).
    if not lot.is_open and not request.membership.has_role('manager'):
        raise PermissionDenied('Lote fechado: remoção é ação de gerente.')
    # InventoryEntry.objects EXPLÍCITO: escopo por empresa (o default é cru).
    entry = get_object_or_404(InventoryEntry.objects, pk=pk, lot=lot)
    qty   = max(1, int(request.POST.get('qty') or 1))

    if qty >= entry.quantity:
        entry.delete()
    else:
        InventoryEntry.objects.filter(pk=entry.pk).update(quantity=F('quantity') - qty)

    entries   = _entries_for_display(lot)
    total_qty = sum(e.quantity for e in entries)
    entries, page_obj = _paginate_entries(request, entries)   # F11.0b: pág. 1
    from tenancy.access import is_unmasked
    masked = not is_unmasked(request)                         # F12
    if masked:
        _masked_entry_labels(entries)

    return render(request, ui(request, 'estoque/partials/table_body.html'), {
        'lot':       lot,
        'entries':   entries,
        'page_obj':  page_obj,
        'masked':    masked,
        'total_qty': total_qty,
    })


# ─── export xls ──────────────────────────────────────────────────────────────

# Estados sem valor no export (mesmo vocabulário canônico do PriceQuote).
_EXPORT_PRICE_LABEL = {
    'NO_BUY':   'não compra',
    'NOT_MADE': 'não fabricado',
    'UNQUOTED': 'sem cotação',
}


def _export_price_maps(request, lot):
    """[(nome_do_comprador, {pn: linha})] para as colunas de preço do export.

    SÓ papel ADMIN da empresa (matriz §8 do PLANO_MULTITENANT: gerente exporta a
    planilha SEM as colunas de preço — mesma regra do card e da valoração).
    Lote FECHADO usa o CONGELADO da F8 (LotPricing.lines — "vendi com qual
    tabela"); aberto (ou fechado sem congelado) precifica ON-READ da tabela viva.
    """
    if getattr(request, 'company_role', None) != 'admin':
        return []
    from pricing.engine import price_lot_multi
    from pricing.models import Buyer, LotPricing

    # 1ª passada: quem tem CONGELADO usa o congelado; os demais entram na
    # precificação viva EM GRUPO (F11.0: classify 1× por PN pra todos).
    frozen_by, live_buyers = {}, []
    buyers = list(Buyer.objects.filter(active=True).order_by('name'))
    for buyer in buyers:
        frozen = None
        if lot.status == Lot.STATUS_CLOSED:
            frozen = (LotPricing.objects.filter(lot=lot, buyer=buyer)
                      .order_by('-created_at').first())
        if frozen:
            frozen_by[buyer.pk] = {l.get('pn'): l for l in frozen.lines}
        else:
            live_buyers.append(buyer)

    live_by = {}
    if live_buyers:
        for buyer, rep in price_lot_multi(lot, live_buyers):
            live_by[buyer.pk] = {
                l.part_number: {
                    'status': l.quote.status,
                    'min': str(l.quote.price_min) if l.quote.price_min is not None else None,
                    'max': str(l.quote.price_max) if l.quote.price_max is not None else None,
                    # PLANO_FX (Fase A): o ¥ é o unitário PRIMÁRIO do export.
                    'rmb': str(l.quote.value_rmb()) if l.quote.status == 'PRICED' else None,
                }
                for l in rep.lines
            }

    # F11.3 (sigilo): para a empresa a contraparte é o WhatTheChip —
    # nenhuma identidade de comprador no header da planilha.
    return [('WhatTheChip',
             frozen_by.get(buyer.pk) or live_by.get(buyer.pk, {}))
            for buyer in buyers]


def _export_price_cells(line, qty):
    """(¥ unitário, US$ unitário, US$ total) p/ a planilha (PLANO_FX Fase A:
    ¥ primeiro). PRICED → números (cenário default da config já embutido no
    min/max congelado do report); sem valor → rótulo curto no ¥ e resto
    vazio. Congelados LEGADOS (sem 'rmb') mostram '—' no ¥."""
    from decimal import Decimal
    if line and line.get('status') == 'PRICED' and line.get('min') is not None:
        low  = Decimal(str(line['min']))
        high = Decimal(str(line.get('max') or line['min']))
        unit = (low + high) / 2
        rmb = float(Decimal(str(line['rmb']))) if line.get('rmb') else '—'
        return (rmb, float(round(unit, 2)),
                float(round(unit * (qty or 0), 2)))
    status = line.get('status') if line else None
    return _EXPORT_PRICE_LABEL.get(status, 'sem preço'), None, None


@role_required('manager')   # §8: exportar lote é de gerente+
def export_xls(request, lot_pk):
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError:
        return HttpResponse('openpyxl não instalado.', status=500)

    lot     = _get_lot(request, lot_pk)
    entries_list = list(_entries_qs(lot))
    # Colunas de preço: vazio para gerente (só admin vê preço — matriz §8).
    price_maps = _export_price_maps(request, lot)

    wb = openpyxl.Workbook()
    ws = wb.active
    # Origem no título da aba (2026-08-01): 'PHONE'/'PCB' — o rótulo que o
    # comprador confere no recebimento (canônico, nunca traduz).
    ws.title = f'Lote {lot.number:03d} {lot.origin.upper()}'

    header_fill  = PatternFill('solid', fgColor='0F62FE')
    header_font  = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_align = Alignment(horizontal='left', vertical='center')
    cell_border  = Border(bottom=Side(style='thin', color='E0E0E0'))
    mono_font    = Font(name='Courier New', size=10)

    # F12: export mascarado p/ empresa-cliente — specs viram o código C-###.
    from tenancy.access import is_unmasked
    masked = not is_unmasked(request)
    if masked:
        _masked_entry_labels(entries_list)
        headers    = ['Part Number', 'Category', 'Qty.', 'Last Added']
        col_widths = [22, 12, 8, 20]
    else:
        headers    = ['Part Number', 'Brand', 'Type', 'Capacity', 'Interface', 'Qty.', 'Source', 'Last Added']
        col_widths = [22, 16, 12, 20, 16, 8, 18, 20]
    base_n  = len(headers)
    qty_col = 3 if masked else 6
    # Colunas de preço por comprador (só chegam aqui para admin). "Unit." é o
    # ponto médio da faixa em USD (cenário default); fechado = congelado (F8).
    # PLANO_FX (Fase A): ¥ primeiro — o unitário em ¥ ganha coluna própria;
    # o USD segue como tradução (≈ mercado no aberto; congelado no fechado).
    for buyer_name, _lines in price_maps:
        headers    += [f'Preço unit. — {buyer_name} (¥ RMB)',
                       f'Preço unit. — {buyer_name} (US$ ≈)',
                       f'Total — {buyer_name} (US$ ≈)']
        col_widths += [24, 24, 20]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = w

    ws.row_dimensions[1].height = 28

    buyer_totals = [0.0] * len(price_maps)   # soma da coluna Total por comprador

    for row_idx, entry in enumerate(entries_list, start=2):
        _quando = (timezone.localtime(entry.last_updated)
                   .strftime('%d/%m/%Y %H:%M:%S') if entry.last_updated else '—')
        if masked:
            data = [entry.part_number, entry.category_label,
                    entry.quantity, _quando]
        else:
            data = [
                entry.part_number,
                entry.brand or '—',
                entry.chip_type or '—',
                entry.display_capacity,
                entry.interface or '—',
                entry.quantity,
                entry.classification_source or '—',
                _quando,
            ]
        for i, (_buyer_name, lines) in enumerate(price_maps):
            rmb_unit, unit, line_total = _export_price_cells(
                lines.get(entry.part_number), entry.quantity)
            data += [rmb_unit, unit, line_total]
            if line_total is not None:
                buyer_totals[i] += line_total
        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border    = cell_border
            cell.alignment = Alignment(vertical='center')
            if col_idx == 1:
                cell.font = mono_font
            if col_idx > base_n and isinstance(value, float):
                cell.number_format = '#,##0.00'
        ws.row_dimensions[row_idx].height = 20

    total_row  = len(entries_list) + 2
    total_font = Font(name='Calibri', bold=True, size=10)
    ws.cell(row=total_row, column=1, value='TOTAL').font = total_font
    ws.cell(row=total_row, column=qty_col, value=sum(e.quantity for e in entries_list)).font = total_font
    # Total geral em USD por comprador (na coluna "Total" dele).
    for i, total in enumerate(buyer_totals):
        # PLANO_FX: 3 colunas por comprador (¥ unit · US$ unit · US$ total)
        cell = ws.cell(row=total_row, column=base_n + 3 + i * 3, value=round(total, 2))
        cell.font = total_font
        cell.number_format = '#,##0.00'

    wb.properties.creator = 'WhatTheChip?'
    wb.properties.title   = f'Lote #{lot.number:03d} — {request.user.username}'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f'lote_{lot.number:03d}_{timezone.localtime(timezone.now()).strftime("%Y%m%d_%H%M")}.xlsx'
    response = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
