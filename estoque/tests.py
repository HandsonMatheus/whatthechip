"""
Testes do gateway de triagem do estoque (Fase 1).

Rodar sempre com o settings de teste (SQLite em memória, Gemini off):

    python manage.py test estoque --settings=core.settings_test

Dois blocos:
  - GatewayDestinationTests: _compute_gateway como função pura (dicts sintéticos),
    sem banco — prova a ordem identificação → fonte → rentabilidade e a regra
    conservadora (INDETERMINADO → aprovado).
  - AddChipHardBlockTests: integra a view add_chip com classify() mockado, para
    isolar a lógica de roteamento (estoque / fila / reprovado / desconhecido) do
    engine de classificação.
"""

import threading
from datetime import timedelta
from unittest import skipUnless
from unittest.mock import patch
from uuid import uuid4

from django.contrib.auth import get_user_model
from django.db import connection
from django.test import TestCase, TransactionTestCase
from django.urls import reverse
from django.utils import timezone

from chips.engine import is_dead_by_generation
from chips.models import UnknownChip
from tenancy.models import Company, Membership
from tenancy.scope import (CompanyScopedManager, CompanyScopeMissing,
                           company_scope, set_current_company)

from .models import (InventoryEntry, Lot, PendingEntry, RejectedEntry,
                     SubmitToken)
from .views import _compute_gateway


def _grant(user, role=Membership.ROLE_OPERATOR):
    """Vincula o usuário de teste a uma empresa com papel (T1: as views do
    estoque exigem Membership ativo — tenancy.access.role_required)."""
    # v3.1 (dono 2026-07-23): is_platform NÃO desmascara mais (só superuser).
    # A eMiner dos testes segue is_platform (campo mantido p/ outros usos), mas
    # os testes que precisam da VISÃO COMPLETA marcam o usuário is_superuser.
    company, _ = Company.objects.get_or_create(
        name='eMiner', defaults={'slug': 'eminer', 'is_platform': True})
    Membership.objects.update_or_create(
        user=user, company=company, defaults={'role': role, 'active': True})
    return company


def _scope(testcase, company):
    """Escopo AMBIENTE de empresa para o corpo do teste (T3: os managers do
    estoque são fail-closed — asserts diretos no ORM precisam de escopo, como
    um comando precisaria). Cleanup força None (comandos chamados no meio do
    teste também setam o contextvar; o reset por token não se aplica)."""
    set_current_company(getattr(company, 'pk', company))
    testcase.addCleanup(set_current_company, None)


def _result(**over):
    """Dict de classificação mínimo, sobrescrevível por teste."""
    base = {
        'chip_type': '', 'subtype': '', 'brand': '',
        'capacity': '', 'emcp_ram': '', 'emcp_nand': '', 'is_emcp': False,
        'interface': '', 'dram_density': '',
        'classification_source': '', 'confidence': '', 'fuzzy_suggestions': [],
    }
    base.update(over)
    return base


class GatewayDestinationTests(TestCase):
    """_compute_gateway: destino correto + estados das 3 etapas.

    TestCase (não SimpleTestCase): assess_profitability / is_dead_by_generation
    leem ProfitabilityConfig do banco (get_config cria o singleton default)."""

    def test_confirmado_rentavel_aprovado(self):
        r = _result(chip_type='eMMC', capacity='16GB',
                    classification_source='banco de dados', confidence='confirmed')
        g = _compute_gateway(r, has_cap=True)
        self.assertEqual(g['destination'], 'aprovado')
        self.assertEqual(g['profitable'], 'RENTÁVEL')
        self.assertEqual(g['profitable_key'], 'rentavel')
        self.assertEqual([s['status'] for s in g['steps']], ['pass', 'pass', 'pass'])

    def test_confirmado_nao_rentavel_reprovado(self):
        r = _result(chip_type='eMMC', capacity='2GB',  # < 4GB (default) → NÃO RENTÁVEL
                    classification_source='banco de dados', confidence='confirmed')
        g = _compute_gateway(r, has_cap=True)
        self.assertEqual(g['destination'], 'reprovado')
        self.assertEqual(g['profitable'], 'NÃO RENTÁVEL')
        self.assertEqual([s['status'] for s in g['steps']], ['pass', 'pass', 'fail'])

    def test_confirmado_indeterminado_aprovado(self):
        # Tipo de catálogo sem regra de rentabilidade → INDETERMINADO → aprovado
        # (regra conservadora). Antes usava chip_type='NOR'; agora 'NOR' é
        # reconhecido como NOR Flash = sucata (dead) pela fonte única
        # (chips/chip_types.py), então usamos 'SoC', que é genuinamente INDETERMINADO.
        r = _result(chip_type='SoC', capacity='8GB', confidence='manual')
        g = _compute_gateway(r, has_cap=True)
        self.assertEqual(g['destination'], 'aprovado')       # não é descarte…
        self.assertEqual(g['profitable'], 'INDETERMINADO')
        self.assertEqual(g['profitable_key'], 'indeterminado')
        # …mas NÃO lança (dono 2026-07-31): botão desabilitado no card.
        self.assertFalse(g['can_add'])
        # F1b: o 3º passo NÃO é 'pass' (verde "sim") — é 'warn' (âmbar "indeterminado").
        # Antes mostrava "Rentável: sim" mentiroso num chip não avaliado.
        self.assertEqual([s['status'] for s in g['steps']], ['pass', 'pass', 'warn'])

    def test_nao_confirmado_vai_para_fila(self):
        # Mesmo sendo eMMC 16GB (rentável), a fonte falha antes da rentabilidade.
        r = _result(chip_type='eMMC', capacity='16GB',
                    classification_source='gramática', confidence='estimated')
        g = _compute_gateway(r, has_cap=True)
        self.assertEqual(g['destination'], 'fila')
        self.assertEqual([s['status'] for s in g['steps']], ['pass', 'fail', 'skip'])
        self.assertEqual(g['profitable'], '')  # rentabilidade nem avaliada

    def test_distribuidor_entra_no_estoque(self):
        # Decisão do dono (2026-07-08): registro de DISTRIBUIDOR é elegível pro
        # estoque (tem registro no banco), mesmo classificando pela gramática. As
        # specs seguem vindo da gramática (distribuidor não vence o engine).
        r = _result(chip_type='eMMC', capacity='16GB',
                    classification_source='gramática', confidence='distributor')
        g = _compute_gateway(r, has_cap=True)
        self.assertEqual(g['destination'], 'aprovado')
        self.assertEqual([s['status'] for s in g['steps']], ['pass', 'pass', 'pass'])
        self.assertTrue(g['can_add'])                        # rentável → lança

    def test_gramatica_pura_sem_registro_vai_para_fila(self):
        # "GRAMÁTICA JAMAIS": sem registro no banco (confidence vazio), mesmo com
        # specs decodificadas pela gramática, NÃO entra no estoque → fila.
        r = _result(chip_type='eMMC', capacity='16GB',
                    classification_source='gramática', confidence='')
        g = _compute_gateway(r, has_cap=True)
        self.assertEqual(g['destination'], 'fila')
        self.assertEqual([s['status'] for s in g['steps']], ['pass', 'fail', 'skip'])

    def test_sem_specs_desconhecido(self):
        g = _compute_gateway(_result(), has_cap=False)
        self.assertEqual(g['destination'], 'desconhecido')
        self.assertEqual([s['status'] for s in g['steps']], ['fail', 'skip', 'skip'])

    def test_typo_e_paralelo_ao_destino(self):
        # Sugestões fuzzy aparecem independentemente do destino (rede de segurança).
        r = _result(fuzzy_suggestions=['KLMAG1JETD', 'KLMAG1JET4'])
        g = _compute_gateway(r, has_cap=False)
        self.assertEqual(g['destination'], 'desconhecido')
        self.assertTrue(g['typo']['has'])
        self.assertEqual(len(g['typo']['suggestions']), 2)

    def test_sem_typo(self):
        g = _compute_gateway(_result(chip_type='eMMC', capacity='16GB',
                                     confidence='confirmed'), has_cap=True)
        self.assertFalse(g['typo']['has'])

    # ── Morto por geração (DERIVADO da rentabilidade) ───────────────────────
    def test_is_dead_by_generation(self):
        # Geração morta — independe da capacidade (eMCP exige RAM e NAND, como no real).
        self.assertTrue(is_dead_by_generation(_result(
            chip_type='eMCP', is_emcp=True,
            emcp_ram='LPDDR2 ⚠ cap. não mapeada', emcp_nand='eMMC ⚠ cap. não mapeada')))
        self.assertTrue(is_dead_by_generation(_result(chip_type='DDR', subtype='DDR2')))
        self.assertTrue(is_dead_by_generation(_result(chip_type='MCP')))
        # Geração viva
        self.assertFalse(is_dead_by_generation(_result(
            chip_type='eMCP', is_emcp=True, emcp_ram='LPDDR4 2GB', emcp_nand='eMMC 16GB')))
        # Não-rentável por CAPACIDADE NÃO conta como morto por geração
        self.assertFalse(is_dead_by_generation(_result(chip_type='eMMC', capacity='2GB')))
        self.assertFalse(is_dead_by_generation(_result(
            chip_type='eMCP', is_emcp=True, emcp_ram='LPDDR3 0.5GB', emcp_nand='eMMC 16GB')))

    def test_gen_dead_nao_confirmado_vai_reprovado(self):
        # KMN5W000ZM-like: eMCP LPDDR2 por gramática, capacidade não mapeada (has_cap
        # False). Antes caía em DESCONHECIDO; agora vai direto pra REPROVADO por geração.
        r = _result(chip_type='eMCP', is_emcp=True, subtype='LPDDR2 + eMMC',
                    emcp_ram='LPDDR2 ⚠ cap. não mapeada', emcp_nand='eMMC ⚠ cap. não mapeada',
                    classification_source='gramática', confidence='estimated')
        g = _compute_gateway(r, has_cap=False)
        self.assertEqual(g['destination'], 'reprovado')
        self.assertTrue(g['reject_by_generation'])
        self.assertEqual([s['status'] for s in g['steps']], ['pass', 'fail', 'fail'])

    def test_gen_dead_confirmado_segue_fluxo_normal(self):
        # eMCP LPDDR2 CONFIRMADO → reprovado normal (sem rótulo de geração).
        r = _result(chip_type='eMCP', is_emcp=True, subtype='LPDDR2 + eMMC',
                    emcp_ram='LPDDR2 1GB', emcp_nand='eMMC 8GB',
                    classification_source='banco de dados', confidence='confirmed')
        g = _compute_gateway(r, has_cap=True)
        self.assertEqual(g['destination'], 'reprovado')
        self.assertFalse(g['reject_by_generation'])
        self.assertEqual([s['status'] for s in g['steps']], ['pass', 'pass', 'fail'])

    # ── Caixa física: DRAM = geração(subtype, LITERAL)+densidade; NAND por GB ──
    def test_caixa_dram_geracao_mais_densidade(self):
        from estoque.views import _compute_destination
        # Geração vem do subtype, literal — MENOS o ruído "SDRAM" (toda DDR é SDRAM).
        # A variante (DDR3/DDR3L/LPDDR3) é preservada. interface pode ser config de
        # barramento ("x16 @ 800MHz") → NÃO vira rótulo. Densidade do dram_density (Gb).
        # Categoria CSS (split jun/2026): DDR/DDR3L → 'ddr' (marrom); LPDDR* → 'lpddr'
        # (azul). O rótulo (DDR3L+2G) é o mesmo; só a cor da caixa muda.
        self.assertEqual(_compute_destination(_result(
            chip_type='RAM', subtype='DDR3L SDRAM', interface='x16 @ 800MHz',
            capacity='256MB', dram_density='2Gb = por die [✓]')), ('DDR3L+2G', 'ddr'))
        # LPDDR móvel = pacote MULTI-DIE → caixa pela CAPACIDADE do pacote em GB
        # (sufixo "GB", convenção CLAUDE.md §6), não pela densidade do die. H9CC 4GB
        # (4 dies de 8Gb) → "LPDDR3+4GB", NUNCA "LPDDR3+32G" (bug do fallback
        # bytes→Gbit assumindo 1 die).
        self.assertEqual(_compute_destination(_result(
            chip_type='RAM', subtype='LPDDR3', interface='LPDDR3',
            capacity='4GB', dram_density=None)), ('LPDDR3+4GB', 'lpddr'))
        # P/ LPDDR a capacidade VENCE a densidade do die (prova que usa capacity):
        # 6GB de pacote, dram_density de 8Gb por die → "LPDDR4+6GB" (6, não 8).
        self.assertEqual(_compute_destination(_result(
            chip_type='LPDDR4', subtype='LPDDR4',
            capacity='6GB', dram_density='8Gb = 1GB por die [✓]')), ('LPDDR4+6GB', 'lpddr'))
        # SK Hynix H5TQ confirmado: dram_density VAZIO + subtype "DDR3 SDRAM".
        # Strip de SDRAM + fallback bytes→Gb (256MB→2G) → 'DDR3+2G'.
        self.assertEqual(_compute_destination(_result(
            chip_type='RAM', subtype='DDR3 SDRAM', interface='DDR3',
            capacity='256MB', dram_density=None)), ('DDR3+2G', 'ddr'))
        # 1GB por chip → 8G (teto físico DDR3)
        self.assertEqual(_compute_destination(_result(
            chip_type='RAM', subtype='DDR3 SDRAM',
            capacity='1GB', dram_density=None)), ('DDR3+8G', 'ddr'))
        # NAND continua por capacidade em GB (não afetado pelo fallback de densidade)
        self.assertEqual(_compute_destination(_result(
            chip_type='eMMC', capacity='16GB')), ('EMMC16GB', 'emmc'))


class AddChipHardBlockTests(TestCase):
    """add_chip: roteamento estoque / fila / reprovado / desconhecido."""

    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user(username='op', password='x')
        self.company = _grant(self.user)       # T1: operador precisa de vínculo
        _scope(self, self.company)             # T3: asserts diretos no ORM
        self.lot = Lot.objects.create(number=0, origin='phone', operator=self.user,
                                      company=self.company)
        self.client.login(username='op', password='x')
        self.url = reverse('estoque:add', args=[self.lot.pk])

    @patch('estoque.views.classify')
    def test_confirmado_nao_rentavel_vira_rejected(self, mock_classify):
        mock_classify.return_value = _result(
            chip_type='eMMC', capacity='2GB',  # < 4GB → NÃO RENTÁVEL por capacidade
            classification_source='banco de dados', confidence='confirmed')
        self.client.post(self.url, {'pn': 'TESTREJECT01', 'qty': '2', 'has_cap': 'true'})
        self.assertEqual(
            RejectedEntry.objects.filter(lot=self.lot, part_number='TESTREJECT01').count(), 1)
        self.assertFalse(
            InventoryEntry.objects.filter(lot=self.lot, part_number='TESTREJECT01').exists())

    @patch('estoque.views.classify')
    def test_reprovado_nao_emcp_com_none_nao_quebra(self, mock_classify):
        # Chip não-eMCP (LPDDR2): classify devolve emcp_ram/emcp_nand = None.
        # Antes do fix, o insert no RejectedEntry dava NotNullViolation (500).
        mock_classify.return_value = _result(
            chip_type='LPDDR2', capacity='1GB', interface='LPDDR2',
            emcp_ram=None, emcp_nand=None,
            classification_source='banco de dados', confidence='confirmed')
        resp = self.client.post(self.url, {'pn': 'K4P8G304EQ', 'qty': '1', 'has_cap': 'true'})
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(
            RejectedEntry.objects.filter(lot=self.lot, part_number='K4P8G304EQ').count(), 1)

    @patch('estoque.views.classify')
    def test_confirmado_rentavel_entra_no_estoque(self, mock_classify):
        mock_classify.return_value = _result(
            chip_type='eMMC', capacity='16GB',
            classification_source='banco de dados', confidence='confirmed')
        self.client.post(self.url, {
            'pn': 'TESTOK16', 'qty': '1', 'has_cap': 'true',
            'chip_type': 'eMMC', 'capacity': '16GB',
            'classification_source': 'banco de dados'})
        self.assertTrue(
            InventoryEntry.objects.filter(lot=self.lot, part_number='TESTOK16').exists())
        self.assertFalse(RejectedEntry.objects.filter(part_number='TESTOK16').exists())

    @patch('estoque.views.classify')
    def test_confirmado_indeterminado_nao_lanca(self, mock_classify):
        # Dono 2026-07-31: sem rentabilidade AVALIADA não entra no estoque —
        # o card desabilita o botão e ESTA é a barreira real (POST forjado).
        # Nada é gravado (nem Rejected, nem Pending — o chip fica na bancada).
        mock_classify.return_value = _result(
            chip_type='SoC', capacity='8GB',
            classification_source='banco de dados', confidence='confirmed')
        resp = self.client.post(self.url, {'pn': 'TESTINDET01', 'qty': '1',
                                           'has_cap': 'true'})
        self.assertContains(resp, 'Sem avaliação de rentabilidade')
        self.assertFalse(InventoryEntry.objects.filter(
            part_number='TESTINDET01').exists())
        self.assertFalse(RejectedEntry.objects.filter(
            part_number='TESTINDET01').exists())
        self.assertFalse(PendingEntry.objects.filter(
            part_number='TESTINDET01').exists())

    @patch('estoque.views.classify')
    def test_intake_carimba_snapshot_catalog_version(self, mock_classify):
        """Passo 2: a entrada no estoque carimba a edição ATUAL do catálogo (data de atualização)."""
        from chips.models import CatalogVersion
        mock_classify.return_value = _result(
            chip_type='eMMC', capacity='16GB',
            classification_source='banco de dados', confidence='confirmed')
        self.client.post(self.url, {
            'pn': 'TESTVER01', 'qty': '1', 'has_cap': 'true',
            'chip_type': 'eMMC', 'capacity': '16GB',
            'classification_source': 'banco de dados'})
        entry = InventoryEntry.objects.get(lot=self.lot, part_number='TESTVER01')
        self.assertEqual(entry.snapshot_catalog_version, CatalogVersion.current())
        self.assertGreaterEqual(entry.snapshot_catalog_version, 1)

    @patch('estoque.views.classify')
    def test_intake_grava_chave_de_preco(self, mock_classify):
        """F11.1: o lançamento materializa a CHAVE DE PREÇO (kind/gen/tier) —
        a valoração resolve por join na tabela viva, sem reclassificar."""
        from decimal import Decimal
        mock_classify.return_value = _result(
            chip_type='eMMC', capacity='16GB', cap_gb=16.0,
            classification_source='banco de dados', confidence='confirmed')
        self.client.post(self.url, {'pn': 'TESTKEY16', 'qty': '1', 'has_cap': 'true'})
        e = InventoryEntry.objects.get(lot=self.lot, part_number='TESTKEY16')
        self.assertEqual((e.price_kind, e.price_gen, e.price_tier_unit),
                         ('emmc', '', 'GB'))
        self.assertEqual(e.price_tier_value, Decimal('16'))
        self.assertEqual(e.price_key_reason, '')

    def test_sem_chave_produz_o_motivo(self):
        """F11.1: chip fora do mercado de preço carrega o MOTIVO do NO_KEY
        nos campos de chave. Era um teste de add_chip com SoC INDETERMINADO —
        mas desde a regra do dono (2026-07-31) indeterminado NÃO entra mais
        no estoque, então o caminho vivo desse campo é o LEGADO (resnapshot);
        aqui provamos a função que o produz."""
        from estoque.views import _price_key_fields
        campos = _price_key_fields(_result(chip_type='SoC', capacity='8GB',
                                           confidence='manual'))
        self.assertIsNone(campos.get('price_tier_value'))
        self.assertIn('fora do mercado', campos['price_key_reason'])

    @patch('estoque.views.classify')
    def test_nao_confirmado_vai_para_pending(self, mock_classify):
        mock_classify.return_value = _result(
            chip_type='eMMC', capacity='16GB',
            classification_source='gramática', confidence='estimated')
        self.client.post(self.url, {'pn': 'TESTGRAM01', 'qty': '1', 'has_cap': 'true'})
        self.assertTrue(
            PendingEntry.objects.filter(lot=self.lot, part_number='TESTGRAM01').exists())
        self.assertFalse(InventoryEntry.objects.filter(part_number='TESTGRAM01').exists())
        self.assertFalse(RejectedEntry.objects.filter(part_number='TESTGRAM01').exists())

    def test_has_cap_false_vai_para_unknown(self):
        # has_cap=false e chip não-reconhecido → UnknownChip (classify roda, mas
        # não é morto por geração nem confirmado).
        self.client.post(self.url, {'pn': 'TESTUNK01', 'qty': '1', 'has_cap': 'false'})
        self.assertTrue(UnknownChip.objects.filter(part_number='TESTUNK01').exists())

    @patch('estoque.views.classify')
    def test_gen_dead_nao_confirmado_vira_rejected_geracao(self, mock_classify):
        # eMCP LPDDR2 por gramática, capacidade não mapeada, has_cap=false:
        # antes ia pra UnknownChip; agora vai direto pra reprovado POR GERAÇÃO.
        mock_classify.return_value = _result(
            chip_type='eMCP', is_emcp=True, subtype='LPDDR2 + eMMC',
            emcp_ram='LPDDR2 ⚠ cap. não mapeada', emcp_nand='eMMC ⚠ cap. não mapeada',
            classification_source='gramática', confidence='estimated')
        self.client.post(self.url, {'pn': 'KMN5W000ZM', 'qty': '1', 'has_cap': 'false'})
        rej = RejectedEntry.objects.filter(lot=self.lot, part_number='KMN5W000ZM')
        self.assertEqual(rej.count(), 1)
        self.assertEqual(rej.first().rejection_reason, 'NÃO RENTÁVEL (geração)')
        self.assertFalse(UnknownChip.objects.filter(part_number='KMN5W000ZM').exists())

    @patch('estoque.views.classify')
    def test_distribuidor_entra_com_source_banco_de_dados(self, mock_classify):
        """Lote 41 (2026-07-13): registro DISTRIBUIDOR cujas specs a gramática
        completou (classification_source='gramática') É elegível (decisão 2026-07-08)
        e, por TER registro no banco, entra com o rótulo 'banco de dados' — nunca
        'gramática' (que parecia chip sem registro e disparava falso alarme)."""
        mock_classify.return_value = _result(
            chip_type='eMMC', capacity='16GB',
            classification_source='gramática', confidence='distributor')
        self.client.post(self.url, {'pn': 'THGBMBG7D4KBAIW', 'qty': '1', 'has_cap': 'true'})
        e = InventoryEntry.objects.get(lot=self.lot, part_number='THGBMBG7D4KBAIW')
        self.assertEqual(e.classification_source, 'banco de dados')


class AddChipIdempotencyTests(TestCase):
    """Idempotência do add_chip (bug Mundo Metal LOT/002/08/26, 2026-08-10).

    POST duplicado — duplo clique, re-clique em rede lenta (Venezuela), retry
    após queda de conexão — NÃO pode somar duas vezes. O card embute um
    ``submit_token`` por render; o 1º POST o reivindica (``SubmitToken``,
    unique no banco) e escreve; o reenvio do MESMO token re-renderiza o estado
    atual sem escrever nada (estoque, fila E log de descarte)."""

    RENTAVEL = dict(chip_type='eMMC', capacity='16GB',
                    classification_source='banco de dados',
                    confidence='confirmed')

    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user(username='op_idem', password='x')
        self.company = _grant(self.user)
        _scope(self, self.company)
        self.lot = Lot.objects.create(number=0, origin='phone',
                                      operator=self.user, company=self.company)
        self.client.login(username='op_idem', password='x')
        self.url = reverse('estoque:add', args=[self.lot.pk])

    def _post(self, pn, qty, token):
        data = {'pn': pn, 'qty': str(qty), 'has_cap': 'true'}
        if token is not None:
            data['submit_token'] = token
        return self.client.post(self.url, data)

    @patch('estoque.views.classify')
    def test_token_repetido_nao_soma_no_estoque(self, mock_classify):
        # O caso do operador: lança 15, a resposta demora, clica de novo.
        mock_classify.return_value = _result(**self.RENTAVEL)
        tok = uuid4().hex
        r1 = self._post('TESTDUP01', 15, tok)
        r2 = self._post('TESTDUP01', 15, tok)
        self.assertEqual(r1.status_code, 200)
        self.assertEqual(r2.status_code, 200)
        e = InventoryEntry.objects.get(lot=self.lot, part_number='TESTDUP01')
        self.assertEqual(e.quantity, 15)          # e NÃO 30
        # A duplicata REPLICA a resposta original (tabela + HX-Trigger/toast) —
        # é a confirmação que o operador perdeu quando a conexão caiu.
        self.assertIn('HX-Trigger', r2)
        self.assertEqual(SubmitToken.objects.filter(token=tok).count(), 1)

    @patch('estoque.views.classify')
    def test_tokens_diferentes_somam(self, mock_classify):
        # 2 cards distintos = 2 cliques legítimos: tem que somar como sempre.
        mock_classify.return_value = _result(**self.RENTAVEL)
        self._post('TESTDUP02', 10, uuid4().hex)
        self._post('TESTDUP02', 5, uuid4().hex)
        e = InventoryEntry.objects.get(lot=self.lot, part_number='TESTDUP02')
        self.assertEqual(e.quantity, 15)

    @patch('estoque.views.classify')
    def test_sem_token_mantem_comportamento_antigo(self, mock_classify):
        # Aba aberta ANTES do deploy (form sem o hidden): sem dedupe — cada
        # POST soma, como sempre. Janela curta e aceita de propósito
        # (exigir o token quebraria a bancada de quem estava logado na virada).
        mock_classify.return_value = _result(**self.RENTAVEL)
        self._post('TESTDUP03', 7, None)
        self._post('TESTDUP03', 7, None)
        e = InventoryEntry.objects.get(lot=self.lot, part_number='TESTDUP03')
        self.assertEqual(e.quantity, 14)

    @patch('estoque.views.classify')
    def test_token_repetido_fila_nao_soma(self, mock_classify):
        # A fila de conferência sofria do MESMO bug (F('quantity') + qty).
        mock_classify.return_value = _result(
            chip_type='eMMC', capacity='16GB',
            classification_source='gramática', confidence='estimated')
        tok = uuid4().hex
        self._post('TESTDUPFILA', 8, tok)
        r2 = self._post('TESTDUPFILA', 8, tok)
        self.assertEqual(r2.status_code, 200)
        p = PendingEntry.objects.get(lot=self.lot, part_number='TESTDUPFILA')
        self.assertEqual(p.quantity, 8)           # e NÃO 16

    @patch('estoque.views.classify')
    def test_token_repetido_descarte_loga_uma_vez(self, mock_classify):
        # RejectedEntry é append-only (1 linha POR TENTATIVA REAL — sinal de
        # calibração): reenvio do mesmo clique não pode virar 2 linhas.
        mock_classify.return_value = _result(
            chip_type='eMMC', capacity='2GB',     # < 4GB → NÃO RENTÁVEL
            classification_source='banco de dados', confidence='confirmed')
        tok = uuid4().hex
        self._post('TESTDUPREJ', 3, tok)
        self._post('TESTDUPREJ', 3, tok)
        self.assertEqual(RejectedEntry.objects.filter(
            lot=self.lot, part_number='TESTDUPREJ').count(), 1)

    @patch('estoque.views.classify')
    def test_token_invalido_e_ignorado(self, mock_classify):
        # Token fora do formato uuid4().hex não deduplica nem explode (e não
        # deixa lixo na tabela).
        mock_classify.return_value = _result(**self.RENTAVEL)
        self._post('TESTDUP04', 2, 'nao-e-token')
        self._post('TESTDUP04', 2, 'nao-e-token')
        e = InventoryEntry.objects.get(lot=self.lot, part_number='TESTDUP04')
        self.assertEqual(e.quantity, 4)
        self.assertEqual(SubmitToken.objects.count(), 0)

    @patch('estoque.views.classify')
    def test_poda_tokens_velhos(self, mock_classify):
        # Poda lazy: um lançamento novo apaga tokens com mais de 48h.
        mock_classify.return_value = _result(**self.RENTAVEL)
        velho = SubmitToken.objects.create(token=uuid4().hex)
        SubmitToken.objects.filter(pk=velho.pk).update(
            created_at=timezone.now() - timedelta(hours=72))
        self._post('TESTDUP05', 1, uuid4().hex)
        self.assertFalse(SubmitToken.objects.filter(pk=velho.pk).exists())

    def test_preview_embute_token_e_travas(self):
        # O card renderizado (mascarado — usuário comum, v3.1) embute o hidden
        # submit_token e as travas anti-duplo-clique (classify REAL, sem mock:
        # qualquer destino renderiza o form com a proteção).
        url = reverse('estoque:preview', args=[self.lot.pk])
        resp = self.client.get(url, {'pn': 'ZZZZ9999XX'})
        self.assertContains(resp, 'name="submit_token"')
        self.assertContains(resp, 'hx-sync="this:drop"')
        self.assertContains(resp, 'hx-disabled-elt')


class DisplaySourceTests(TestCase):
    """_display_source: rótulo 'Source' do ESTOQUE = 'banco de dados' para tudo que
    tem registro no banco (confirmed/manual/distributor) — inclusive distribuidor com
    specs por gramática. Gramática PURA (sem registro) preserva 'gramática'. O motor
    e o site NÃO mudam. (Diagnóstico do lote 41, 2026-07-13.)"""

    def test_distribuidor_com_specs_por_gramatica_vira_banco(self):
        from estoque.views import _display_source
        self.assertEqual(
            _display_source(_result(classification_source='gramática',
                                    confidence='distributor')),
            'banco de dados')

    def test_confirmado_e_banco(self):
        from estoque.views import _display_source
        self.assertEqual(
            _display_source(_result(classification_source='banco de dados',
                                    confidence='confirmed')),
            'banco de dados')

    def test_confirmado_sem_familia_source_vazio_vira_banco(self):
        # Micron JZ###: known_exact True, source vazio → 'banco de dados' (não apaga).
        from estoque.views import _display_source
        self.assertEqual(
            _display_source(_result(classification_source='', confidence='confirmed',
                                    known_exact=True)),
            'banco de dados')

    def test_gramatica_pura_preserva_rotulo(self):
        # Sem registro (estimated) → NÃO elegível → mantém 'gramática'.
        from estoque.views import _display_source
        self.assertEqual(
            _display_source(_result(classification_source='gramática',
                                    confidence='estimated')),
            'gramática')


class RefreshLoteRelabelTests(TestCase):
    """refresh_lote reescreve o Source das entradas JÁ gravadas ao vivo: distribuidor
    com specs por gramática ('gramática' antigo) → 'banco de dados' (fix do lote 41)."""

    def setUp(self):
        self.user = get_user_model().objects.create_user(username='op41', password='x')
        self.company = _grant(self.user)
        _scope(self, self.company)
        self.lot = Lot.objects.create(number=41, origin='phone', operator=self.user,
                                      company=self.company)

    @patch('estoque.management.commands.refresh_lote.classify')
    def test_relabela_distribuidor_gramatica_para_banco(self, mock_classify):
        from django.core.management import call_command
        e = InventoryEntry.objects.create(
            lot=self.lot, part_number='THGBMBG7D4KBAIW', chip_type='eMMC',
            capacity='16GB', classification_source='gramática')
        mock_classify.return_value = _result(
            chip_type='eMMC', capacity='16GB',
            classification_source='gramática', confidence='distributor')
        call_command('refresh_lote', '--lot', '41', '--commit')
        e.refresh_from_db()
        self.assertEqual(e.classification_source, 'banco de dados')


class PreviewFuzzyPopupTests(TestCase):
    """O popup de sugestões (fuzzy) renderiza no card e ABRE SOZINHO quando o chip
    cai em fila/desconhecido — as operadoras ignoravam o inline e mandavam typo pra
    fila (pedido 2026-07-13). Em aprovado NÃO auto-abre (não incomoda decode válido)."""

    def setUp(self):
        # v3.1 (dono 2026-07-23): o popup fuzzy vive no card COMPLETO (não
        # mascarado). Só o superuser vê o card completo → o usuário de
        # visão-plena aqui é superuser (não afrouxa a máscara; ver is_unmasked).
        self.user = get_user_model().objects.create_user(
            username='op55', password='x', is_superuser=True)
        self.company = _grant(self.user)
        _scope(self, self.company)
        self.client.force_login(self.user)
        self.lot = Lot.objects.create(number=55, origin='phone', operator=self.user,
                                      company=self.company)

    @patch('estoque.views.classify')
    def test_popup_abre_sozinho_em_fila(self, mock_classify):
        # eMMC 16GB por gramática (não confirmado) → fila; com sugestão fuzzy.
        mock_classify.return_value = _result(
            chip_type='eMMC', capacity='16GB',
            classification_source='gramática', confidence='estimated',
            fuzzy_suggestions=['KLMAG1JETD'])
        url = reverse('estoque:preview', args=[self.lot.pk])
        body = self.client.get(url, {'pn': 'KLMAG1JET0'}).content.decode()
        self.assertIn('fuzzy-modal-overlay', body)          # popup presente
        self.assertIn("ov.style.display='flex'", body)      # abre sozinho (fila)
        self.assertIn('KLMAG1JETD', body)                   # sugestão aparece
        # trava: comentário de template NÃO pode vazar na página (bug do {# #}
        # multi-linha, corrigido 2026-07-13 — {# #} do Django é só de UMA linha).
        self.assertNotIn('força a conferência', body)

    @patch('estoque.views.classify')
    def test_popup_nao_auto_abre_em_aprovado(self, mock_classify):
        # Confirmado + rentável → aprovado; mesmo com sugestão, popup NÃO auto-abre.
        mock_classify.return_value = _result(
            chip_type='eMMC', capacity='16GB',
            classification_source='banco de dados', confidence='confirmed',
            fuzzy_suggestions=['KLMAG1JETD'])
        url = reverse('estoque:preview', args=[self.lot.pk])
        body = self.client.get(url, {'pn': 'KLMAG1JETD'}).content.decode()
        self.assertIn('fuzzy-modal-overlay', body)          # markup existe
        self.assertNotIn("ov.style.display='flex'", body)   # mas NÃO auto-abre


class ResnapshotLoteTests(TestCase):
    """Passo 2: o resnapshot_lote revalua as entradas DEFASADAS (catálogo melhorou)."""

    def setUp(self):
        self.user = get_user_model().objects.create_user(username='op2', password='x')
        self.company = _grant(self.user)
        _scope(self, self.company)
        self.lot = Lot.objects.create(number=77, origin='phone', operator=self.user,
                                      company=self.company)

    @patch('chips.engine.classify')
    def test_revalua_entrada_defasada(self, mock_classify):
        from chips.models import CatalogVersion
        from django.core.management import call_command
        v0 = CatalogVersion.current()
        e = InventoryEntry.objects.create(
            lot=self.lot, part_number='MTSTALE1', chip_type='RAM', capacity='48GB',
            snapshot_catalog_version=v0)
        CatalogVersion.bump()                       # catálogo melhora → versão sobe
        cur = CatalogVersion.current()
        self.assertGreater(cur, v0)
        mock_classify.return_value = {
            'chip_type': 'LPDDR4', 'capacity': '6GB',
            'classification_source': 'banco de dados'}
        call_command('resnapshot_lote', '--lot', '77', '--commit')
        e.refresh_from_db()
        self.assertEqual(e.snapshot_catalog_version, cur)   # saiu da defasagem
        self.assertEqual(e.capacity, '6GB')                 # 48GB → 6GB
        self.assertEqual(e.chip_type, 'LPDDR4')

    @patch('chips.engine.classify')
    def test_nao_apaga_source_de_confirmado_sem_familia(self, mock_classify):
        """JZ###: confirmado SEM família casada → classify não devolve Source, mas o
        resnapshot deriva 'banco de dados' (não apaga o rótulo)."""
        from chips.models import CatalogVersion
        from django.core.management import call_command
        v0 = CatalogVersion.current()
        e = InventoryEntry.objects.create(
            lot=self.lot, part_number='JZ109', chip_type='eMMC',
            classification_source='banco de dados', snapshot_catalog_version=v0)
        CatalogVersion.bump()
        mock_classify.return_value = {
            'chip_type': 'eMMC', 'confidence': 'confirmed', 'classification_source': ''}
        call_command('resnapshot_lote', '--lot', '77', '--commit')
        e.refresh_from_db()
        self.assertEqual(e.classification_source, 'banco de dados')  # NÃO apagou

    @patch('chips.engine.classify')
    def test_backfill_da_chave_de_preco(self, mock_classify):
        """F11.1: entrada LEGADA (sem chave — pré-F11.1, aprovação de
        pendência, restore) ganha a chave no resnapshot (backfill)."""
        from decimal import Decimal
        from chips.models import CatalogVersion
        from django.core.management import call_command
        v0 = CatalogVersion.current()
        e = InventoryEntry.objects.create(
            lot=self.lot, part_number='LEGADO16', chip_type='eMMC',
            capacity='16GB', snapshot_catalog_version=v0)
        self.assertEqual(e.price_kind, '')                  # nasceu legada
        CatalogVersion.bump()
        cur = CatalogVersion.current()
        # Legada com o carimbo EM DIA: o backfill tem que pegar mesmo assim
        # (o filtro do resnapshot inclui "sem chave", não só "defasada").
        e2 = InventoryEntry.objects.create(
            lot=self.lot, part_number='LEGADOCUR', chip_type='eMMC',
            capacity='16GB', snapshot_catalog_version=cur)
        mock_classify.return_value = {
            'chip_type': 'eMMC', 'capacity': '16GB', 'cap_gb': 16.0,
            'classification_source': 'banco de dados'}
        call_command('resnapshot_lote', '--lot', '77', '--commit')
        e.refresh_from_db()
        e2.refresh_from_db()
        self.assertEqual((e.price_kind, e.price_tier_unit), ('emmc', 'GB'))
        self.assertEqual(e.price_tier_value, Decimal('16'))
        self.assertEqual(e2.price_kind, 'emmc')             # backfill sem bump

    @patch('chips.engine.classify')
    def test_dry_run_explicito_nao_grava(self, mock_classify):
        """`--dry-run` (explícito) NÃO grava — mesmo com uma entrada defasada."""
        from chips.models import CatalogVersion
        from django.core.management import call_command
        v0 = CatalogVersion.current()
        e = InventoryEntry.objects.create(
            lot=self.lot, part_number='MTDRY1', chip_type='RAM', capacity='48GB',
            snapshot_catalog_version=v0)
        CatalogVersion.bump()
        mock_classify.return_value = {'chip_type': 'LPDDR4', 'capacity': '6GB'}
        call_command('resnapshot_lote', '--lot', '77', '--dry-run')  # NÃO deve gravar
        e.refresh_from_db()
        self.assertEqual(e.capacity, '48GB')                  # intacto
        self.assertEqual(e.snapshot_catalog_version, v0)      # ainda defasado


class OnReadDisplayTests(TestCase):
    """Passo 2: a TELA do estoque mostra o valor ATUAL (cálculo na leitura/on-read)
    das entradas defasadas — sem gravar — e exibe a data de última atualização."""

    def setUp(self):
        # v3.1 (dono 2026-07-23): visão completa (specs on-read) = só superuser.
        self.user = get_user_model().objects.create_user(username='op3', password='x',
                                                         is_superuser=True)
        self.company = _grant(self.user)       # T1: operador precisa de vínculo
        _scope(self, self.company)             # T3
        self.client.force_login(self.user)
        self.lot = Lot.objects.create(number=88, origin='phone', operator=self.user,
                                      company=self.company)

    def _get_table(self):
        resp = self.client.get(reverse('estoque:lot_detail', args=[self.lot.pk]))
        self.assertEqual(resp.status_code, 200)
        return resp.content.decode()

    @patch('estoque.views.classify')
    def test_onread_mostra_valor_atual_de_entrada_defasada(self, mock_classify):
        """Catálogo melhorou desde o intake → a tela mostra o valor ATUAL, não o
        snapshot velho — e NÃO persiste (quem grava é o resnapshot_lote)."""
        from chips.models import CatalogVersion
        v0 = CatalogVersion.current()
        e = InventoryEntry.objects.create(
            lot=self.lot, part_number='MTSTALE9', chip_type='RAM', capacity='48GB',
            snapshot_catalog_version=v0)
        CatalogVersion.bump()                       # catálogo melhora → entrada defasa
        mock_classify.return_value = _result(
            chip_type='LPDDR4', capacity='6GB', classification_source='banco de dados')

        body = self._get_table()
        self.assertIn('6GB', body)                  # valor ATUAL na tela
        self.assertNotIn('48GB', body)              # o defasado não aparece

        e.refresh_from_db()                         # banco INTACTO (on-read não grava)
        self.assertEqual(e.capacity, '48GB')
        self.assertEqual(e.snapshot_catalog_version, v0)

    @patch('estoque.views.classify')
    def test_onread_nao_recalcula_entrada_em_dia(self, mock_classify):
        """Entrada já na versão atual do catálogo → nem chama o classify."""
        from chips.models import CatalogVersion
        InventoryEntry.objects.create(
            lot=self.lot, part_number='MTFRESH9', chip_type='eMMC', capacity='16GB',
            snapshot_catalog_version=CatalogVersion.current())
        body = self._get_table()
        mock_classify.assert_not_called()
        self.assertIn('16GB', body)

    def test_tabela_mostra_data_de_ultima_atualizacao(self):
        from chips.models import CatalogVersion
        InventoryEntry.objects.create(
            lot=self.lot, part_number='MTDATE9', chip_type='eMMC', capacity='16GB',
            snapshot_catalog_version=CatalogVersion.current())
        self.assertIn('atualizado', self._get_table())


# ═══════════════════════════════════════════════════════════════════════════
# T1 (PLANO_MULTITENANT.md) — matriz PAPEL × VIEW (prova do objetivo O1)
# ═══════════════════════════════════════════════════════════════════════════

class RoleMatrixTests(TestCase):
    """§8 do plano vira parametrização: cada view sensível × cada papel →
    status esperado. Esconder botão NUNCA é a única barreira — o gate é a view.

    operador → busca/preview/adicionar/lista/detalhe (403 no resto)
    gerente+ → abrir/fechar/reabrir/exportar
    sem vínculo (logado) → 403 em tudo do estoque
    anônimo → redirect ao login
    """

    @classmethod
    def setUpTestData(cls):
        User = get_user_model()
        cls.company = Company.objects.create(name='eMiner', slug='eminer')
        cls.users = {}
        for username, role in [('mx_op', Membership.ROLE_OPERATOR),
                               ('mx_mgr', Membership.ROLE_MANAGER),
                               ('mx_adm', Membership.ROLE_ADMIN)]:
            u = User.objects.create_user(username=username, password='x')
            Membership.objects.create(user=u, company=cls.company, role=role)
            cls.users[role] = u
        # Logado mas SEM vínculo com empresa (ex.: conta órfã) → 403 em tudo.
        cls.users['none'] = User.objects.create_user(username='mx_orfao', password='x')
        # Lote aberto por um GERENTE — o operador trabalha nele (lote é da empresa).
        cls.lot = Lot.all_companies.create(number=900, origin='phone', operator=cls.users['manager'],
                                           company=cls.company)

    def setUp(self):
        _scope(self, self.company)   # T3: asserts diretos no ORM do teste

    def _as(self, who):
        self.client.logout()
        if who is not None:
            self.client.force_login(self.users[who])

    # ── A matriz em si ───────────────────────────────────────────────────────
    def test_matriz_papel_view(self):
        lot_pk = self.lot.pk
        # (nome, método, url, dados, {papel: status esperado})
        # 200 = ok · 302 = redirect pós-ação (ok) · 403 = barrado
        matrix = [
            ('painel', 'get', reverse('painel'), {},
             {'operator': 200, 'manager': 200, 'admin': 200, 'none': 403}),
            ('lot_list', 'get', reverse('estoque:index'), {},
             {'operator': 200, 'manager': 200, 'admin': 200, 'none': 403}),
            ('lot_detail', 'get', reverse('estoque:lot_detail', args=[lot_pk]), {},
             {'operator': 200, 'manager': 200, 'admin': 200, 'none': 403}),
            ('preview', 'get', reverse('estoque:preview', args=[lot_pk]), {'pn': 'KM'},
             {'operator': 200, 'manager': 200, 'admin': 200, 'none': 403}),
            ('export', 'get', reverse('estoque:export', args=[lot_pk]), {},
             {'operator': 403, 'manager': 200, 'admin': 200, 'none': 403}),
            ('lot_close', 'post', reverse('estoque:lot_close', args=[lot_pk]), {},
             {'operator': 403, 'manager': 302, 'admin': 302, 'none': 403}),
            ('lot_reopen', 'post', reverse('estoque:lot_reopen', args=[lot_pk]), {},
             {'operator': 403, 'manager': 302, 'admin': 302, 'none': 403}),
            # excluir é gerente+ (operador barrado); sem confirm_code o POST só
            # redireciona (não apaga) — o efeito real é o LotDeleteTests.
            ('lot_delete', 'post', reverse('estoque:lot_delete', args=[lot_pk]), {},
             {'operator': 403, 'manager': 302, 'admin': 302, 'none': 403}),
            ('lot_create', 'post', reverse('estoque:lot_create'),
             {'description': 't', 'origin': 'phone'},
             {'operator': 403, 'manager': 302, 'admin': 302, 'none': 403}),
        ]
        for name, method, url, data, expected in matrix:
            for who, status in expected.items():
                with self.subTest(view=name, papel=who):
                    self._as(who)
                    resp = getattr(self.client, method)(url, data)
                    self.assertEqual(resp.status_code, status)
            # Estado do lote pode ter mudado (close/reopen) — restaura.
            Lot.objects.filter(pk=lot_pk).update(
                status=Lot.STATUS_OPEN, closed_at=None)

    def test_anonimo_redireciona_ao_login(self):
        self._as(None)
        resp = self.client.get(reverse('estoque:index'))
        self.assertEqual(resp.status_code, 302)
        self.assertIn('/login/', resp['Location'])

    @patch('estoque.views.classify')
    def test_operador_adiciona_chip_em_lote_aberto(self, mock_classify):
        """O trabalho do operador continua intacto (O2): adicionar a lote aberto
        — inclusive lote aberto pelo GERENTE (lote é ativo da empresa)."""
        mock_classify.return_value = _result(
            chip_type='eMMC', capacity='16GB',
            classification_source='banco de dados', confidence='confirmed')
        self._as('operator')
        resp = self.client.post(reverse('estoque:add', args=[self.lot.pk]),
                                {'pn': 'MATRIXOK16', 'qty': '1'})
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(InventoryEntry.objects.filter(
            lot=self.lot, part_number='MATRIXOK16').exists())

    @patch('estoque.views.classify')
    def test_operador_remove_so_de_lote_aberto(self, mock_classify):
        """Remoção: p/ operador é correção de lançamento (só lote ABERTO);
        gerente também mexe em lote fechado (comportamento de hoje preservado)."""
        mock_classify.return_value = _result(   # p/ o on-read do render da tabela
            chip_type='eMMC', capacity='16GB',
            classification_source='banco de dados', confidence='confirmed')
        entry = InventoryEntry.objects.create(
            lot=self.lot, part_number='MATRIXRM1', quantity=5)
        url = reverse('estoque:remove', args=[self.lot.pk, entry.pk])

        self._as('operator')
        self.assertEqual(self.client.post(url, {'qty': '1'}).status_code, 200)

        Lot.objects.filter(pk=self.lot.pk).update(status=Lot.STATUS_CLOSED)
        self.assertEqual(self.client.post(url, {'qty': '1'}).status_code, 403)

        self._as('manager')
        self.assertEqual(self.client.post(url, {'qty': '1'}).status_code, 200)
        Lot.objects.filter(pk=self.lot.pk).update(status=Lot.STATUS_OPEN)

    def test_template_esconde_acoes_de_gerente_do_operador(self):
        """Navegação por papel (§9): operador não vê 'Novo Lote' nem Fechar/
        Exportar; gerente vê. (UX — a barreira real é a matriz acima.)"""
        self._as('operator')
        body = self.client.get(reverse('estoque:index')).content.decode()
        self.assertNotIn('Novo lote', body)
        detail = self.client.get(
            reverse('estoque:lot_detail', args=[self.lot.pk])).content.decode()
        self.assertNotIn('Fechar Lote', detail)
        self.assertNotIn('Exportar', detail)

        self._as('manager')
        body = self.client.get(reverse('estoque:index')).content.decode()
        self.assertIn('Novo lote', body)
        detail = self.client.get(
            reverse('estoque:lot_detail', args=[self.lot.pk])).content.decode()
        self.assertIn('Fechar Lote', detail)
        self.assertIn('Exportar', detail)

    def test_operador_ve_lote_de_outro_usuario(self):
        """Lote é ativo da EMPRESA (não do usuário): o operador enxerga e
        trabalha no lote que o gerente abriu — mudança intencional da T1
        (antes cada usuário só via os próprios lotes; ver _get_lot)."""
        self._as('operator')
        resp = self.client.get(reverse('estoque:index'))
        self.assertContains(resp, 'LOT/900/')   # nomenclatura F11.2


class LotDeleteTests(TestCase):
    """Exclusão DEFINITIVA do lote (dono 2026-08-05). O gate de PAPEL (operador
    barrado; gerente+ passa) está na RoleMatrixTests; aqui vão o EFEITO (apaga +
    cascata), o type-to-confirm (código do lote) e o require_POST."""

    def setUp(self):
        self.mgr = get_user_model().objects.create_user('del_mgr', password='x')
        self.company = _grant(self.mgr, role=Membership.ROLE_MANAGER)
        _scope(self, self.company)
        self.client.force_login(self.mgr)
        self.lot = Lot.objects.create(number=930, origin='phone',
                                      operator=self.mgr, company=self.company)
        InventoryEntry.objects.create(lot=self.lot, part_number='DELPN1',
                                      quantity=3)
        self.url = reverse('estoque:lot_delete', args=[self.lot.pk])

    def test_codigo_certo_apaga_e_redireciona_pra_lista(self):
        resp = self.client.post(self.url, {'confirm_code': self.lot.code})
        self.assertRedirects(resp, reverse('estoque:index'),
                             fetch_redirect_response=False)
        self.assertFalse(Lot.all_companies.filter(pk=self.lot.pk).exists())
        # Cascata FK: as entradas do lote somem junto.
        self.assertFalse(InventoryEntry.all_companies.filter(
            part_number='DELPN1').exists())

    def test_codigo_errado_nao_apaga(self):
        resp = self.client.post(self.url, {'confirm_code': 'ERRADO'})
        self.assertEqual(resp.status_code, 302)          # volta ao detalhe c/ erro
        self.assertTrue(Lot.all_companies.filter(pk=self.lot.pk).exists())

    def test_get_nao_apaga(self):
        # require_POST: nunca apaga por GET (crawler/prefetch/link).
        self.assertEqual(self.client.get(self.url).status_code, 405)
        self.assertTrue(Lot.all_companies.filter(pk=self.lot.pk).exists())


class LotPaginationTests(TestCase):
    """F11.0b (2026-07-16): a página do lote renderizava TODAS as entradas
    (lote 42 = ~700KB de HTML). Agora pagina em 100/página; filtros e POSTs
    resetam pra página 1 (não enviam ?p=); valoração/export seguem cobrindo
    o lote inteiro (testado em ExportPriceColumnsTests/BenchAndLot)."""

    @classmethod
    def setUpTestData(cls):
        from chips.models import CatalogVersion
        User = get_user_model()
        cls.company = Company.objects.create(name='PgCo', slug='pg-co')
        cls.op = User.objects.create_user('pg_op', password='x')
        Membership.objects.create(user=cls.op, company=cls.company,
                                  role=Membership.ROLE_OPERATOR)
        cls.lot = Lot.all_companies.create(number=700, origin='phone', operator=cls.op,
                                           company=cls.company)
        cur = CatalogVersion.current()
        for i in range(105):
            # snapshot_catalog_version atual → o on-read NÃO reclassifica.
            InventoryEntry.all_companies.create(
                lot=cls.lot, part_number=f'PG{i:04d}', quantity=1,
                chip_type='eMMC', company=cls.company,
                snapshot_catalog_version=cur)

    def setUp(self):
        _scope(self, self.company)

    def test_pagina_1_e_2(self):
        self.client.force_login(self.op)
        url = reverse('estoque:lot_detail', args=[self.lot.pk])
        resp = self.client.get(url)
        # class=" — o seletor .wtc-stock-row do CSS não entra na conta.
        self.assertEqual(resp.content.decode().count('class="erow"'), 100)
        self.assertContains(resp, 'página 1 de 2')
        self.assertContains(resp, 'Próxima')
        resp2 = self.client.get(url, {'p': '2'})
        self.assertEqual(resp2.content.decode().count('class="erow"'), 5)
        self.assertContains(resp2, 'página 2 de 2')
        self.assertContains(resp2, 'Anterior')

    def test_filtro_htmx_reseta_e_esconde_paginacao_com_pouco_resultado(self):
        self.client.force_login(self.op)
        url = reverse('estoque:lot_detail', args=[self.lot.pk])
        resp = self.client.get(url, {'q': 'PG0001'}, HTTP_HX_REQUEST='true')
        self.assertEqual(resp.content.decode().count('class="erow"'), 1)
        self.assertNotContains(resp, 'página 1 de')     # 1 página → sem rodapé


class PainelTests(TestCase):
    """/painel/ (lançadeira pós-login): hero = lote aberto → 1 clique para a
    bancada; empty-state orienta por papel; stats do dia como contexto."""

    @classmethod
    def setUpTestData(cls):
        User = get_user_model()
        cls.company = Company.objects.create(name='eMiner', slug='eminer')
        cls.op = User.objects.create_user('pn_op', password='x')
        cls.mgr = User.objects.create_user('pn_mgr', password='x')
        Membership.objects.create(user=cls.op, company=cls.company,
                                  role=Membership.ROLE_OPERATOR)
        Membership.objects.create(user=cls.mgr, company=cls.company,
                                  role=Membership.ROLE_MANAGER)

    def setUp(self):
        _scope(self, self.company)   # T3

    def test_hero_mostra_lote_aberto_com_cta(self):
        lot = Lot.objects.create(number=500, origin='phone', operator=self.mgr,
                                 company=self.company,
                                 description='Compra Jul/26')
        self.client.force_login(self.op)
        resp = self.client.get(reverse('painel'))
        self.assertContains(resp, 'LOT/500/')   # nomenclatura F11.2
        self.assertContains(resp, 'Continuar triagem')
        self.assertContains(resp, reverse('estoque:lot_detail', args=[lot.pk]))

    def test_lote_fechado_nao_vira_hero(self):
        Lot.objects.create(number=501, origin='phone', operator=self.mgr, company=self.company,
                           status=Lot.STATUS_CLOSED)
        self.client.force_login(self.op)
        resp = self.client.get(reverse('painel'))
        self.assertContains(resp, 'Nenhum lote aberto')
        self.assertNotContains(resp, 'Continuar triagem')

    def test_empty_state_orienta_por_papel(self):
        """Sem lote aberto: gerente ganha CTA de abrir; operador, a instrução
        de pedir ao gerente (ele não pode abrir — matriz §8)."""
        self.client.force_login(self.mgr)
        resp = self.client.get(reverse('painel'))
        self.assertContains(resp, 'Abrir um lote')

        self.client.force_login(self.op)
        resp = self.client.get(reverse('painel'))
        self.assertNotContains(resp, 'Abrir um lote')
        self.assertContains(resp, 'Peça ao gerente')

    def test_stats_do_dia(self):
        """ATUALIZADO ao redesenho do painel (f9da4a9, "limpeza do painel"): a
        seção "Números de hoje" SAIU do template — as métricas viraram contexto
        do HERÓI (Unidades = lot.total_qty; Tipos/Categorias = lot.chip_count,
        rótulo conforme access.is_unmasked). O teste valida o layout novo com
        DADOS reais do lote (1 entrada de qty 1 → "1" nas duas métricas) — e
        documenta que o layout antigo não volta por acidente."""
        lot = Lot.objects.create(number=502, origin='phone', operator=self.mgr,
                                 company=self.company)
        InventoryEntry.objects.create(lot=lot, part_number='PNHOJE1')
        PendingEntry.objects.create(lot=lot, part_number='PNFILA1',
                                    operator=self.op)
        self.client.force_login(self.op)
        resp = self.client.get(reverse('painel'))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, 'Unidades')
        self.assertContains(resp, 'Categorias')     # rótulo mascarado (operador)
        # As métricas do hero carregam os dados do LOTE: 1 un., 1 PN.
        self.assertContains(resp, '<div class="mv">1</div>', count=2)
        self.assertNotContains(resp, 'Números de hoje')   # layout antigo saiu


# ═══════════════════════════════════════════════════════════════════════════
# T2 (PLANO_MULTITENANT.md §7) — numeração de lote atômica (prova do O3)
# ═══════════════════════════════════════════════════════════════════════════

class LotNumberSequenceTests(TestCase):
    """open_for_company em série: contador incrementa, herda seed, auto-cura
    drift (lote criado por fora do contador). Roda em qualquer banco."""

    def setUp(self):
        self.user = get_user_model().objects.create_user('seq_user')
        self.company = Company.objects.create(name='SeqCo', slug='seqco')

    def test_empresa_nova_comeca_no_lote_1(self):
        lot = Lot.open_for_company(self.company, self.user, 'primeiro', origin='phone')
        self.assertEqual(lot.number, 1)          # "Lote #001" (§5.2 do plano)
        self.assertEqual(Lot.open_for_company(self.company, self.user, origin='phone').number, 2)

    def test_herda_seed_do_bootstrap(self):
        Company.objects.filter(pk=self.company.pk).update(last_lot_number=40)
        self.company.refresh_from_db()
        self.assertEqual(Lot.open_for_company(self.company, self.user, origin='phone').number, 41)

    def test_auto_cura_drift_do_contador(self):
        """Lote criado POR FORA do contador (legado/manual) não gera colisão:
        o guard max(contador, Max(number)) pula para depois dele."""
        Lot.all_companies.create(number=77, origin='phone', operator=self.user,
                                 company=self.company)      # fora do contador
        lot = Lot.open_for_company(self.company, self.user, origin='phone')
        self.assertEqual(lot.number, 78)
        self.company.refresh_from_db()
        self.assertEqual(self.company.last_lot_number, 78)  # contador curado


class LotNumberRaceTests(TransactionTestCase):
    """O3 (§7 do plano): N threads abrindo lote JUNTAS → números sequenciais sem
    buraco, zero IntegrityError.

    ⚠ Postgres-only: select_for_update é NO-OP no SQLite e o settings_test usa
    SQLite em memória — aqui a prova passaria sem provar nada. Rodar de verdade:

        python manage.py test estoque.tests.LotNumberRaceTests
        # (settings default → seu Postgres local)
    """

    N_THREADS = 8

    @skipUnless(connection.vendor == 'postgresql',
                'select_for_update é no-op no SQLite — rodar contra Postgres')
    def test_corrida_de_numeracao(self):
        user = get_user_model().objects.create_user('race_user')
        company = Company.objects.create(name='RaceCo', slug='raceco')

        barrier = threading.Barrier(self.N_THREADS)   # todas partem juntas
        numbers, errors = [], []

        def worker():
            try:
                barrier.wait()
                # company_scope (T4): além do contextvar, seta o GUC do RLS na
                # conexão DESTA thread — com FORCE RLS o insert seria rejeitado.
                with company_scope(company):
                    lot = Lot.open_for_company(company, user, origin='phone')
                numbers.append(lot.number)
            except Exception as exc:                  # IntegrityError incluso
                errors.append(exc)
            finally:
                connection.close()                    # conexão da thread

        threads = [threading.Thread(target=worker) for _ in range(self.N_THREADS)]
        for t in threads:
            t.start()
        for t in threads:
            t.join()

        self.assertEqual(errors, [])                             # zero quebra
        self.assertEqual(sorted(numbers),
                         list(range(1, self.N_THREADS + 1)))     # sem buraco
        company.refresh_from_db()
        self.assertEqual(company.last_lot_number, self.N_THREADS)


# ═══════════════════════════════════════════════════════════════════════════
# T3 (PLANO_MULTITENANT.md §12.1) — HANDSHAKE DE TENANCY (prova do O4)
# A fronteira comercial absoluta: a empresa A JAMAIS vê lote/estoque/fila da B.
# ═══════════════════════════════════════════════════════════════════════════

_PROBE_ROLE = 'wtc_rls_probe'


def enter_non_superuser(testcase, cur):
    """Armadilha §6.2.1: role SUPERUSER bypassa RLS por completo (até com
    FORCE) — e o dev local costuma conectar como super (o Render não). Se o
    role atual é super, troca para um role de sondagem SEM superuser (membro
    do role original → mesmos privilégios, RLS aplica); cleanup faz RESET+DROP.
    Reutilizado pelos testes de RLS do estoque e do pricing."""
    cur.execute('SELECT rolsuper FROM pg_roles WHERE rolname = current_user')
    if not cur.fetchone()[0]:
        return          # já é não-super (prod-like): RLS vale direto
    cur.execute('SELECT current_user')
    original = cur.fetchone()[0]

    def _leave():
        with connection.cursor() as c:
            c.execute('RESET ROLE')
            c.execute(f'DROP ROLE IF EXISTS {_PROBE_ROLE}')
    testcase.addCleanup(_leave)

    cur.execute(f'DROP ROLE IF EXISTS {_PROBE_ROLE}')
    cur.execute(f'CREATE ROLE {_PROBE_ROLE}')            # sem SUPER/LOGIN
    cur.execute(f'GRANT "{original}" TO {_PROBE_ROLE}')  # herda privilégios
    cur.execute(f'SET ROLE {_PROBE_ROLE}')

class TenancyHandshakeTests(TestCase):
    """Permanente na suíte (no espírito do golden/handshake de rentabilidade):
    cria empresas A e B com dados e prova que A não lê/edita/deleta/exporta
    NADA de B — via manager padrão (Camada A) e via views. A camada B (RLS,
    SQL cru) entra na T4 com o teste próprio."""

    @classmethod
    def setUpTestData(cls):
        User = get_user_model()
        cls.a = Company.objects.create(name='eMiner', slug='eminer')
        cls.b = Company.objects.create(name='Brasil Reciclagem', slug='brasil')
        cls.mgr_a = User.objects.create_user('hs_mgr_a', password='x')
        cls.mgr_b = User.objects.create_user('hs_mgr_b', password='x')
        Membership.objects.create(user=cls.mgr_a, company=cls.a,
                                  role=Membership.ROLE_MANAGER)
        Membership.objects.create(user=cls.mgr_b, company=cls.b,
                                  role=Membership.ROLE_MANAGER)
        cls.lot_a = Lot.open_for_company(cls.a, cls.mgr_a, 'lote da eMiner', origin='phone')
        cls.lot_b = Lot.open_for_company(cls.b, cls.mgr_b, 'lote da Brasil', origin='phone')
        # company dos filhos herda do lote no save() (CompanyBoundByLot).
        cls.entry_a = InventoryEntry.all_companies.create(
            lot=cls.lot_a, part_number='PN_DA_EMINER', quantity=3)
        cls.entry_b = InventoryEntry.all_companies.create(
            lot=cls.lot_b, part_number='PN_DA_BRASIL', quantity=5)
        PendingEntry.all_companies.create(
            lot=cls.lot_b, part_number='PEND_DA_BRASIL', operator=cls.mgr_b)
        RejectedEntry.all_companies.create(
            lot=cls.lot_b, part_number='REJ_DA_BRASIL', operator=cls.mgr_b)

    # ── Numeração por empresa ────────────────────────────────────────────────
    def test_cada_empresa_tem_sua_sequencia(self):
        self.assertEqual(self.lot_a.number, 1)
        self.assertEqual(self.lot_b.number, 1)   # mesmo nº, empresas diferentes
        self.assertEqual(self.entry_b.company_id, self.b.pk)  # denormalizado

    # ── Fail-closed (Camada A) ───────────────────────────────────────────────
    def test_sem_escopo_explode_nunca_vaza_todos(self):
        for Model in (Lot, InventoryEntry, PendingEntry, RejectedEntry):
            with self.subTest(model=Model.__name__):
                with self.assertRaises(CompanyScopeMissing):
                    list(Model.objects.all())

    # ── ORM escopado: leitura e escrita ──────────────────────────────────────
    def test_orm_da_a_nao_le_nem_edita_a_b(self):
        with company_scope(self.a):
            self.assertEqual(
                list(Lot.objects.values_list('description', flat=True)),
                ['lote da eMiner'])
            self.assertFalse(
                InventoryEntry.objects.filter(part_number='PN_DA_BRASIL').exists())
            self.assertEqual(PendingEntry.objects.count(), 0)
            self.assertEqual(RejectedEntry.objects.count(), 0)
            # Escrita cross-company pelo manager padrão = 0 linhas afetadas.
            self.assertEqual(
                InventoryEntry.objects.filter(pk=self.entry_b.pk).update(quantity=99), 0)
            self.assertEqual(Lot.objects.filter(pk=self.lot_b.pk).delete()[0], 0)
        with company_scope(self.b):
            self.entry_b.refresh_from_db()
            self.assertEqual(self.entry_b.quantity, 5)         # intacta
            self.assertTrue(Lot.objects.filter(pk=self.lot_b.pk).exists())

    # ── Views: gerente da B tentando o lote da A ─────────────────────────────
    def test_views_da_a_devolvem_404_para_a_b(self):
        self.client.force_login(self.mgr_b)
        casos = [
            ('detail', 'get', reverse('estoque:lot_detail', args=[self.lot_a.pk]), {}),
            ('export', 'get', reverse('estoque:export', args=[self.lot_a.pk]), {}),
            ('preview', 'get', reverse('estoque:preview', args=[self.lot_a.pk]),
             {'pn': 'KMQX10013M'}),
            ('add', 'post', reverse('estoque:add', args=[self.lot_a.pk]),
             {'pn': 'KMQX10013M', 'qty': '1'}),
            ('close', 'post', reverse('estoque:lot_close', args=[self.lot_a.pk]), {}),
            ('remove', 'post',
             reverse('estoque:remove', args=[self.lot_a.pk, self.entry_a.pk]),
             {'qty': '1'}),
        ]
        for nome, metodo, url, data in casos:
            with self.subTest(view=nome):
                resp = getattr(self.client, metodo)(url, data)
                self.assertEqual(resp.status_code, 404)   # nem existe para a B

    def test_lista_e_painel_mostram_so_a_propria_empresa(self):
        self.client.force_login(self.mgr_b)
        body = self.client.get(reverse('estoque:index')).content.decode()
        self.assertIn('lote da Brasil', body)
        self.assertNotIn('lote da eMiner', body)
        painel = self.client.get(reverse('painel')).content.decode()
        self.assertNotIn('lote da eMiner', painel)


class RLSHandshakeTests(TransactionTestCase):
    """T4 (§12.1) — handshake da CAMADA B: prova por SQL CRU que o Postgres
    filtra sozinho (RLS + FORCE), sem confiar em nenhuma linha de Python.
    É o teste que remove a confiança no código: query bugada não cruza empresa.

    ⚠ Postgres-only (no SQLite as policies nem existem; a Camada A cobre lá):

        python manage.py test estoque.tests.RLSHandshakeTests

    ⚠ Armadilha §6.2.1: role SUPERUSER bypassa RLS por completo (até com
    FORCE) — e o dev local costuma conectar como super (o Render não). Se o
    role atual é super, o teste troca para um role de sondagem SEM superuser
    (membro do role atual → mesmos privilégios, RLS aplica) só nas asserções.
    """

    @skipUnless(connection.vendor == 'postgresql', 'RLS é Postgres-only')
    def test_sql_cru_respeita_o_rls(self):
        User = get_user_model()
        a = Company.objects.create(name='RlsA', slug='rlsa')
        b = Company.objects.create(name='RlsB', slug='rlsb')
        ua = User.objects.create_user('rls_ua')
        ub = User.objects.create_user('rls_ub')
        with company_scope(a):
            lot_a = Lot.open_for_company(a, ua, 'lote RLS A', origin='phone')
            InventoryEntry.objects.create(lot=lot_a, part_number='RLSPN_A')
        with company_scope(b):
            lot_b = Lot.open_for_company(b, ub, 'lote RLS B', origin='phone')

        def _clear_gucs():
            with connection.cursor() as c:
                c.execute("SELECT set_config('app.company_id', '', false)")
                c.execute("SELECT set_config('app.platform', '', false)")
        self.addCleanup(_clear_gucs)
        _clear_gucs()

        with connection.cursor() as cur:
            enter_non_superuser(self, cur)   # §6.2.1: super bypassa RLS
            # 1) SEM GUC → FORCE RLS vale até para o dono da tabela: 0 linhas.
            cur.execute('SELECT count(*) FROM estoque_lot')
            self.assertEqual(cur.fetchone()[0], 0)
            cur.execute('SELECT count(*) FROM estoque_inventoryentry')
            self.assertEqual(cur.fetchone()[0], 0)

            # 2) GUC da empresa A → SÓ as linhas da A; a da B "não existe"
            #    nem pedindo explicitamente por WHERE.
            cur.execute("SELECT set_config('app.company_id', %s, false)",
                        [str(a.pk)])
            cur.execute('SELECT company_id FROM estoque_lot')
            self.assertEqual([r[0] for r in cur.fetchall()], [a.pk])
            cur.execute('SELECT count(*) FROM estoque_lot WHERE company_id = %s',
                        [b.pk])
            self.assertEqual(cur.fetchone()[0], 0)

            # 3) Escrita cruzada barrada NO BANCO: insert com company da B sob
            #    o GUC da A viola o WITH CHECK da policy.
            from django.db import Error as DBError
            with self.assertRaises(DBError):
                cur.execute(
                    "INSERT INTO estoque_rejectedentry "
                    "(lot_id, part_number, quantity, chip_type, brand, capacity,"
                    " emcp_ram, emcp_nand, is_emcp, interface,"
                    " classification_source, confidence, rejection_reason,"
                    " operator_id, created_at, company_id) "
                    "VALUES (%s, 'HACK', 1, '', '', '', '', '', false, '', '',"
                    " '', 'X', %s, now(), %s)",
                    [lot_b.pk, ub.pk, b.pk])

            # 4) GUC de plataforma → enxerga as duas (Django admin/superuser).
            cur.execute("SELECT set_config('app.company_id', '', false)")
            cur.execute("SELECT set_config('app.platform', '1', false)")
            cur.execute('SELECT count(*) FROM estoque_lot')
            self.assertEqual(cur.fetchone()[0], 2)


class PlatformAdminFormTests(TestCase):
    """Regressão do bug 2026-07-09 (produção): superuser de PLATAFORMA (sem
    Membership) criando registro tenant-scoped pelo Django admin explodia com
    CompanyScopeMissing — o Django 5 valida UniqueConstraint de formulário via
    `_default_manager`, que era o fail-closed. Com `default_manager_name =
    'all_companies'`, o form valida e salva; o escopo das VIEWS segue explícito
    (Model.objects)."""

    def setUp(self):
        User = get_user_model()
        self.company = Company.objects.create(name='eMiner', slug='eminer')
        self.platform = User.objects.create_superuser('plat_admin', password='x')
        self.client.force_login(self.platform)   # sem Membership de propósito

    def test_plataforma_cria_lote_pelo_admin_sem_escopo(self):
        operador = get_user_model().objects.create_user('adm_form_op')
        resp = self.client.post('/admin/estoque/lot/add/', {
            'number': '77', 'company': str(self.company.pk),
            'operator': str(operador.pk), 'description': 'via admin',
            'status': 'open', 'origin': 'phone',
        })
        # Sucesso = redirect pro changelist (antes: CompanyScopeMissing/500).
        self.assertEqual(resp.status_code, 302)
        self.assertTrue(Lot.all_companies.filter(
            number=77, company=self.company).exists())

    def test_plataforma_cria_comprador_pelo_admin_sem_escopo(self):
        """O caso EXATO do bug em produção (/admin/pricing/buyer/add/)."""
        from pricing.models import Buyer
        resp = self.client.post('/admin/pricing/buyer/add/', {
            'company': str(self.company.pk), 'name': 'Wu Quan',
            'slug': 'wu-quan', 'active': 'on', 'notes': '',
            'fx_usd_rate': '0.14',   # F10.1: taxa contratual ¥→US$
        })
        self.assertEqual(resp.status_code, 302)
        buyer = Buyer.all_companies.get(slug='wu-quan')
        self.assertEqual(buyer.company_id, self.company.pk)


class ExportPriceColumnsTests(TestCase):
    """Export .xlsx com preço (feature 2026-07-09): colunas "Preço unit./Total —
    <comprador> (USD)" **SÓ para ADMIN** (matriz §8: gerente não vê preço — ele
    exporta a planilha sem as colunas; operador nem exporta, gate manager+ já
    provado na RoleMatrixTests)."""

    @classmethod
    def setUpTestData(cls):
        User = get_user_model()
        cls.company = Company.objects.create(name='eMiner', slug='eminer',
                                             is_platform=True)   # F12
        # v3.1: visão completa (colunas de spec) = só superuser; os papéis
        # (admin/gerente) seguem valendo pro gate SEPARADO de preço.
        cls.adm = User.objects.create_user('xp_adm', password='x', is_superuser=True)
        cls.mgr = User.objects.create_user('xp_mgr', password='x', is_superuser=True)
        Membership.objects.create(user=cls.adm, company=cls.company,
                                  role=Membership.ROLE_ADMIN)
        Membership.objects.create(user=cls.mgr, company=cls.company,
                                  role=Membership.ROLE_MANAGER)
        cls.lot = Lot.all_companies.create(number=600, origin='phone', operator=cls.adm,
                                           company=cls.company)
        InventoryEntry.all_companies.create(
            lot=cls.lot, part_number='KLMAG1JETD',
            chip_type='eMMC', capacity='16GB', quantity=3)
        # Catálogo: PN confirmado (o price_lot reclassifica cada PN on-read).
        from chips.models import Brand as ChipBrand, KnownPart
        brand = ChipBrand.objects.create(name='Samsung', code='SAM')
        KnownPart.objects.create(part_number='KLMAG1JETD', brand=brand,
                                 chip_type='eMMC', capacity='16GB',
                                 confidence='confirmed', review_status='approved')
        # Comprador com preço: eMMC 16GB = ¥ 15 (lista genérica). F10 (RMB
        # canônico): o banco guarda ¥; o export segue em USD DERIVADO pela
        # taxa contratual default 0.14 → US$ 2,10 unitário.
        from pricing.models import Buyer, Price, PriceList
        buyer = Buyer.all_companies.create(name='Wuquan', slug='wuquan',
                                           company=cls.company)
        pl = PriceList.all_companies.create(buyer=buyer, company=cls.company)
        Price.all_companies.create(price_list=pl, kind='emmc', gen='', origin='phone',
                                   tier_value=16, tier_unit='GB',
                                   status='quoted', price_min='15',
                                   price_max='15', company=cls.company)

    def _sheet(self, user):
        import io as _io
        import openpyxl
        self.client.force_login(user)
        resp = self.client.get(reverse('estoque:export', args=[self.lot.pk]))
        self.assertEqual(resp.status_code, 200)
        return openpyxl.load_workbook(_io.BytesIO(resp.content)).active

    def test_admin_exporta_com_preco_unitario_e_total(self):
        ws = self._sheet(self.adm)
        headers = [c.value for c in ws[1]]
        # F11.3: contraparte é o WhatTheChip — nome do comprador nunca sai.
        # PLANO_FX Fase A: ¥ primeiro (coluna própria) + US$ como "≈"
        self.assertIn('Preço unit. — WhatTheChip (¥ RMB)', headers)
        self.assertIn('Preço unit. — WhatTheChip (US$ ≈)', headers)
        self.assertIn('Total — WhatTheChip (US$ ≈)', headers)
        rmb_col = headers.index('Preço unit. — WhatTheChip (¥ RMB)') + 1
        self.assertEqual(ws.cell(row=2, column=rmb_col).value, 15.0)  # ¥15
        unit_col = headers.index('Preço unit. — WhatTheChip (US$ ≈)') + 1
        # USD DERIVADO (F10): ¥15 × 0.14 = 2.10 — o export NUNCA vê ¥.
        self.assertEqual(ws.cell(row=2, column=unit_col).value, 2.1)
        self.assertEqual(ws.cell(row=2, column=unit_col + 1).value, 6.3)  # 3 × 2,10
        self.assertEqual(ws.cell(row=3, column=unit_col + 1).value, 6.3)  # TOTAL geral

    def test_gerente_exporta_sem_colunas_de_preco(self):
        """Gerente continua exportando (papel dele) — mas a planilha vem SEM
        preço (a decisão 'gerente não vê preço' vale em toda superfície)."""
        ws = self._sheet(self.mgr)
        headers = [c.value for c in ws[1]]
        self.assertEqual(len(headers), 8)                 # só as colunas base
        self.assertFalse(any('Preço' in str(h) for h in headers if h))


class MaskingTests(TestCase):
    """F12 (dono, 2026-07-17): o conhecimento "PN → o que é → quanto vale" é
    o ativo do negócio. Empresa-CLIENTE vê só o código de caixa (bancada com
    card whitelist SEM data-debug, tabela, export); só o SUPERUSER (v3.1,
    dono 2026-07-23) vê tudo — nem membro da eMiner mais."""

    @classmethod
    def setUpTestData(cls):
        from decimal import Decimal
        from pricing.models import CategoryCode
        User = get_user_model()
        cls.plat = Company.objects.create(name='eMiner HQ', slug='mask-plat',
                                          is_platform=True)
        cls.cli = Company.objects.create(name='Cliente X', slug='mask-cli')
        cls.users = {}
        for tag, co in (('plat', cls.plat), ('cli', cls.cli)):
            # v3.1 (dono 2026-07-23): SÓ superuser vê tudo — o membro da
            # plataforma vira superuser; o cliente segue mascarado.
            u = User.objects.create_user(f'mask_{tag}', password='x',
                                         is_superuser=(tag == 'plat'))
            Membership.objects.create(user=u, company=co,
                                      role=Membership.ROLE_MANAGER)  # export
            cls.users[tag] = u
        # v3: emmc = letra B (pricing/convention.py) → rótulo 'B-53'.
        cls.code = CategoryCode.objects.create(
            kind='emmc', gen='', tier_value=Decimal('16'), tier_unit='GB',
            code=53)
        # Confirmado no catálogo → bancada aprova (gate "só confirmados").
        from chips.models import Brand as ChipBrand, KnownPart
        b = ChipBrand.objects.create(name='Samsung MK', code='SAMMK')
        KnownPart.objects.create(part_number='MASKPN16', brand=b,
                                 chip_type='eMMC', capacity='16GB',
                                 confidence='confirmed',
                                 review_status='approved')

    def _lot(self, company, user):
        set_current_company(company.pk)
        self.addCleanup(set_current_company, None)
        lot = Lot.all_companies.create(number=800 + company.pk, origin='phone',
                                       operator=user, company=company)
        return lot

    def test_bancada_cliente_mascarada_plataforma_completa(self):
        lot_c = self._lot(self.cli, self.users['cli'])
        self.client.login(username='mask_cli', password='x')
        resp = self.client.get(
            reverse('estoque:preview', args=[lot_c.pk]), {'pn': 'MASKPN16'})
        body = resp.content.decode()
        self.assertIn('B-53', body)                   # código da categoria (v3)
        self.assertIn('Caixa', body)
        self.assertNotIn('eMMC', body)                # specs NÃO vazam
        self.assertNotIn('16GB', body)
        self.assertNotIn('data-debug', body)          # o JSON inteiro tampouco
        self.assertNotIn('Rentável', body)            # veredito nominal some
        # Plataforma: card completo, com specs e debug.
        lot_p = self._lot(self.plat, self.users['plat'])
        self.client.login(username='mask_plat', password='x')
        resp = self.client.get(
            reverse('estoque:preview', args=[lot_p.pk]), {'pn': 'MASKPN16'})
        body = resp.content.decode()
        self.assertIn('eMMC', body)
        self.assertIn('data-debug', body)

    def test_preview_pn_nao_encontrado_nao_quebra(self):
        """Regressão (2026-08-05): PN não classificado → classify devolve dict
        REDUZIDO (in_review_queue, SEM as chaves de spec string). O confirm_card
        acessava result.capacity como ARGUMENTO de filtro → VariableDoesNotExist
        (500). A view agora garante as chaves-padrão; o preview responde 200 sem
        specs (exposto ao digitar o prefixo de um PN em fila de revisão — ESMT M15T)."""
        lot = self._lot(self.plat, self.users['plat'])
        self.client.login(username='mask_plat', password='x')
        resp = self.client.get(
            reverse('estoque:preview', args=[lot.pk]), {'pn': 'ZZUNKNOWNPN9'})
        self.assertEqual(resp.status_code, 200)   # antes do fix: 500

    def test_tabela_e_export_mascarados(self):
        from decimal import Decimal
        from chips.models import CatalogVersion
        lot = self._lot(self.cli, self.users['cli'])
        InventoryEntry.all_companies.create(
            lot=lot, part_number='MASKPN16', quantity=3, brand='Samsung MK',
            chip_type='eMMC', capacity='16GB', company=self.cli,
            snapshot_catalog_version=CatalogVersion.current(),
            price_kind='emmc', price_gen='', price_tier_value=Decimal('16'),
            price_tier_unit='GB')
        self.client.login(username='mask_cli', password='x')
        page = self.client.get(
            reverse('estoque:lot_detail', args=[lot.pk])).content.decode()
        self.assertIn('B-53', page)
        self.assertNotIn('>eMMC<', page)              # badge de tipo sumiu
        import io as _io
        import openpyxl
        resp = self.client.get(reverse('estoque:export', args=[lot.pk]))
        ws = openpyxl.load_workbook(_io.BytesIO(resp.content)).active
        headers = [c.value for c in ws[1]]
        self.assertIn('Category', headers)
        self.assertNotIn('Type', headers)             # colunas de spec sumiram
        self.assertNotIn('Capacity', headers)
        self.assertEqual(ws.cell(row=2, column=2).value, 'B-53')


class SeedCategoryCodesTests(TestCase):
    """F12 v3 (convenção universal, dono 2026-07-23): o seed é DETERMINÍSTICO
    — carrega exatamente a TABELA FUNDADORA de pricing/convention.py (mesmos
    códigos em qualquer deploy do mundo). Idempotente; divergência banco ×
    convenção é ERRO alto; --reset recarrega (pré-deploy). A categoria deriva
    do CHIP — existe SEM preço/grid ("preço até pode ficar sem, categoria
    não"); o Geral/C-000 foi DESFEITO; leitura nunca cunha."""

    def test_seed_deterministico_e_convencao(self):
        from decimal import Decimal
        from io import StringIO
        from django.core.management import call_command
        from pricing.convention import FOUNDING_TABLE, KIND_LETTER
        from pricing.models import CategoryCode

        out = StringIO()
        call_command('seed_category_codes', stdout=out)          # dry-run
        self.assertEqual(CategoryCode.objects.count(), 0)
        call_command('seed_category_codes', commit=True, stdout=out)
        self.assertEqual(CategoryCode.objects.count(), len(FOUNDING_TABLE))
        # Determinístico: re-rodar não muda NADA (e confere linha a linha).
        call_command('seed_category_codes', commit=True, stdout=out)
        self.assertEqual(CategoryCode.objects.count(), len(FOUNDING_TABLE))
        # A convenção do dono (v3.1): eMCP 64GB = A-02 — SÓ pelo NAND.
        self.assertEqual(CategoryCode.label_for_key(
            'emcp', '', Decimal('64'), 'GB'), 'A-02')
        # Qualquer geração de RAM cai na MESMA caixa (fold → vazio):
        self.assertEqual(CategoryCode.label_for_key(
            'emcp', 'LPDDR4X', Decimal('64'), 'GB'), 'A-02')
        self.assertEqual(CategoryCode.label_for_key(
            'emcp', 'LPDDR3', Decimal('64'), 'GB'), 'A-02')
        # Números são POR LETRA: A-06 e B-06 coexistem (emmc 16GB = B-06).
        self.assertEqual(CategoryCode.label_for_key(
            'emmc', '', Decimal('16'), 'GB'), 'B-06')

        # Categoria INÉDITA nasce na APROVAÇÃO, SEM depender de grid/preço
        # (zero Price no banco deste teste) — próximo número livre DA letra:
        n_ufs = CategoryCode.objects.filter(kind='ufs').count()
        lbl = CategoryCode.label_for_key('ufs', '', Decimal('2048'), 'GB')
        self.assertEqual(lbl, f'D-{n_ufs + 1:02d}')

        # LEITURA nunca cunha (create=False) e kind fora da convenção → None:
        antes = CategoryCode.objects.count()
        self.assertIsNone(CategoryCode.label_for_key(
            'ddr', 'DDR2', Decimal('1'), 'Gb', create=False))
        self.assertIsNone(CategoryCode.label_for_key(
            'gddr', 'GDDR5', Decimal('8'), 'Gb'))   # kind extinto
        self.assertEqual(CategoryCode.objects.count(), antes)

        # --reset recarrega a tabela fundadora (some o D-#### inédito local).
        call_command('seed_category_codes', commit=True, reset=True,
                     stdout=out)
        self.assertEqual(CategoryCode.objects.count(), len(FOUNDING_TABLE))
        # Sanidade da convenção: letras batem com KIND_LETTER.
        for c in CategoryCode.objects.all():
            self.assertIn(c.kind, KIND_LETTER)
            self.assertTrue(c.label.startswith(f'{KIND_LETTER[c.kind]}-'))


class DebugButtonGateTests(TestCase):
    """O 📋 (copiar diagnóstico) é ferramenta de suporte da PLATAFORMA — só
    SUPERUSER vê (dono, 2026-07-20): nem admin de empresa, nem operador da
    plataforma. v3.1 (dono 2026-07-23): o card COMPLETO (com data-debug) agora
    também é só do superuser — o admin de empresa passou a receber o card
    MASCARADO (sem data-debug e sem o botão). Superuser recebe os dois."""

    def setUp(self):
        self.user = get_user_model().objects.create_user(username='dbg_op',
                                                         password='x')
        self.company = _grant(self.user, role=Membership.ROLE_ADMIN)
        _scope(self, self.company)
        self.lot = Lot.objects.create(number=571, origin='phone', operator=self.user,
                                      company=self.company)
        self.url = reverse('estoque:preview', args=[self.lot.pk])

    def _body(self, user):
        self.client.force_login(user)
        with patch('estoque.views.classify') as mock_classify:
            mock_classify.return_value = _result(
                chip_type='eMMC', capacity='16GB',
                classification_source='banco de dados', confidence='confirmed')
            return self.client.get(self.url, {'pn': 'KLMAG1JETD'}).content.decode()

    def test_admin_de_empresa_nao_ve_o_botao(self):
        # v3.1: admin de empresa é MASCARADO — não recebe o card completo
        # (data-debug) nem o botão 📋. Antes (is_platform) via o card completo
        # sem o botão; agora não vê nenhum dos dois.
        body = self._body(self.user)
        self.assertNotIn('data-debug', body)         # card MASCARADO (cliente)
        self.assertNotIn('est-debug-btn', body)      # botão 📋 NÃO

    def test_superuser_ve_o_botao(self):
        su = get_user_model().objects.create_superuser(username='dbg_su',
                                                       password='x')
        # Superuser navega o app com Membership REAL (sem bypass nos gates).
        Membership.objects.create(user=su, company=self.company,
                                  role=Membership.ROLE_OPERATOR)
        body = self._body(su)
        self.assertIn('est-debug-btn', body)


class MaskedFuzzyDiffTests(TestCase):
    """Card MASCARADO mantém o diff visual do typo (verde = o que falta do
    digitado) — regressão apontada pelo dono 2026-07-20. Só caracteres de PN:
    nenhuma spec entra na sugestão."""

    def setUp(self):
        User = get_user_model()
        self.cli = Company.objects.create(name='Cliente FZ', slug='cli-fz')
        self.user = User.objects.create_user('fz_cli', password='x')
        Membership.objects.create(user=self.user, company=self.cli,
                                  role=Membership.ROLE_OPERATOR)
        _scope(self, self.cli)
        self.lot = Lot.all_companies.create(number=572, origin='phone', operator=self.user,
                                            company=self.cli)
        self.client.force_login(self.user)

    @patch('estoque.views.classify')
    def test_diff_verde_presente_no_card_mascarado(self, mock_classify):
        mock_classify.return_value = _result(
            fuzzy_suggestions=['KLMAG1JETD'])
        url = reverse('estoque:preview', args=[self.lot.pk])
        body = self.client.get(url, {'pn': 'KLMAG1JET0'}).content.decode()
        self.assertIn('wtc-fuzzy-pn', body)                    # span do diff
        self.assertIn('data-suggestion="KLMAG1JETD"', body)
        self.assertIn('#198038', body)                         # verde do diff
        self.assertNotIn('data-debug', body)                   # máscara intacta
        self.assertNotIn('est-debug-btn', body)


class TemplateMultilineCommentTests(TestCase):
    """PORTÃO (3ª ocorrência do erro, 2026-07-16 — CLAUDE.md §7): `{# #}` do
    Django é SÓ single-line; com quebra de linha o "comentário" VAZA como
    texto puro na página. Comentário longo = {% comment %}. Este teste varre
    todos os templates dos apps e deixa a suíte vermelha na hora."""

    def test_nenhum_template_tem_comentario_hash_multilinha(self):
        import pathlib
        from django.conf import settings
        base = pathlib.Path(settings.BASE_DIR)
        maus = []
        raizes = [base / app for app in
                  ('chips', 'estoque', 'pages', 'tenancy', 'pricing',
                   'vendas')] + [base / 'templates']
        for raiz in raizes:
            for p in raiz.glob('**/templates/**/*.html'):
                self._scan(p, base, maus)
            if raiz.name == 'templates':
                for p in raiz.glob('**/*.html'):
                    self._scan(p, base, maus)
        self.assertEqual(
            maus, [],
            '{# #} multiline VAZA como texto — troque por {% comment %}: '
            + ', '.join(maus))

    @staticmethod
    def _scan(path, base, maus):
        t = path.read_text(encoding='utf-8')
        i = 0
        while True:
            s = t.find('{#', i)
            if s == -1:
                return
            e = t.find('#}', s)
            if e == -1 or '\n' in t[s:e]:
                maus.append(f'{path.relative_to(base)}:{t[:s].count(chr(10)) + 1}')
                if e == -1:
                    return
            i = e + 2


class TenancyDeclarationTests(TestCase):
    """§12.1: "tabela sem decisão de tenancy = suíte vermelha". Todo modelo dos
    apps do projeto ou está na lista GLOBAL declarada (PRECIFICACAO §10), ou é
    escopado (campo company + CompanyScopedManager como manager padrão).
    Criou modelo novo e este teste quebrou? Decida o tenancy dele AQUI."""

    GLOBAL_DECLARADOS = {
        # Catálogo/produto — o "Google dos chips" é um cérebro só (§10).
        'chips.Brand', 'chips.ChipFamily', 'chips.DecodeMap', 'chips.KnownPart',
        'chips.Source', 'chips.SearchLog', 'chips.UnknownChip',
        'chips.CorrectionRequest', 'chips.ChipSubmission',
        'chips.ProfitabilityConfig', 'chips.CatalogVersion',
        # CMS de documentação.
        'pages.Page',
        # O próprio tecido do tenancy.
        'tenancy.Company', 'tenancy.Branch', 'tenancy.Membership',
        # E4 (B4+B7): blob do logo, 1-pra-1 com a Company (o pk É a company).
        # Branding é PÚBLICO — a view company_logo serve ANÔNIMO (tela de
        # login do subdomínio); manager fail-closed ali seria contrassenso.
        # Sem RLS, como o resto do tecido de tenancy (§6 do PLANO).
        'tenancy.CompanyLogo',
        # i18n: a preferência de idioma é da PESSOA, não da empresa (um técnico
        # chinês numa empresa paraguaia lê 中文) — GLOBAL. Ver I18N.md §3.
        'tenancy.UserLanguage',
        # Config do sistema de preços — singleton GLOBAL (padrão
        # ProfitabilityConfig; PRECIFICACAO §3.4). Buyer/PriceList/Price são
        # ESCOPADOS (company + CompanyScopedManager + RLS em pricing/0002).
        'pricing.PricingConfig',
        'pricing.FxRate',           # câmbio mid-market diário — dado de plataforma (PLANO_FX)
        # F12: dicionário GLOBAL código C-### ↔ categoria (chave de
        # preço) — mesma tabela para todo cliente (decisão do dono:
        # auditabilidade > embaralhamento por empresa).
        'pricing.CategoryCode',
        # Idempotência do add_chip (bug Mundo Metal, 2026-08-10): ledger de
        # tokens uuid4 + timestamp, SEM dado de tenant DE PROPÓSITO (nenhum
        # pn/lote/empresa) — não há o que isolar, e um uuid4 não colide entre
        # empresas. Fora do RLS; poda lazy de 48h no próprio add_chip.
        'estoque.SubmitToken',
    }
    # F11.2: 'vendas' entra — SalesOrder/SalesOrderLine/DocSequence são
    # ESCOPADOS (company + CompanyScopedManager + RLS em vendas/0002).
    APPS_DO_PROJETO = {'chips', 'estoque', 'pages', 'tenancy', 'pricing',
                       'vendas'}

    def test_toda_tabela_declara_tenancy(self):
        from django.apps import apps as django_apps
        faltando = []
        for model in django_apps.get_models():
            if model._meta.app_label not in self.APPS_DO_PROJETO:
                continue    # django/pghistory internos
            if hasattr(model, 'pgh_tracked_model'):
                continue    # tabela de evento pghistory (espelha a rastreada;
                            # ganha policy própria na T4)
            label = f'{model._meta.app_label}.{model.__name__}'
            if label in self.GLOBAL_DECLARADOS:
                continue
            has_company = any(f.name == 'company' for f in model._meta.fields)
            # Convenção completa (bug 2026-07-09): `objects` = escopado
            # fail-closed (o caminho EXPLÍCITO das views), e o DEFAULT = cru
            # ('all_companies') — a validação de UniqueConstraint do Django 5
            # e o admin usam _default_manager, que não pode ser fail-closed.
            objects_scoped = isinstance(getattr(model, 'objects', None),
                                        CompanyScopedManager)
            default_cru = model._meta.default_manager_name == 'all_companies'
            if not (has_company and objects_scoped and default_cru):
                faltando.append(label)
        self.assertEqual(
            faltando, [],
            'Modelo(s) SEM decisão de tenancy completa: ou adicione à lista '
            'GLOBAL_DECLARADOS (com justificativa no PR), ou dê a ele campo '
            'company + objects=CompanyScopedManager + all_companies + '
            "Meta.base_manager_name/default_manager_name='all_companies' + "
            f'caso no TenancyHandshakeTests: {faltando}')


class ReplicateLotXlsxTests(TestCase):
    """replicate_lot_xlsx (2026-07-31): réplica de lote a partir do export —
    reclassifica TUDO no engine local (bypass do portão de propósito: é
    material que JÁ ESTÁ em prod) e cria o lote com snapshot+chave daqui."""

    @classmethod
    def setUpTestData(cls):
        import openpyxl
        import tempfile
        from chips.models import Brand as ChipBrand, ChipFamily, DecodeMap
        User = get_user_model()
        cls.su = User.objects.create_superuser('root_rep', password='x')
        cls.company = Company.objects.create(name='RepCo', slug='repco')
        sam = ChipBrand.objects.create(name='Samsung REP', code='SAMREP')
        DecodeMap.objects.create(map_name='CAP_REP', char_key='A',
                                 val_primary='16GB', val_secondary='',
                                 brand=sam)
        ChipFamily.objects.create(brand=sam, prefix='KLM', chip_type='eMMC',
                                  subtype='', decode_cap_pos=3,
                                  decode_cap_map='CAP_REP', is_emcp=False,
                                  active=True, priority=50)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Part Number', 'Category', 'Qty.', 'Last Added',
                   'Preço unit.', 'Total'])
        ws.append(['KLMAG1JETD', 'B-05', 7, '31/07/2026', 1.0, 7.0])
        ws.append(['ZZDESCONHECIDO1', '—', 2, '31/07/2026', 'sem preço', None])
        ws.append(['TOTAL', None, 9, None, None, None])
        fh = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        wb.save(fh.name)
        cls.xlsx = fh.name

    def test_dry_run_projeta_e_nao_grava(self):
        from django.core.management import call_command
        from io import StringIO
        out = StringIO()
        call_command('replicate_lot_xlsx', self.xlsx, '--company', 'repco',
                     '--origin', 'phone', stdout=out)
        self.assertIn('2 PNs · 9 un.', out.getvalue())
        self.assertIn('DRY-RUN', out.getvalue())
        set_current_company(self.company)
        try:
            self.assertFalse(Lot.objects.exists())
        finally:
            set_current_company(None)

    def test_commit_cria_lote_com_chave_local(self):
        from django.core.management import call_command
        from io import StringIO
        out = StringIO()
        call_command('replicate_lot_xlsx', self.xlsx, '--company', 'repco',
                     '--origin', 'phone', '--commit', stdout=out)
        self.assertIn('✅', out.getvalue())
        set_current_company(self.company)
        try:
            lot = Lot.objects.get()
            self.assertEqual(lot.operator, self.su)
            e = InventoryEntry.objects.get(lot=lot, part_number='KLMAG1JETD')
            self.assertEqual(e.quantity, 7)
            # a chave nasce do engine LOCAL (eMMC 16GB — grid novo)
            self.assertEqual((e.price_kind, float(e.price_tier_value)),
                             ('emmc', 16.0))
            # o desconhecido entra MESMO ASSIM (réplica, não lançamento)
            zz = InventoryEntry.objects.get(lot=lot,
                                            part_number='ZZDESCONHECIDO1')
            self.assertEqual(zz.quantity, 2)
            self.assertTrue(zz.price_key_reason)
        finally:
            set_current_company(None)


class LotOriginTests(TestCase):
    """v4 (dono, 2026-08-01): origem do lote OBRIGATÓRIA e sem default na
    abertura — celular × PCB; define a tabela de preço do eMMC."""

    def setUp(self):
        User = get_user_model()
        self.mgr = User.objects.create_user(username='orig_mgr', password='x')
        self.company = _grant(self.mgr, role='manager')
        self.client.login(username='orig_mgr', password='x')

    def test_abrir_sem_origem_nao_cria(self):
        resp = self.client.post(reverse('estoque:lot_create'),
                                {'description': 'sem origem'}, follow=True)
        self.assertContains(resp, 'ORIGEM')
        _scope(self, self.company)
        self.assertFalse(Lot.objects.filter(description='sem origem').exists())

    def test_abrir_com_origem_pcb(self):
        self.client.post(reverse('estoque:lot_create'),
                         {'description': 'lote pcb', 'origin': 'pcb'})
        _scope(self, self.company)
        lot = Lot.objects.get(description='lote pcb')
        self.assertEqual(lot.origin, 'pcb')

    def test_open_for_company_valida_origem(self):
        from django.core.exceptions import ValidationError
        with self.assertRaises(ValidationError):
            Lot.open_for_company(self.company, self.mgr, 'x', origin='')
        with self.assertRaises(ValidationError):
            Lot.open_for_company(self.company, self.mgr, 'x', origin='misto')


class FxLockOnCloseTests(TestCase):
    """PLANO_FX Fase C (2026-08-01): fechar o lote TRAVA a taxa de mercado do
    dia (atômico com o CLOSED); reabrir é EXCLUSIVO do superuser (destrava;
    re-fechar captura taxa nova — pghistory loga); OV herda a taxa travada."""

    def setUp(self):
        from datetime import date
        from decimal import Decimal
        from pricing.models import FxRate
        globals()['Decimal'] = Decimal          # usado pelos testes da classe
        User = get_user_model()
        self.mgr = User.objects.create_user(username='fxc_mgr', password='x')
        self.company = _grant(self.mgr, role='manager')
        self.client.login(username='fxc_mgr', password='x')
        FxRate.objects.create(date=date.today(), rate=Decimal('0.1478'),
                              source='mid-market teste')
        _scope(self, self.company)
        self.lot = Lot.objects.create(number=910, origin='phone',
                                      operator=self.mgr, company=self.company)
        set_current_company(None)

    def _close(self):
        return self.client.post(reverse('estoque:lot_close',
                                        args=[self.lot.pk]),
                                {'confirm_code': self.lot.code})

    def test_fechar_trava_a_taxa_do_dia(self):
        self._close()
        _scope(self, self.company)
        self.lot.refresh_from_db()
        set_current_company(None)
        self.assertEqual(self.lot.status, 'closed')
        self.assertEqual(self.lot.fx_rate, Decimal('0.1478'))
        self.assertEqual(self.lot.fx_source, 'mid-market teste')
        self.assertIsNotNone(self.lot.fx_locked_at)
        self.assertFalse(self.lot.fx_is_fallback)
        # selo na página do lote
        resp = self.client.get(reverse('estoque:lot_detail',
                                       args=[self.lot.pk]))
        self.assertContains(resp, '1 ¥ = US$ 0.1478')

    def test_reabrir_e_so_do_superuser_e_destrava(self):
        self._close()
        # gerente (e admin de empresa) NÃO reabrem mais
        resp = self.client.post(reverse('estoque:lot_reopen',
                                        args=[self.lot.pk]), follow=True)
        self.assertContains(resp, 'exclusiva da plataforma')
        _scope(self, self.company)
        self.lot.refresh_from_db()
        set_current_company(None)
        self.assertEqual(self.lot.status, 'closed')       # seguiu fechado
        # superuser reabre → destrava o câmbio
        User = get_user_model()
        su = User.objects.create_superuser('fxc_root', password='x')
        Membership.objects.create(user=su, company=self.company,
                                  role='manager')
        self.client.logout()
        self.client.login(username='fxc_root', password='x')
        self.client.post(reverse('estoque:lot_reopen', args=[self.lot.pk]))
        _scope(self, self.company)
        self.lot.refresh_from_db()
        set_current_company(None)
        self.assertEqual(self.lot.status, 'open')
        self.assertIsNone(self.lot.fx_rate)
        self.assertIsNone(self.lot.fx_locked_at)

    def test_modal_mostra_a_taxa_que_sera_travada(self):
        resp = self.client.get(reverse('estoque:lot_detail',
                                       args=[self.lot.pk]))
        self.assertContains(resp, 'Câmbio que será TRAVADO')
        self.assertContains(resp, '0.1478')

    def test_ov_herda_a_taxa_travada_do_lote(self):
        from datetime import date
        from pricing.models import Buyer, FxRate, Price, PriceList
        from vendas.models import SalesOrder
        _scope(self, self.company)
        try:
            buyer = Buyer.objects.create(company=self.company, name='Wu C',
                                         slug='wu-fxc')
            pl = PriceList.all_companies.create(buyer=buyer, brand=None)
            Price.all_companies.create(
                price_list=pl, kind='emmc', gen='', origin='phone',
                tier_value=Decimal('16'), tier_unit='GB', status='quoted',
                price_min=Decimal('10'), price_max=Decimal('10'),
                quote_date=date.today())
            InventoryEntry.objects.create(
                lot=self.lot, part_number='FXCEMMC16', quantity=2,
                chip_type='eMMC', price_kind='emmc', price_gen='',
                price_tier_value=Decimal('16'), price_tier_unit='GB')
        finally:
            set_current_company(None)
        self._close()                                    # trava 0.1478
        # a taxa de MERCADO muda DEPOIS do fechamento (dia seguinte)…
        from datetime import timedelta
        FxRate.objects.create(date=date.today() + timedelta(days=1),
                              rate=Decimal('0.1600'), source='t2')
        _scope(self, self.company)
        try:
            # RE-ESPECIFICADO (F11.6/F1, 2026-08-18): o `confirm()` manual saiu
            # — a OV já NASCE congelada no fechamento. O que este teste prova
            # continua o mesmo: a taxa que congela é a TRAVADA no lote.
            so = SalesOrder.all_companies.get(lot=self.lot)
            self.assertEqual(so.status, 'confirmed')
            # …e a OV usa a TRAVADA do lote (não a vigente 0.16):
            self.assertEqual(so.fx_usd_rate, Decimal('0.1478'))
            # ¥10 × 2 un = ¥20 → US$ 2.96 pela taxa travada
            self.assertEqual(so.total_usd, Decimal('2.96'))
        finally:
            set_current_company(None)


class FxHeaderBadgeTests(TestCase):
    """Dono (2026-08-01): "deixar mais claro no frontend o câmbio em tempo
    real" — o header do painel interno estampa a taxa vigente + carimbo."""

    def test_header_mostra_taxa_do_dia(self):
        from datetime import date
        from decimal import Decimal
        from pricing.models import FxRate
        User = get_user_model()
        u = User.objects.create_user(username='fxh_op', password='x')
        _grant(u)
        self.client.login(username='fxh_op', password='x')
        # sem taxa: aviso discreto (bootstrap sem carimbo)
        resp = self.client.get(reverse('estoque:index'))
        self.assertContains(resp, 'rode fetch_fx_rate')
        # com taxa: valor + carimbo mid-market
        FxRate.objects.create(date=date.today(), rate=Decimal('0.1478'),
                              source='mid-market teste')
        resp = self.client.get(reverse('estoque:index'))
        self.assertContains(resp, '1 ¥ ≈ US$ 0.1478')
        self.assertContains(resp, 'mid-market')


class AutoRefreshTests(TestCase):
    """Auto-refresh sem F5 (dono, 2026-08-01): o header faz polling do
    parcial da taxa; a valoração do lote refaz em est:added/60s — valores
    continuam atrás do gate de admin."""

    def setUp(self):
        from datetime import date
        from decimal import Decimal
        from pricing.models import FxRate
        User = get_user_model()
        self.op = User.objects.create_user(username='ar_op', password='x')
        self.company = _grant(self.op)
        self.adm = User.objects.create_user(username='ar_adm', password='x')
        Membership.objects.update_or_create(
            user=self.adm, company=self.company,
            defaults={'role': Membership.ROLE_ADMIN, 'active': True})
        FxRate.objects.create(date=date.today(), rate=Decimal('0.1478'),
                              source='mid-market teste')
        _scope(self, self.company)
        from pricing.models import Buyer
        Buyer.objects.create(company=self.company, name='Wu AR',
                             slug='wu-ar')
        self.lot = Lot.objects.create(number=920, origin='phone',
                                      operator=self.op, company=self.company)
        set_current_company(None)

    def test_fx_badge_para_qualquer_papel(self):
        self.client.login(username='ar_op', password='x')
        resp = self.client.get(reverse('estoque:fx_badge'))
        self.assertContains(resp, '1 ¥ ≈ US$ 0.1478')
        self.assertContains(resp, 'mid-market')

    def test_valoracao_viva_so_admin_ve_valores(self):
        url = reverse('estoque:lot_valuation', args=[self.lot.pk])
        # operador: parcial VAZIO (gate — nunca vaza valor)
        self.client.login(username='ar_op', password='x')
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 200)
        self.assertNotContains(resp, '💰')
        # admin: valoração ¥ ≈ US$
        self.client.logout()
        self.client.login(username='ar_adm', password='x')
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, '💰')

    def test_pagina_do_lote_carrega_os_gatilhos(self):
        self.client.login(username='ar_adm', password='x')
        resp = self.client.get(reverse('estoque:lot_detail',
                                       args=[self.lot.pk]))
        self.assertContains(resp, 'est:added from:body')
        self.assertContains(resp, 'every 60s')

    def test_fmt_card_devolve_valor_compacto(self):
        self.client.login(username='ar_adm', password='x')
        url = reverse('estoque:lot_valuation', args=[self.lot.pk])
        resp = self.client.get(url, {'fmt': 'card'})
        self.assertEqual(resp.status_code, 200)
        self.assertNotContains(resp, '💰')            # formato compacto
        self.assertContains(resp, '¥')

    def test_listagem_tem_polling_e_radios_de_origem(self):
        from django.contrib.auth import get_user_model as _gum
        mgr = _gum().objects.create_user(username='ar_mgr', password='x')
        Membership.objects.update_or_create(
            user=mgr, company=self.company,
            defaults={'role': Membership.ROLE_MANAGER, 'active': True})
        self.client.login(username='ar_mgr', password='x')
        resp = self.client.get(reverse('estoque:index'))
        # radios da origem (cartões, obrigatórios)
        self.assertContains(resp, 'name="origin"')
        self.assertContains(resp, 'value="phone"')
        self.assertContains(resp, 'value="pcb"')
        # o admin vê o valor com polling nos lotes ABERTOS
        self.client.logout()
        self.client.login(username='ar_adm', password='x')
        resp = self.client.get(reverse('estoque:index'))
        self.assertContains(resp, 'fmt=card')
        self.assertContains(resp, 'every 60s')


class K9BenchTests(TestCase):
    """K9 na bancada (dono 2026-08-14, HANDOFF_K9): o operador digita o
    pseudo-código "K9" (exceção ao mínimo de 4 chars — server aqui, client no
    estoque.html) e lança quantidade SEM marca/capacidade. Sem mock: o
    classify real curto-circuita no pseudo-código. Card confirmado (manual) +
    RENTÁVEL; caixa dedicada 'K9' (código K-01); chave (k9, '', 1, '')."""

    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user(username='op_k9', password='x')
        self.company = _grant(self.user)
        _scope(self, self.company)
        self.lot = Lot.objects.create(number=0, origin='phone',
                                      operator=self.user, company=self.company)
        self.client.login(username='op_k9', password='x')

    def test_preview_aceita_k9(self):
        r = self.client.get(reverse('estoque:preview', args=[self.lot.pk]),
                            {'pn': 'k9'})
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, 'K9')

    def test_preview_segue_bloqueando_curtos(self):
        # A exceção é SÓ para o pseudo-código — 3 chars comuns seguem mudos.
        r = self.client.get(reverse('estoque:preview', args=[self.lot.pk]),
                            {'pn': 'K4B'})
        self.assertEqual(r.content, b'')

    def test_add_chip_k9_entra_com_chave_materializada(self):
        from decimal import Decimal
        r = self.client.post(reverse('estoque:add', args=[self.lot.pk]),
                             {'pn': 'K9', 'qty': '50', 'has_cap': 'true',
                              'submit_token': uuid4().hex})
        self.assertEqual(r.status_code, 200)
        e = InventoryEntry.objects.get(lot=self.lot, part_number='K9')
        self.assertEqual(e.quantity, 50)
        self.assertEqual(e.chip_type, 'K9')
        self.assertEqual((e.brand, e.capacity), ('', ''))   # plano de propósito
        self.assertEqual(e.price_kind, 'k9')
        self.assertEqual(e.price_gen, '')
        self.assertEqual(e.price_tier_value, Decimal('1'))
        self.assertEqual(e.price_tier_unit, '')
        self.assertEqual(e.price_key_reason, '')

    def test_add_chip_curto_segue_invalido(self):
        r = self.client.post(reverse('estoque:add', args=[self.lot.pk]),
                             {'pn': 'K4B', 'qty': '1',
                              'submit_token': uuid4().hex})
        self.assertContains(r, 'PN inválido')
        self.assertFalse(InventoryEntry.objects.filter(lot=self.lot).exists())

    def test_caixa_k9(self):
        from estoque.views import _compute_destination
        self.assertEqual(_compute_destination({'chip_type': 'K9'}),
                         ('K9', 'k9'))

    def test_categoria_mascarada_k01(self):
        # Empresa-cliente vê o código da caixa: K-01 (cunhada na aprovação —
        # mesma chave plana da convenção fundadora; nunca H-00).
        from estoque.views import _masked_category
        self.assertEqual(_masked_category({'chip_type': 'K9'}),
                         ('K-01', False))


class QtyCeilingTests(TestCase):
    """Teto de quantidade por lançamento (dono 2026-08-17).

    O input do card nascia com ``max="9999"`` e o navegador barrava lote grande
    de verdade — o caso concreto foi 22.000 chips de um PN só ("Value must be
    less than or equal to 9999"). O teto agora é ``MAX_QTY_POR_LANCAMENTO``
    (1 milhão), publicado no HTML **e** aplicado no servidor: quem ignora o
    formulário (curl, aba antiga em cache, extensão) não fura o limite, e qty
    não-numérico vira 1 em vez de derrubar a request em 500."""

    RENTAVEL = dict(chip_type='eMMC', capacity='16GB',
                    classification_source='banco de dados',
                    confidence='confirmed')

    def setUp(self):
        # superuser p/ o card COMPLETO no preview (mesma razão de
        # PreviewFuzzyPopupTests); o vínculo de operador é o que libera o add.
        self.user = get_user_model().objects.create_user(
            username='op_qty', password='x', is_superuser=True)
        self.company = _grant(self.user)
        _scope(self, self.company)
        self.client.force_login(self.user)
        self.lot = Lot.objects.create(number=0, origin='phone',
                                      operator=self.user, company=self.company)
        self.url = reverse('estoque:add', args=[self.lot.pk])

    def _post(self, pn, qty):
        return self.client.post(self.url, {'pn': pn, 'qty': str(qty),
                                           'has_cap': 'true',
                                           'submit_token': uuid4().hex})

    @patch('estoque.views.classify')
    def test_bancada_lanca_22_mil_de_uma_vez(self, mock_classify):
        # O caso que motivou o fix: o lançamento tem que entrar INTEIRO.
        mock_classify.return_value = _result(**self.RENTAVEL)
        r = self._post('TESTQTY22K', 22000)
        self.assertEqual(r.status_code, 200)
        e = InventoryEntry.objects.get(lot=self.lot, part_number='TESTQTY22K')
        self.assertEqual(e.quantity, 22000)

    @patch('estoque.views.classify')
    def test_servidor_prende_acima_do_teto(self, mock_classify):
        from .views import MAX_QTY_POR_LANCAMENTO
        mock_classify.return_value = _result(**self.RENTAVEL)
        self._post('TESTQTYMAX', MAX_QTY_POR_LANCAMENTO * 9)
        e = InventoryEntry.objects.get(lot=self.lot, part_number='TESTQTYMAX')
        self.assertEqual(e.quantity, MAX_QTY_POR_LANCAMENTO)

    @patch('estoque.views.classify')
    def test_qty_lixo_nao_derruba_a_bancada(self, mock_classify):
        # Antes: int('vinte mil') → ValueError → 500 na cara da operadora.
        mock_classify.return_value = _result(**self.RENTAVEL)
        r = self._post('TESTQTYLIXO', 'vinte mil')
        self.assertEqual(r.status_code, 200)
        e = InventoryEntry.objects.get(lot=self.lot, part_number='TESTQTYLIXO')
        self.assertEqual(e.quantity, 1)

    @patch('estoque.views.classify')
    def test_card_publica_o_teto_no_html(self, mock_classify):
        from .views import MAX_QTY_POR_LANCAMENTO
        mock_classify.return_value = _result(**self.RENTAVEL)
        body = self.client.get(reverse('estoque:preview', args=[self.lot.pk]),
                               {'pn': 'TESTQTYCARD'}).content.decode()
        self.assertIn('id="confirm-qty"', body)
        self.assertIn(f'max="{MAX_QTY_POR_LANCAMENTO}"', body)

    def test_os_dois_cards_seguem_a_constante(self):
        """PORTÃO: HTML e servidor não podem divergir. Mexeu no teto de um
        lado, mexa no outro — inclusive no card MASCARADO (cliente), que não
        aparece no teste de render acima."""
        import pathlib
        from django.conf import settings
        from .views import MAX_QTY_POR_LANCAMENTO
        base = pathlib.Path(settings.BASE_DIR) / 'estoque' / 'templates' / 'estoque' / 'partials'
        for nome in ('confirm_card.html', 'confirm_card_masked.html'):
            t = (base / nome).read_text(encoding='utf-8')
            self.assertIn('id="confirm-qty"', t, nome)
            self.assertIn(f'max="{MAX_QTY_POR_LANCAMENTO}"', t,
                          f'{nome}: teto do input divergiu de MAX_QTY_POR_LANCAMENTO')

    @patch('estoque.views.classify')
    def test_remocao_com_qty_lixo_nao_quebra(self, mock_classify):
        # Mesmo parser na baixa: lixo = 1 unidade, não 500.
        mock_classify.return_value = _result(**self.RENTAVEL)
        self._post('TESTQTYDEL', 5)
        e = InventoryEntry.objects.get(lot=self.lot, part_number='TESTQTYDEL')
        r = self.client.post(reverse('estoque:remove', args=[self.lot.pk, e.pk]),
                             {'qty': ''})
        self.assertEqual(r.status_code, 200)
        e.refresh_from_db()
        self.assertEqual(e.quantity, 4)


class TemplateCsrfTokenTests(TestCase):
    """PORTÃO (2026-08-18): todo `<form method="post">` de template NOSSO tem
    que carregar `{% csrf_token %}`. Sem ele o clique morre em 403 "CSRF token
    missing" — e o sintoma é indistinguível de problema de cookie/sessão, o que
    manda o depurador para o lado errado (foi o que aconteceu: os botões Fechar
    e Reabrir da LISTA de lotes nunca funcionaram, e a caçada foi parar em
    DEBUG/SESSION_COOKIE_DOMAIN).

    O portão é barato: 22 formulários no projeto inteiro. Se um dia existir um
    POST legítimo para fora (outro domínio), a exceção entra aqui explicitada."""

    def test_todo_form_post_tem_csrf_token(self):
        import pathlib
        import re
        from django.conf import settings
        base = pathlib.Path(settings.BASE_DIR)
        raizes = [base / app for app in
                  ('chips', 'estoque', 'pages', 'tenancy', 'pricing',
                   'vendas')] + [base / 'templates']
        maus = []
        for raiz in raizes:
            arquivos = list(raiz.glob('**/templates/**/*.html'))
            if raiz.name == 'templates':
                arquivos += list(raiz.glob('**/*.html'))
            for path in arquivos:
                if 'venv' in path.parts or '_to_delete' in path.parts:
                    continue
                t = path.read_text(encoding='utf-8')
                for m in re.finditer(r'<form\b[^>]*>', t, re.I):
                    if not re.search(r'method\s*=\s*["\']post["\']',
                                     m.group(0), re.I):
                        continue
                    fim = t.find('</form>', m.end())
                    corpo = t[m.end(): fim if fim != -1 else len(t)]
                    if '{% csrf_token %}' not in corpo:
                        maus.append(f'{path.relative_to(base)}:'
                                    f'{t[:m.start()].count(chr(10)) + 1}')
        self.assertEqual(
            maus, [],
            'form method="post" SEM {% csrf_token %} (dá 403 no clique): '
            + ', '.join(maus))
