# -*- coding: utf-8 -*-
"""
A COMPRA EM PLANILHA — as abas Resumo e Chips, uma em cada aba do arquivo.

Substitui o CSV por aba (dono, 2026-09-02: "o botao de exportar esta
exportando um CSV com os chips, isso nao serve ao comprador"). Dois motivos
para o CSV não servir:

  1. Ele entregava UMA aba — a que estivesse aberta. Quem exporta uma compra
     quer a compra, não o recorte em que o cursor parou.
  2. CSV não tem abas, então "Resumo e Chips" viravam dois downloads que
     ninguém junta depois.

⚠ O CSV também não era fiel à tela: a aba Chips mostra `Tipo` e o CSV não
  exportava essa coluna. Aqui as colunas são as MESMAS da tela, na mesma
  ordem — é o que o dono pediu ("assim EXATAMENTE como elas sao lá no
  sistema"), e é o que torna a planilha conferível contra o que ele viu.

A fonte é o `_detalhe(so)`, o MESMO dicionário que renderiza a ficha. Não é
economia de código: recalcular aqui criaria uma segunda conta para o mesmo
fato, e a primeira vez que as duas divergissem o comprador teria uma planilha
que discorda da tela sem nada que diga qual está certa.

Números saem como NÚMERO, com formato `¥`. Texto formatado ("¥ 1.234,00")
parece igual e não soma — e planilha que não soma é print.
"""

import io

from django.utils.translation import gettext as _, ngettext

# Tokens do design system, para a planilha não parecer de outro produto:
# 0F62FE é o --blue-60 (mesma faixa de cabeçalho do export do estoque),
# E0E0E0 é o --line e F2F4F8 é o --ink-10, a tinta da faixa de grupo.
AZUL, LINHA, CINZA = '0F62FE', 'E0E0E0', 'F2F4F8'

FMT_QTD = '#,##0'
FMT_RMB = '"¥" #,##0.00'
#: rascunho: o mesmo número, com o sinal de estimativa que a tela mostra.
FMT_RMB_EST = '"≈ ¥" #,##0.00'

XLSX = ('application/vnd.openxmlformats-officedocument.'
        'spreadsheetml.sheet')


def _estilos():
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    return {
        'h_fill': PatternFill('solid', fgColor=AZUL),
        'h_font': Font(name='Calibri', bold=True, color='FFFFFF', size=11),
        'h_alin': Alignment(horizontal='left', vertical='center'),
        'borda': Border(bottom=Side(style='thin', color=LINHA)),
        'mono': Font(name='Courier New', size=10),
        'g_fill': PatternFill('solid', fgColor=CINZA),
        'g_font': Font(name='Calibri', bold=True, size=10),
        't_font': Font(name='Calibri', bold=True, size=11),
        't_borda': Border(top=Side(style='medium', color='161616')),
    }


def _titulo(ws, so, colunas, e):
    """Identificação em UMA linha acima da tabela.

    A tabela abaixo é a da tela, coluna por coluna. Esta linha não é parte
    dela — existe porque o arquivo sai da tela e vira anexo de e-mail: sem o
    código da ordem, dois downloads na mesma pasta são indistinguíveis, e foi
    exatamente essa colisão que fez o código do lote ganhar ano e a lista
    trocar a coluna do lote pela da ordem.
    """
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    ws.cell(row=1, column=1, value='%s · %s · %s' % (
        so.code, so.buyer.name if so.buyer_id else '—',
        so.lot.code if so.lot_id else '—')).font = Font(
            name='Calibri', bold=True, size=12)
    ws.freeze_panes = 'A4'
    for i, (rotulo, largura) in enumerate(colunas, start=1):
        c = ws.cell(row=3, column=i, value=rotulo)
        c.font, c.fill, c.alignment = e['h_font'], e['h_fill'], e['h_alin']
        ws.column_dimensions[get_column_letter(i)].width = largura
    ws.row_dimensions[3].height = 26


def _pinta(ws, linha, dados, e, fonte=None, fill=None, borda=None):
    from openpyxl.styles import Alignment
    for i, valor in enumerate(dados, start=1):
        c = ws.cell(row=linha, column=i, value=valor)
        c.border = borda if borda is not None else e['borda']
        c.alignment = Alignment(vertical='center')
        if fonte is not None:
            c.font = fonte
        if fill is not None:
            c.fill = fill
    return ws[linha]


def _num(ws, linha, coluna, formato):
    ws.cell(row=linha, column=coluna).number_format = formato


def _dinheiro(valor, legado):
    """O que a CÉLULA recebe quando não há preço.

    A tela escreve "sem preço" no unitário e "—" no total, e a planilha diz o
    mesmo: um vazio aqui seria lido como zero, que é a única leitura errada
    possível — zero é um preço, ausência de preço não é.
    """
    if valor is not None:
        return float(valor)
    return '—' if legado else _('sem preço')


# ── ABA RESUMO ───────────────────────────────────────────────────────────
def _aba_resumo(ws, so, ctx):
    e = _estilos()
    ws.title = _('Resumo')
    acerto = ctx['pode_acertar'] or ctx['tem_resultado']
    legado = ctx['registro_legado']

    colunas = [(_('Marca'), 16), (_('Tipo'), 14), (_('Capacidade'), 18),
               (_('Caixa WTC'), 14), (_('Enviados'), 11),
               (_('¥ unit.'), 13), (_('¥ esperado'), 15)]
    if acerto:
        colunas += [(_('Recusados'), 12), (_('Aprovados'), 12),
                    (_('¥ resultado'), 15)]
    _titulo(ws, so, colunas, e)

    r = 4
    for g in ctx['grupos']:
        # A FAIXA DA MARCA, como na tela: a marca e o subtotal dela. A marca
        # se repete na coluna 1 de cada linha abaixo — na tela quem diz a
        # marca é a faixa, mas numa planilha uma coluna vazia impede filtrar e
        # ordenar, que é metade do motivo de exportar.
        faixa = [g['brand'], '', '', '', g['qty'],
                 '', '—' if legado else float(g['rmb'])]
        if acerto:
            faixa += [-g['rejected'] if g['rejected'] else 0, g['accepted'],
                      '—' if legado else float(g['pago_rmb'])]
        _pinta(ws, r, faixa, e, fonte=e['g_font'], fill=e['g_fill'])
        _num(ws, r, 5, FMT_QTD)
        if not legado:
            _num(ws, r, 7, FMT_RMB)
            if acerto:
                _num(ws, r, 10, FMT_RMB)
        r += 1

        for l in g['lines']:
            dados = [g['brand'], l['type'], l['capacity'], l['wtc'], l['qty'],
                     _dinheiro(l['unit_rmb'], legado),
                     float(l['total_rmb']) if l['total_rmb'] is not None else '—']
            if acerto:
                dados += [
                    -l['rejected'] if l['rejected'] else 0, l['accepted'],
                    float(l['pago_rmb']) if l['pago_rmb'] is not None else '—']
            _pinta(ws, r, dados, e)
            _num(ws, r, 5, FMT_QTD)
            if l['unit_rmb'] is not None:
                _num(ws, r, 6, FMT_RMB)
            if l['total_rmb'] is not None:
                _num(ws, r, 7, FMT_RMB)
            if acerto and l['pago_rmb'] is not None:
                _num(ws, r, 10, FMT_RMB)
            r += 1

    # RODAPÉ — o mesmo da tela: "Total · N marcas", enviados, esperado e,
    # quando há acerto, recusados/aprovados/resultado. O esperado sai do
    # CONGELADO da ordem quando existe; no rascunho é a soma viva, com "≈".
    # ngettext com as MESMAS strings do `blocktrans count` da tela: reusa a
    # entrada que já existe no catálogo, em vez de criar um msgid paralelo que
    # alguém teria de traduzir de novo — e que perderia o plural.
    n_marcas = len(ctx['grupos'])
    total = [ngettext('Total · %(n)s marca', 'Total · %(n)s marcas',
                      n_marcas) % {'n': n_marcas},
             '', '', '', ctx['total_qty'], '']
    if ctx['estimado']:
        total.append(float(ctx['total_estimado']))
    elif so.total_rmb:
        total.append(float(so.total_rmb))
    else:
        total.append('—')
    if acerto:
        rej = sum(g['rejected'] for g in ctx['grupos'])
        total += [-rej if rej else 0,
                  sum(g['accepted'] for g in ctx['grupos']),
                  float(sum(g['pago_rmb'] for g in ctx['grupos']))]
    _pinta(ws, r, total, e, fonte=e['t_font'], borda=e['t_borda'])
    _num(ws, r, 5, FMT_QTD)
    # O "≈" do rascunho entra pelo FORMATO, não no texto: a tela escreve
    # "≈ ¥ 3.750,00" e o sinal não é enfeite — diz que o número é estimativa
    # contra a tabela de hoje, não o congelado da ordem. Escrever "≈ ¥ ..."
    # como texto mostraria a mesma coisa e mataria a célula: quem exporta
    # planilha soma a coluna, e texto não soma.
    _num(ws, r, 7, FMT_RMB_EST if ctx['estimado'] else FMT_RMB)
    if acerto:
        _num(ws, r, 10, FMT_RMB)


# ── ABA CHIPS ────────────────────────────────────────────────────────────
def _aba_chips(ws, so, ctx):
    e = _estilos()
    ws.title = _('Chips')
    legado = ctx['registro_legado']
    chips = ctx['chips']

    colunas = [(_('Part Number'), 24), (_('Marca'), 16), (_('Tipo'), 14),
               (_('Spec'), 20), (_('Caixa WTC'), 14), (_('Qtd.'), 11),
               (_('¥ unit.'), 13), (_('¥ total'), 15)]
    _titulo(ws, so, colunas, e)

    r = 4
    for c in chips['linhas']:
        _pinta(ws, r, [c['pn'], c['brand'], c['type'], c['spec'], c['wtc'],
                       c['qty'], _dinheiro(c['unit_rmb'], legado),
                       float(c['total_rmb']) if c['total_rmb'] is not None
                       else '—'], e)
        ws.cell(row=r, column=1).font = e['mono']     # PN é código, não texto
        _num(ws, r, 6, FMT_QTD)
        if c['unit_rmb'] is not None:
            _num(ws, r, 7, FMT_RMB)
        if c['total_rmb'] is not None:
            _num(ws, r, 8, FMT_RMB)
        r += 1

    _pinta(ws, r, [_('Total'), '', '', '', '', chips['qty'], '',
                   '—' if legado else float(chips['rmb'])],
           e, fonte=e['t_font'], borda=e['t_borda'])
    _num(ws, r, 6, FMT_QTD)
    if not legado:
        _num(ws, r, 8, FMT_RMB)


# ── A ENTRADA ────────────────────────────────────────────────────────────
def compra_em_planilha(so, ctx):
    """Devolve ``(bytes, nome_do_arquivo)``.

    `ctx` é o `_detalhe(so)` da ficha — quem chama passa o MESMO dicionário
    que a tela usou, e não um recalculado.
    """
    from openpyxl import Workbook
    wb = Workbook()
    _aba_resumo(wb.active, so, ctx)
    _aba_chips(wb.create_sheet(), so, ctx)
    wb.properties.creator = 'WhatTheChip?'
    wb.properties.title = so.code

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    # O nome leva o código da ORDEM, e não o do lote como fazia o CSV. O
    # código do lote perdeu o prefixo da empresa em 2026-09-02: dois clientes
    # com o lote 7 dariam o MESMO nome de arquivo na pasta de Downloads. É a
    # mesma colisão que tirou a coluna do lote da lista de compras.
    return buf.read(), '%s.xlsx' % so.code.replace('/', '-')
