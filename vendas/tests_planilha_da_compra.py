# -*- coding: utf-8 -*-
"""
O EXPORT DA FICHA: planilha com Resumo e Chips (dono, 2026-09-02).

  "o botao de exportar esta exportando um CSV com os chips, isso nao serve ao
   comprador, serve mais que seja exportado a aba de RESUMO e CHIPS inteira,
   1 em cada aba de planilha, assim EXATAMENTE como elas sao lá no sistema."

Três defeitos no que existia, e cada um vira teste aqui:

  1. Exportava UMA aba — a que estivesse aberta. Quem exporta uma compra quer
     a compra.
  2. CSV não tem abas, então Resumo e Chips seriam dois downloads.
  3. O CSV NÃO era fiel à tela: a aba Chips mostra `Tipo` e o CSV não
     exportava essa coluna. "Exatamente como na tela" não é figura de
     linguagem — é o que torna a planilha conferível contra o que ele viu.

⚠ Há um quarto, mais sutil, e ele tem teste próprio: o CSV decidia mostrar as
  colunas de recusa por `so.received_at`, e a TELA decide por
  `pode_acertar or tem_resultado`. São condições diferentes. Duas regras para
  a mesma pergunta divergem — é o mesmo defeito do selo que saiu do topo da
  ficha, em outro lugar.
"""

import io
from decimal import Decimal as D
from datetime import date

from django.conf import settings
from django.contrib.auth import get_user_model
from django.test import Client, TestCase
from django.urls import reverse
from django.utils import timezone

from estoque.models import Lot
from pricing.models import Buyer
from tenancy.models import Company, Membership
from tenancy.scope import company_scope
from vendas.models import (DocSequence, SEQ_SO, SalesOrder, SalesOrderLine,
                           STATUS_CONFIRMED)

User = get_user_model()


class _Base(TestCase):

    def setUp(self):
        self.emp = Company.objects.create(name='eMiner', slug='eminer', code='')
        self.buyer = Buyer.all_companies.create(company=None, name='Wu Quan',
                                                slug='wu-quan')
        self.parceiro = User.objects.create_user('u_wq', password='x')
        self.buyer.users.add(self.parceiro)
        self.gerente = User.objects.create_user('g', password='x')
        Membership.objects.create(user=self.gerente, company=self.emp,
                                  role=Membership.ROLE_MANAGER)
        with company_scope(self.emp.id):
            self.lot = Lot.all_companies.create(
                company=self.emp, number=7, description='x', status='closed',
                operator=self.gerente, origin='pcb')
            self.so = SalesOrder(
                lot=self.lot, buyer=self.buyer, status=STATUS_CONFIRMED,
                fx_usd_rate=D('0.1400'), total_rmb=D('3000.00'),
                total_usd=D('420.00'), shipped_at=date(2026, 8, 27),
                number=DocSequence.next_number(self.emp, SEQ_SO))
            self.so.save()
            for marca, cap, qtd in (('Samsung', D('64'), 200),
                                    ('Hynix', D('4'), 50)):
                SalesOrderLine.all_companies.create(
                    order=self.so, company=self.emp, brand=marca,
                    kind='emmc', gen='', tier_value=cap, tier_unit='GB',
                    quantity=qtd, unit_rmb=D('10.00'))
        self.client.force_login(self.parceiro)

    def _baixar(self):
        r = self.client.get(reverse('compras:planilha', args=[self.so.pk]))
        self.assertEqual(r.status_code, 200)
        return r

    def _wb(self, r=None):
        import openpyxl
        return openpyxl.load_workbook(io.BytesIO((r or self._baixar()).content))

    def _linha(self, ws, n):
        return [c.value for c in ws[n]]


class ArquivoTests(_Base):

    def test_sai_um_xlsx_de_verdade(self):
        r = self._baixar()
        self.assertEqual(
            r['Content-Type'],
            'application/vnd.openxmlformats-officedocument.'
            'spreadsheetml.sheet')
        self.assertIn('attachment;', r['Content-Disposition'])
        self.assertTrue(r.content[:2] == b'PK', 'não é um zip/xlsx')

    def test_duas_abas_resumo_e_chips(self):
        """O pedido literal: "1 em cada aba de planilha"."""
        self.assertEqual(self._wb().sheetnames, ['Resumo', 'Chips'])

    def test_o_nome_do_arquivo_leva_o_codigo_da_ORDEM(self):
        """E não o do lote, como fazia o CSV. O código do lote perdeu o
        prefixo da empresa em 2026-09-02: o lote 7 de dois clientes daria o
        MESMO nome de arquivo na pasta de Downloads. É a mesma colisão que
        tirou a coluna do lote da lista de compras."""
        nome = self._baixar()['Content-Disposition']
        self.assertIn(self.so.code, nome)
        self.assertNotIn(self.lot.code, nome)

    def test_a_planilha_diz_de_que_compra_e(self):
        """Ela vira anexo de e-mail. Sem identificação, dois arquivos na mesma
        pasta são indistinguíveis — e o cabeçalho da tabela não diz nada sobre
        a compra."""
        cab = self._wb()['Resumo'].cell(row=1, column=1).value
        self.assertIn(self.so.code, cab)
        self.assertIn('Wu Quan', cab)
        self.assertIn(self.lot.code, cab)


class ColunasIguaisAsDaTelaTests(_Base):

    #: exatamente os `<th>` da tabela do Resumo, na ordem, mais a Marca —
    #: que na tela é a faixa de grupo e aqui vira coluna, para dar filtro.
    RESUMO = ['Marca', 'Tipo', 'Capacidade', 'Caixa WTC', 'Enviados',
              '¥ unit.', '¥ esperado']
    #: exatamente os `<th>` da tabela de Chips, na ordem.
    CHIPS = ['Part Number', 'Marca', 'Tipo', 'Spec', 'Caixa WTC', 'Qtd.',
             '¥ unit.', '¥ total']

    def test_as_colunas_do_resumo_sao_as_da_tela(self):
        cab = [c for c in self._linha(self._wb()['Resumo'], 3) if c]
        self.assertEqual(cab[:len(self.RESUMO)], self.RESUMO)

    def test_as_colunas_de_chips_sao_as_da_tela(self):
        self.assertEqual([c for c in self._linha(self._wb()['Chips'], 3) if c],
                         self.CHIPS)

    def test_chips_traz_a_coluna_TIPO_que_o_csv_esquecia(self):
        """O defeito nº 3. A tela sempre mostrou `Tipo` na aba Chips; o CSV
        exportava sem ela, e ninguém tinha como notar sem comparar as duas
        lado a lado."""
        self.assertIn('Tipo', self._linha(self._wb()['Chips'], 3))


class NumeroEhNumeroTests(_Base):

    def test_dinheiro_sai_como_numero_com_formato(self):
        """O motivo de ser planilha e não print. Texto "¥ 2.000,00" parece
        igual na tela e não soma — e quem exporta soma a coluna."""
        ws = self._wb()['Resumo']
        achou = False
        for linha in range(4, ws.max_row + 1):
            c = ws.cell(row=linha, column=7)           # ¥ esperado
            if isinstance(c.value, (int, float)):
                self.assertIn('¥', c.number_format)
                achou = True
        self.assertTrue(achou, 'nenhuma célula de dinheiro virou número')

    def test_quantidade_sai_como_inteiro(self):
        ws = self._wb()['Resumo']
        self.assertIsInstance(ws.cell(row=4, column=5).value, int)

    def test_o_total_bate_com_o_congelado_da_ordem(self):
        ws = self._wb()['Resumo']
        self.assertEqual(ws.cell(row=ws.max_row, column=7).value,
                         float(self.so.total_rmb))


class ColunasDeAcertoTests(_Base):
    """As colunas Recusados / Aprovados / ¥ resultado.

    ⚠ CORREÇÃO DE UMA AFIRMAÇÃO ERRADA (2026-09-02). Escrevi antes que "uma
    ordem confirmada e sem fatura já mostra as colunas mesmo antes do
    recebimento". É falso, e o teste que nasceu dessa frase quebrou na
    primeira execução — bem quebrado. O `pode_acertar` EXIGE o recebimento
    ("ele deve acusar como recebido primeiro", dono 2026-08-18): não se
    confere caixa que ainda não chegou.

    A divergência entre a tela e o CSV antigo existe, mas é o caso OPOSTO:

        tela .. (confirmada E sem fatura E recebida)  OU  (tem fatura)
        csv .... recebida

    Elas só discordam quando há FATURA e o recebimento nunca foi registrado —
    o estado `pulado` do trilho, em que o resultado fechou sem ninguém marcar
    a chegada. Aí a tela mostra as colunas (a conferência aconteceu, e é o que
    o comprador precisa reler) e o CSV as escondia. A planilha segue a tela.
    """

    def _cabecalho(self):
        return [c for c in self._linha(self._wb()['Resumo'], 3) if c]

    ACERTO = ['Recusados', 'Aprovados', '¥ resultado']

    def test_sem_recebimento_as_colunas_nao_aparecem(self):
        """Não há o que relatar: nada foi conferido. Coluna vazia num export
        é pergunta sem resposta."""
        self.assertIsNone(self.so.received_at)
        cab = self._cabecalho()
        for coluna in self.ACERTO:
            self.assertNotIn(coluna, cab)

    def test_marcado_o_recebimento_as_colunas_entram(self):
        self.so.received_at = timezone.now()
        self.so.save(update_fields=['received_at'])
        cab = self._cabecalho()
        for coluna in self.ACERTO:
            self.assertIn(coluna, cab)

    def test_a_planilha_usa_a_MESMA_condicao_que_a_tela(self):
        """A trava. Se alguém trocar a condição do exportador por
        `received_at` — que é o que o CSV antigo fazia —, este teste continua
        passando aqui e falha no caso `pulado`; por isso ele compara a
        CONDIÇÃO, e não o efeito: os dois lados têm de ler o mesmo
        `pode_acertar or tem_resultado` do `_detalhe`.

        ⚠ `_detalhe` toca `InventoryEntry`, que é multi-empresa: fora de um
        `company_scope` o próprio ORM levanta `CompanyScopeMissing`. Na view
        quem abre o escopo é o `services.buyer_order`; aqui tem de ser
        explícito. Foi assim que este teste falhou da primeira vez — e a falha
        provou, de graça, que a view está certa em chamar `_detalhe` DENTRO do
        `with`.
        """
        from vendas.views_partner import _detalhe
        for recebido in (False, True):
            self.so.received_at = timezone.now() if recebido else None
            self.so.save(update_fields=['received_at'])
            with company_scope(self.emp.id):
                ctx = _detalhe(self.so)
            esperado = ctx['pode_acertar'] or ctx['tem_resultado']
            self.assertEqual(esperado, recebido,
                             'a condição da tela mudou — reveja a planilha')
            cab = self._cabecalho()
            self.assertEqual('Recusados' in cab, esperado,
                             'a planilha discorda da tela com recebido=%s'
                             % recebido)


class BotaoDaFichaTests(_Base):

    def test_o_botao_aponta_para_a_planilha(self):
        html = self.client.get(
            reverse('compras:detail', args=[self.so.pk])).content.decode()
        self.assertIn(reverse('compras:planilha', args=[self.so.pk]), html)

    def test_o_botao_nao_segue_mais_a_aba(self):
        """Some junto a reescrita do href no JS. Sem destino variável não há o
        que reescrever — e era ela que podia entregar o arquivo da aba
        anterior com o nome da atual."""
        html = self.client.get(
            reverse('compras:detail', args=[self.so.pk])).content.decode()
        self.assertNotIn("exp.href", html)
        self.assertNotIn('id="exp-aba"', html)


class SoDoDonoTests(_Base):

    def test_outro_comprador_nao_baixa(self):
        """A planilha carrega preço e quantidade de um lote inteiro. Quem
        entra pela URL sem ser o dono da ordem não pode receber nada."""
        outro_buyer = Buyer.all_companies.create(company=None, name='Outro',
                                                 slug='outro')
        outro = User.objects.create_user('u_outro', password='x')
        outro_buyer.users.add(outro)
        self.client.force_login(outro)
        r = self.client.get(reverse('compras:planilha', args=[self.so.pk]))
        self.assertIn(r.status_code, (403, 404))


class IdiomaTests(_Base):
    """A planilha herda o idioma ATIVO do usuário (dono, 2026-09-02: "preciso
    que o idioma dela herde do idioma do sistema do usuario no momento, é
    possivel?").

    É — e sai de graça, desde que os rótulos sejam resolvidos na HORA da
    requisição. `gettext` (e não `gettext_lazy` guardado em constante de
    módulo) faz exatamente isso: a constante seria resolvida no import e
    congelaria o idioma do primeiro processo que carregasse o módulo, que é o
    bug clássico e o motivo de o `_stage_labels` ser função e não dicionário.

    Estes testes existem para que uma "otimização" futura — subir a lista de
    colunas para o topo do arquivo, por exemplo — não passe despercebida.
    """

    #: (nome da aba, coluna Marca, coluna Enviados) — os valores SÃO os do
    #: catálogo versionado, conferidos um a um. Cravar aqui é o ponto: se o
    #: `.mo` deixar de ser compilado, o rótulo cai no msgid em português e
    #: este teste é quem avisa.
    ESPERADO = {
        'pt-br':   ('Resumo', 'Marca', 'Enviados'),
        'en':      ('Summary', 'Brand', 'Sent'),
        'es':      ('Resumen', 'Marca', 'Enviados'),
        'zh-hans': ('汇总', '品牌', '发出'),
    }

    def _em(self, idioma):
        c = Client()
        c.force_login(self.parceiro)
        c.cookies[settings.LANGUAGE_COOKIE_NAME] = idioma
        r = c.get(reverse('compras:planilha', args=[self.so.pk]))
        self.assertEqual(r.status_code, 200)
        import openpyxl
        return openpyxl.load_workbook(io.BytesIO(r.content))

    def test_o_nome_das_abas_segue_o_idioma(self):
        for idioma in self.ESPERADO:
            wb = self._em(idioma)
            self.assertEqual(wb.sheetnames[0], self.ESPERADO[idioma][0],
                             'aba Resumo em %s' % idioma)

    def test_os_titulos_das_colunas_seguem_o_idioma(self):
        for idioma, (_aba, marca, enviados) in self.ESPERADO.items():
            cab = [c.value for c in self._em(idioma)[_aba][3]]
            self.assertIn(marca, cab, 'coluna Marca em %s' % idioma)
            self.assertIn(enviados, cab, 'coluna Enviados em %s' % idioma)

    def test_trocar_de_idioma_troca_o_arquivo(self):
        """A prova de que nada ficou preso no import: dois downloads na MESMA
        sessão do processo, em idiomas diferentes, saem diferentes."""
        self.assertNotEqual(self._em('pt-br').sheetnames,
                            self._em('en').sheetnames)

    def test_os_numeros_nao_traduzem(self):
        """Rótulo traduz; DADO não. A quantidade e o dinheiro são os mesmos em
        qualquer idioma — se um dia alguém formatar o número no Python em vez
        de deixar no formato da célula, é aqui que aparece.

        ⚠ Compara a COLUNA INTEIRA entre os idiomas, e não uma célula fixa. A
        primeira versão cravava "linha 4 vale 200" e falhou: o `result_rows`
        devolve os grupos ordenados por marca, então a linha 4 é a faixa do
        Hynix (50), não a do Samsung que o cenário cria primeiro. Cravar
        posição num teste é depender de uma ordenação que o teste não declara
        — e comparar entre idiomas é o que a garantia realmente diz.
        """
        colunas = {}
        for idioma, (aba, _m, _e) in self.ESPERADO.items():
            ws = self._em(idioma)[aba]
            colunas[idioma] = [ws.cell(row=r, column=5).value
                               for r in range(4, ws.max_row + 1)]
        referencia = colunas['pt-br']
        self.assertTrue(all(isinstance(v, int) for v in referencia),
                        'a coluna Enviados deixou de ser número: %r'
                        % referencia)
        # 200 + 50 nas faixas, os mesmos nas linhas, e 250 no rodapé.
        self.assertEqual(referencia[-1], 250, 'o total de enviados mudou')
        self.assertEqual(sorted(referencia), [50, 50, 200, 200, 250])
        for idioma, valores in colunas.items():
            self.assertEqual(valores, referencia,
                             'a coluna Enviados mudou em %s' % idioma)
